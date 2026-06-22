import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
import pickle
import os
import pulp
import numpy as np
import math
import hashlib
from collections import defaultdict
import json
import psycopg2
import threading

# 页面配置必须放在最前面
st.set_page_config(
    page_title="电价数据查询系统",
    page_icon="⚡",
    layout="wide"
)

st.title("⚡ 站点电价数据查询系统")

# 全局表格样式：数据与标题列居中对齐
st.markdown("""
<style>
    .stDataFrame th, .stDataFrame td {
        text-align: center !important;
    }
</style>
""", unsafe_allow_html=True)

# 定义脚本目录和电价数据目录
SCRIPT_DIR = Path(__file__).resolve().parent
PRICE_DATA_DIR = SCRIPT_DIR / "电价数据"
CACHE_DIR = SCRIPT_DIR / ".cache"
CACHE_DIR.mkdir(exist_ok=True)
RANKING_CACHE_FILE = CACHE_DIR / ".price_spread_rank_cache.pkl"
PROVINCIAL_AVG_CACHE_FILE = CACHE_DIR / ".provincial_avg_cache.pkl"
SHARED_OPT_CACHE_FILE = CACHE_DIR / ".shared_storage_opt_cache.pkl"
SHARED_OPT_CACHE_LOCK = threading.Lock()
RANKING_CACHE_VERSION = 2
PROVINCIAL_AVG_CACHE_VERSION = 2
SHARED_OPT_CACHE_VERSION = 1
FACTORY_GROUP_COLUMN = "厂站类型"
FACTORY_STATION_LABEL = "电厂"
NON_FACTORY_STATION_LABEL = "电站"
GUANGDONG_SHARP_PEAK_MONTHS = {7, 8, 9}

# TimescaleDB 数据库连接配置
DB_CONFIG = {
    'dbname': 'postgres',
    'user': 'postgres',
    'password': '123456',
    'host': 'localhost',
    'port': '5432'
}

# 储能优化参数配置
STORAGE_CONFIG = {
    'P': 200,  # 储能逆变器功率，单位 MW
    'battery_capacity': 400,  # 电池容量上限，单位 MWh
    'initial_soc': 0,  # 初始电量，单位 MWh
    'efficiency': 0.85,  # 放电效率
    'dt': 0.25,  # 时间间隔，小时
    'num': 96  # 每天时段数
}

# 区域节点映射
REGION_MAPPING = {
    '珠三角': ['东莞', '佛山', '广州', '惠州', '江门', '深圳', '珠海', '肇庆', '中山'],
    '粤北': ['韶关', '清远', '河源', '梅州'],
    '粤东': ['汕头', '汕尾', '揭阳', '潮州'],
    '粤西': ['湛江', '茂名', '阳江', '云浮'],
}

# 使用单选切换视图，避免隐藏页面也执行耗时计算
view_mode = st.radio(
    "功能切换",
    options=["📊 电价数据查询", "📈 多站点电价对比", "📈 电价差排名", "🔋 储能配储优化", "📊 区域节点电价对比", "💰 光伏上网电费计算"],
    horizontal=True,
    label_visibility="collapsed"
)

# 加载电站名.xlsx文件
@st.cache_data
def load_station_info():
    """加载电站名.xlsx文件，获取站点与母线的对应关系"""
    try:
        # 尝试多个可能的位置
        possible_paths = [
            SCRIPT_DIR / "电站名.xlsx",
            SCRIPT_DIR / "电价名.xlsx",
            PRICE_DATA_DIR / "电站名.xlsx",
            PRICE_DATA_DIR / "电价名.xlsx",
            SCRIPT_DIR.parent / "电站名.xlsx",
            Path.home() / "Desktop" / "电站名.xlsx",
            Path.home() / "Documents" / "电站名.xlsx",
        ]

        for path in possible_paths:
            if path.exists():
                df = pd.read_excel(path)
                return df

        return None
    except Exception:
        return None


# 数据库查询函数
def get_db_connection():
    """获取数据库连接"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        st.error(f"数据库连接失败: {e}")
        return None


@st.cache_data(ttl=300, show_spinner=False)  # 缓存5分钟
def check_db_available():
    """检测数据库是否可用，返回(True, None)或(False, 错误信息)"""
    try:
        conn = psycopg2.connect(**DB_CONFIG, connect_timeout=5)
        conn.close()
        return True, None
    except Exception as e:
        return False, str(e)


@st.cache_data(ttl=300, show_spinner=False)  # 缓存5分钟
def get_db_station_list():
    """从数据库获取所有站点列表"""
    conn = get_db_connection()
    if conn is None:
        return []
    
    try:
        query = """
            SELECT DISTINCT station
            FROM real_time_electricity_price
            ORDER BY station
        """
        df = pd.read_sql_query(query, conn)
        return df['station'].tolist()
    except Exception as e:
        st.error(f"查询站点列表失败: {e}")
        return []
    finally:
        conn.close()


@st.cache_data(ttl=300, show_spinner=False)  # 缓存5分钟
def get_db_available_years():
    """从数据库获取可用的年份列表"""
    conn = get_db_connection()
    if conn is None:
        return []
    
    try:
        query = """
            SELECT DISTINCT EXTRACT(YEAR FROM time)::integer AS year
            FROM real_time_electricity_price
            ORDER BY year DESC
        """
        df = pd.read_sql_query(query, conn)
        return df['year'].tolist()
    except Exception as e:
        st.error(f"查询年份列表失败: {e}")
        return []
    finally:
        conn.close()


@st.cache_data(ttl=600, show_spinner=False)  # 缓存10分钟
def load_price_data_from_db(station_name, year):
    """从数据库加载指定站点和年份的电价数据
    
    Args:
        station_name: 站点名称
        year: 年份 (int)
    
    Returns:
        DataFrame: 包含日期列和96个时段电价列的宽格式数据
    """
    conn = get_db_connection()
    if conn is None:
        return None
    
    try:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
        
        query = """
            SELECT time, price
            FROM real_time_electricity_price
            WHERE station = %s
              AND time >= %s AND time < %s
            ORDER BY time
        """
        df = pd.read_sql_query(query, conn, params=(station_name, start_date, end_date))
        
        if df.empty:
            return None
        
        # 转换时间格式
        df['time'] = pd.to_datetime(df['time'])
        df['date'] = df['time'].dt.date
        df['time_str'] = df['time'].dt.strftime('%H:%M')
        
        # 透视表：将长格式转换为宽格式
        pivot_df = df.pivot_table(
            index='date',
            columns='time_str',
            values='price',
            aggfunc='first'
        )
        
        # 重置索引并重命名列
        pivot_df = pivot_df.reset_index()
        pivot_df.columns.name = None
        
        # 按时间排序列（00:00, 00:15, 00:30, ...）
        time_cols = [col for col in pivot_df.columns if col != 'date']
        time_cols_sorted = sorted(time_cols)
        pivot_df = pivot_df[['date'] + time_cols_sorted]
        
        # 将date列转换为datetime类型
        pivot_df['date'] = pd.to_datetime(pivot_df['date'])
        
        return pivot_df
    except Exception as e:
        st.error(f"从数据库加载数据失败: {e}")
        return None
    finally:
        conn.close()


@st.cache_data(show_spinner="正在计算全省平均电价...")
def compute_provincial_average_from_db(station_year_pairs):
    """从数据库计算全省所有站点各时段平均电价
    
    Args:
        station_year_pairs: [(station_name, year), ...] 站点和年份的元组列表
    
    Returns:
        (date_list, time_columns, avg_by_date, failed_stations)
    """
    # 先检测数据库连接
    db_available, db_error = check_db_available()
    if not db_available:
        st.error(f"数据库连接断开: {db_error}")
        return None, None, None, []
    
    all_daily_arrays = []
    failed_stations = []
    
    for station_name, year in station_year_pairs:
        df = load_price_data_from_db(station_name, year)
        if df is None:
            failed_stations.append(station_name)
            continue
        
        if len(df.columns) < 2:
            failed_stations.append(station_name)
            continue
        
        date_col = df.columns[0]
        price_cols = df.columns[1:]
        
        if len(price_cols) < 96:
            failed_stations.append(station_name)
            continue
        price_cols = price_cols[:96]
        
        for _, row in df.iterrows():
            date_val = row[date_col]
            try:
                date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
            except Exception:
                date_str = str(date_val)[:10]
            prices = pd.to_numeric(row[price_cols], errors='coerce').values.astype(float)
            if len(prices) == 96 and not np.all(np.isnan(prices)):
                all_daily_arrays.append((date_str, prices))
    
    if not all_daily_arrays:
        return None, None, None, failed_stations
    
    # Group by date and compute mean
    date_grouped = {}
    for date_str, prices in all_daily_arrays:
        if date_str not in date_grouped:
            date_grouped[date_str] = []
        date_grouped[date_str].append(prices)
    
    date_list = sorted(date_grouped.keys())
    avg_by_date = np.array([np.mean(date_grouped[d], axis=0) for d in date_list])
    
    time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]
    
    return date_list, time_columns, avg_by_date, failed_stations


def _sliding_window_spread(prices, window_size):
    """用滑动窗口计算峰谷价差：连续window_size个点的均值最大值 - 连续window_size个点的均值最小值

    Args:
        prices: 一维价格数组
        window_size: 窗口大小

    Returns:
        (peak_avg, valley_avg, spread)
    """
    arr = np.asarray(prices, dtype=float)
    n = len(arr)
    if n < window_size:
        return np.nan, np.nan, np.nan
    # 用cumsum高效计算所有连续窗口的均值
    cumsum = np.concatenate([[0], np.cumsum(arr)])
    window_means = (cumsum[window_size:] - cumsum[:-window_size]) / window_size
    peak_avg = float(np.nanmax(window_means))
    valley_avg = float(np.nanmin(window_means))
    return peak_avg, valley_avg, peak_avg - valley_avg


def _calculate_station_stats_from_db(station_name, year):
    """从数据库计算单个站点的电价差统计"""
    df = load_price_data_from_db(station_name, year)
    if df is None or len(df.columns) < 2:
        return None
    
    price_data = df.iloc[:, 1:]  # 除日期列外的所有列
    if price_data.empty:
        return None
    
    # 计算日均峰谷价差（滑动窗口：连续N个时段均值最高 - 连续N个时段均值最低）
    n_peak = 8  # 窗口大小

    def _calc_spread(row):
        p, v, s = _sliding_window_spread(row.values, n_peak)
        return pd.Series({'peak': p, 'valley': v, 'spread': s})

    spread_df = price_data.apply(_calc_spread, axis=1)
    daily_peak_avg = spread_df['peak']
    daily_valley_avg = spread_df['valley']
    daily_spread = spread_df['spread']
    
    all_prices = price_data.to_numpy().flatten()
    
    return {
        '日均峰谷价差': round(daily_spread.mean(), 4),
        '全年最高峰谷价差': round(daily_spread.max(), 4),
        '全年最低峰谷价差': round(daily_spread.min(), 4),
        '全年平均电价': round(all_prices.mean(), 4),
        '数据天数': len(df)
    }


@st.cache_data(show_spinner=False)
def calculate_all_stations_price_spread_from_db(station_year_pairs):
    """从数据库计算所有站点的电价差统计数据
    
    Args:
        station_year_pairs: [(station_name, year), ...] 站点和年份的元组列表
    
    Returns:
        (stats_df, failed_stations, cache_summary)
    """
    if not station_year_pairs:
        return pd.DataFrame(), (), {"cached_count": 0, "recomputed_count": 0}
    
    station_stats = []
    failed_stations = []
    
    for station_name, year in station_year_pairs:
        try:
            stats = _calculate_station_stats_from_db(station_name, year)
            if stats is None:
                failed_stations.append((station_name, "无数据"))
                continue
            
            station_stats.append({
                '站点名称': station_name,
                **stats
            })
        except Exception as e:
            failed_stations.append((station_name, str(e)))
    
    if not station_stats:
        return pd.DataFrame(), tuple(failed_stations), {
            "cached_count": 0,
            "recomputed_count": len(station_year_pairs)
        }
    
    stats_df = pd.DataFrame(station_stats)
    stats_df = stats_df.sort_values('日均峰谷价差', ascending=False).reset_index(drop=True)
    stats_df['排名'] = range(1, len(stats_df) + 1)
    stats_df = stats_df[['排名', '站点名称', '日均峰谷价差', '全年最高峰谷价差', '全年最低峰谷价差', '全年平均电价', '数据天数']]
    
    return stats_df, tuple(failed_stations), {
        "cached_count": 0,
        "recomputed_count": len(station_year_pairs)
    }


def export_db_data_to_excel(station_name, year, progress_callback=None):
    """将数据库中的电价数据导出到本地Excel文件
    
    Args:
        station_name: 站点名称
        year: 年份
        progress_callback: 进度回调函数 callback(current, total, message)
    
    Returns:
        (success, message, file_path)
    """
    try:
        # 从数据库加载数据
        if progress_callback:
            progress_callback(0, 100, f"正在从数据库加载 {station_name} 的数据...")
        
        df = load_price_data_from_db(station_name, year)
        if df is None:
            return False, f"无法从数据库加载 {station_name} 的 {year} 年数据", None
        
        if progress_callback:
            progress_callback(30, 100, "数据加载完成，正在准备保存...")
        
        # 创建保存目录
        save_dir = PRICE_DATA_DIR / str(year)
        save_dir.mkdir(parents=True, exist_ok=True)
        
        # 生成文件名（只包含站点名）
        file_name = f"{station_name}.xlsx"
        file_path = save_dir / file_name
        
        if progress_callback:
            progress_callback(50, 100, f"正在保存到 {file_name}...")
        
        # 保存到Excel（覆盖已有文件）
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        if progress_callback:
            progress_callback(100, 100, "保存完成！")
        
        return True, f"成功导出到 {file_path}", str(file_path)
    except Exception as e:
        return False, f"导出失败: {str(e)}", None


def export_all_db_data_to_excel(station_list, year, progress_placeholder=None):
    """批量将数据库中的电价数据导出到本地Excel文件
    
    Args:
        station_list: 站点名称列表
        year: 年份
        progress_placeholder: Streamlit placeholder用于显示进度
    
    Returns:
        (success_count, fail_count, failed_stations)
    """
    success_count = 0
    fail_count = 0
    failed_stations = []
    total = len(station_list)
    
    # 创建进度条
    if progress_placeholder:
        progress_bar = progress_placeholder.progress(0)
        status_text = progress_placeholder.empty()
    
    for i, station_name in enumerate(station_list):
        # 更新进度
        progress = int((i / total) * 100)
        if progress_placeholder:
            progress_bar.progress(progress)
            status_text.text(f"正在处理 ({i+1}/{total}): {station_name}")
        
        # 导出单个站点
        success, message, _ = export_db_data_to_excel(station_name, year)
        if success:
            success_count += 1
        else:
            fail_count += 1
            failed_stations.append((station_name, message))
    
    # 完成
    if progress_placeholder:
        progress_bar.progress(100)
        status_text.text(f"导出完成！成功: {success_count}, 失败: {fail_count}")
    
    return success_count, fail_count, failed_stations


def sort_price_data_dirs(directories):
    """按年份倒序排列目录，非纯数字目录排在后面。"""
    def sort_key(path):
        folder_name = path.name.strip()
        if folder_name.isdigit():
            return (0, -int(folder_name))
        return (1, folder_name)

    return sorted(directories, key=sort_key)


def get_price_data_dir_options():
    """获取可用的电价数据年份目录；如果没有子目录，则回退到根目录。"""
    if not PRICE_DATA_DIR.exists():
        return []

    year_dirs = sort_price_data_dirs(
        [path for path in PRICE_DATA_DIR.iterdir() if path.is_dir()]
    )

    if year_dirs:
        return [(path.name, path) for path in year_dirs]

    return [("当前目录", PRICE_DATA_DIR)]


# 扫描所有电价数据文件
def scan_price_files(price_data_dir):
    """扫描指定目录下的所有电价Excel文件"""
    price_files = {}

    if not price_data_dir.exists():
        st.error(f"电价数据目录不存在: {price_data_dir}")
        return price_files

    for file_path in sorted(price_data_dir.iterdir()):
        if not file_path.is_file():
            continue
        if file_path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        if file_path.name.startswith("~$"):
            continue

        # 从文件名提取站点名称，兼容旧命名和“站点名.xlsx”两种格式
        file_name = file_path.stem
        if file_name in {"电站名", "电价名"}:
            continue

        station_name = (
            file_name
            .replace("电价数据（一年）", "")
            .replace("电价数据(一年)", "")
            .replace("电价数据", "")
            .strip()
        )

        if station_name:
            price_files[station_name] = file_path

    return price_files

# 加载电价数据
def get_file_cache_key(file_path):
    """生成文件缓存键，文件更新后自动失效缓存。"""
    file_stat = file_path.stat()
    return str(file_path), file_stat.st_mtime_ns, file_stat.st_size


@st.cache_data(show_spinner=False)
def load_price_data(file_path_str, modified_time_ns, file_size):
    """加载指定站点的电价数据"""
    try:
        df = pd.read_excel(file_path_str)
        return df
    except Exception as e:
        st.error(f"加载数据失败: {e}")
        return None


@st.cache_data(show_spinner=False)
def load_station_all_years_data(station_name, data_source, _price_dir_labels=None, selected_year=None):
    """加载指定站点所有年份的数据并合并为一个 DataFrame。

    Args:
        station_name: 站点名称
        data_source: 'database' 或 'excel'
        _price_dir_labels: Excel模式下所有年份目录的 (label, path) 列表（用于缓存失效）
        selected_year: 数据库模式下的年份

    Returns:
        合并后的 DataFrame，或 None
    """
    dfs = []
    if data_source == 'database':
        for year in get_db_available_years():
            df_year = load_price_data_from_db(station_name, year)
            if df_year is not None and len(df_year.columns) >= 2:
                dfs.append(df_year)
    else:
        price_dir_options = get_price_data_dir_options()
        for dir_label, dir_path in price_dir_options:
            station_file = dir_path / f"{station_name}.xlsx"
            if not station_file.exists():
                continue
            file_key = get_file_cache_key(station_file)
            df_year = load_price_data(*file_key)
            if df_year is not None and len(df_year.columns) >= 2:
                dfs.append(df_year)

    if not dfs:
        return None

    combined = pd.concat(dfs, ignore_index=True)
    date_col = combined.columns[0]
    combined[date_col] = pd.to_datetime(combined[date_col], errors='coerce')
    combined = combined.dropna(subset=[date_col])
    combined = combined.sort_values(date_col).drop_duplicates(subset=[date_col], keep='first').reset_index(drop=True)
    return combined


@st.cache_data(show_spinner="正在计算全省平均电价...")
def compute_provincial_average(file_keys):
    """计算全省所有站点各时段平均电价，返回(date_list, time_columns, avg_by_date)"""
    all_daily_arrays = []  # list of (date_string, 96-element array)

    for file_path_str, modified_time_ns, file_size in file_keys:
        df = load_price_data(file_path_str, modified_time_ns, file_size)
        if df is None or len(df.columns) < 2:
            continue

        date_col_name = df.columns[0]
        price_cols = df.columns[1:]
        if len(price_cols) < 96:
            continue
        price_cols = price_cols[:96]

        for _, row in df.iterrows():
            date_val = row[date_col_name]
            try:
                date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
            except Exception:
                date_str = str(date_val)[:10]
            prices = pd.to_numeric(row[price_cols], errors='coerce').values.astype(float)
            if len(prices) == 96 and not np.all(np.isnan(prices)):
                all_daily_arrays.append((date_str, prices))

    if not all_daily_arrays:
        return None, None, None

    # Group by date and compute mean
    date_grouped = {}
    for date_str, prices in all_daily_arrays:
        if date_str not in date_grouped:
            date_grouped[date_str] = []
        date_grouped[date_str].append(prices)

    date_list = sorted(date_grouped.keys())
    avg_by_date = np.array([np.mean(date_grouped[d], axis=0) for d in date_list])

    time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]

    return date_list, time_columns, avg_by_date

def build_price_file_index(price_files_dict):
    """构造稳定的文件索引，便于缓存排名结果。"""
    file_index = []
    for station_name in sorted(price_files_dict):
        file_path = price_files_dict[station_name]
        file_path_str, modified_time_ns, file_size = get_file_cache_key(file_path)
        file_index.append((station_name, file_path_str, modified_time_ns, file_size))
    return tuple(file_index)


def load_provincial_avg_cache():
    """读取全省平均电价的磁盘缓存。"""
    if not PROVINCIAL_AVG_CACHE_FILE.exists():
        return {}
    
    try:
        with open(PROVINCIAL_AVG_CACHE_FILE, "rb") as cache_file:
            cache_payload = pickle.load(cache_file)
        
        if cache_payload.get("version") != PROVINCIAL_AVG_CACHE_VERSION:
            return {}
        
        return cache_payload.get("entries", {})
    except Exception:
        return {}


def save_provincial_avg_cache(cache_data):
    """将全省平均电价缓存写入磁盘。"""
    cache_payload = {
        "version": PROVINCIAL_AVG_CACHE_VERSION,
        "entries": cache_data
    }
    temp_cache_file = PROVINCIAL_AVG_CACHE_FILE.with_suffix(".tmp")
    
    with open(temp_cache_file, "wb") as cache_file:
        pickle.dump(cache_payload, cache_file, protocol=pickle.HIGHEST_PROTOCOL)
    
    os.replace(temp_cache_file, PROVINCIAL_AVG_CACHE_FILE)


def _provincial_avg_cache_key(file_keys):
    """根据文件索引生成稳定的缓存键（MD5），用于区分不同年份/目录。"""
    key_parts = [f"{fp}|{mtime}|{size}" for fp, mtime, size in file_keys]
    raw = "\n".join(key_parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _is_provincial_avg_entry_valid(cached_entry, file_keys):
    """校验单条全省平均电价缓存条目是否仍然有效。"""
    if not isinstance(cached_entry, dict):
        return False

    cached_file_keys = cached_entry.get("file_keys", [])
    if len(cached_file_keys) != len(file_keys):
        return False

    for cached_key, current_key in zip(cached_file_keys, file_keys):
        if cached_key[1] != current_key[1] or cached_key[2] != current_key[2]:
            return False

    return True


def compute_provincial_average_with_cache(file_keys):
    """带磁盘缓存的全省平均电价计算，支持多年份缓存共存。"""
    cache_key = _provincial_avg_cache_key(file_keys)
    disk_cache = load_provincial_avg_cache()

    # 按哈希键查找对应条目
    cached_entry = disk_cache.get(cache_key)
    if cached_entry is not None and _is_provincial_avg_entry_valid(cached_entry, file_keys):
        date_list = cached_entry["date_list"]
        time_columns = cached_entry["time_columns"]
        avg_by_date = np.array(cached_entry["avg_by_date"])
        return date_list, time_columns, avg_by_date

    # 缓存未命中，重新计算
    date_list, time_columns, avg_by_date = compute_provincial_average(file_keys)

    if date_list is not None and time_columns is not None and avg_by_date is not None:
        # 将新条目合并到已有缓存中（保留其他年份的缓存）
        disk_cache[cache_key] = {
            "file_keys": file_keys,
            "date_list": date_list,
            "time_columns": time_columns,
            "avg_by_date": avg_by_date.tolist()
        }
        save_provincial_avg_cache(disk_cache)

    return date_list, time_columns, avg_by_date


def load_ranking_cache_entries():
    """读取磁盘缓存，支持跨 Streamlit 重启复用。"""
    if not RANKING_CACHE_FILE.exists():
        return {}

    try:
        with open(RANKING_CACHE_FILE, "rb") as cache_file:
            cache_payload = pickle.load(cache_file)

        if cache_payload.get("version") != RANKING_CACHE_VERSION:
            return {}

        cache_entries = cache_payload.get("entries", {})
        return cache_entries if isinstance(cache_entries, dict) else {}
    except Exception:
        return {}


def is_valid_ranking_cache_entry(cached_entry, modified_time_ns, file_size):
    """判断缓存条目是否仍可直接复用。"""
    return (
        cached_entry
        and cached_entry.get("modified_time_ns") == modified_time_ns
        and cached_entry.get("file_size") == file_size
        and isinstance(cached_entry.get("stats"), dict)
    )


def save_ranking_cache_entries(cache_entries):
    """将排名统计缓存写入磁盘。"""
    cache_payload = {
        "version": RANKING_CACHE_VERSION,
        "entries": cache_entries
    }
    temp_cache_file = RANKING_CACHE_FILE.with_suffix(".tmp")

    with open(temp_cache_file, "wb") as cache_file:
        pickle.dump(cache_payload, cache_file, protocol=pickle.HIGHEST_PROTOCOL)

    os.replace(temp_cache_file, RANKING_CACHE_FILE)


def merge_ranking_cache_entries(disk_cache_entries, current_cache_entries, price_file_index):
    """合并当前批次缓存，同时保留其他年份目录的缓存条目。"""
    merged_cache_entries = dict(disk_cache_entries)
    current_file_paths = {file_path_str for _, file_path_str, _, _ in price_file_index}
    current_parent_dirs = {str(Path(file_path_str).parent) for file_path_str in current_file_paths}

    if len(current_parent_dirs) == 1:
        current_parent_dir = next(iter(current_parent_dirs))
        merged_cache_entries = {
            cache_path: cache_entry
            for cache_path, cache_entry in merged_cache_entries.items()
            if str(Path(cache_path).parent) != current_parent_dir or cache_path in current_file_paths
        }

    merged_cache_entries.update(current_cache_entries)
    return merged_cache_entries


def _coerce_numeric(value):
    """将单元格值尽量转为浮点数。"""
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _calculate_station_stats_with_openpyxl(file_path_str):
    """使用只读方式扫描 Excel，避免为排名统计创建整张 DataFrame。"""
    workbook = None
    try:
        workbook = load_workbook(file_path_str, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)  # 跳过表头

        n_peak = 8  # 取最高的8个时段
        spread_sum = 0.0
        spread_max = None
        spread_min = None
        day_count = 0
        total_price_sum = 0.0
        total_price_count = 0

        for row in rows:
            # 收集该行所有电价数值
            row_prices = []
            row_sum = 0.0
            row_count = 0

            for cell_value in row[1:]:
                numeric_value = _coerce_numeric(cell_value)
                if numeric_value is None:
                    continue
                row_prices.append(numeric_value)
                row_sum += numeric_value
                row_count += 1

            if row_count < n_peak * 2:
                continue

            # 滑动窗口取连续N个时段均值最高和最低
            peak_avg, valley_avg, daily_spread = _sliding_window_spread(row_prices, n_peak)

            spread_sum += daily_spread
            spread_max = daily_spread if spread_max is None else max(spread_max, daily_spread)
            spread_min = daily_spread if spread_min is None else min(spread_min, daily_spread)
            total_price_sum += row_sum
            total_price_count += row_count
            day_count += 1

        if day_count == 0 or total_price_count == 0:
            return None

        return {
            '日均峰谷价差': round(spread_sum / day_count, 4),
            '全年最高峰谷价差': round(spread_max, 4),
            '全年最低峰谷价差': round(spread_min, 4),
            '全年平均电价': round(total_price_sum / total_price_count, 4),
            '数据天数': day_count
        }
    finally:
        if workbook is not None:
            workbook.close()


def _calculate_station_stats_with_pandas(file_path_str):
    """兼容 .xls 文件的统计逻辑。"""
    df = pd.read_excel(file_path_str)
    price_data = df.iloc[:, 1:]

    if price_data.empty:
        return None

    # 计算日均峰谷价差（滑动窗口：连续N个时段均值最高 - 连续N个时段均值最低）
    n_peak = 8  # 窗口大小

    def _calc_spread(row):
        p, v, s = _sliding_window_spread(row.values, n_peak)
        return pd.Series({'peak': p, 'valley': v, 'spread': s})

    spread_df = price_data.apply(_calc_spread, axis=1)
    daily_peak_avg = spread_df['peak']
    daily_valley_avg = spread_df['valley']
    daily_spread = spread_df['spread']
    
    all_prices = price_data.to_numpy().flatten()

    return {
        '日均峰谷价差': round(daily_spread.mean(), 4),
        '全年最高峰谷价差': round(daily_spread.max(), 4),
        '全年最低峰谷价差': round(daily_spread.min(), 4),
        '全年平均电价': round(all_prices.mean(), 4),
        '数据天数': len(df)
    }


def calculate_station_price_stats(file_meta):
    """计算单个站点的统计值。"""
    station_name, file_path_str, modified_time_ns, file_size = file_meta

    try:
        file_suffix = Path(file_path_str).suffix.lower()
        if file_suffix == ".xlsx":
            stats = _calculate_station_stats_with_openpyxl(file_path_str)
        else:
            stats = _calculate_station_stats_with_pandas(file_path_str)

        if stats is None:
            return None

        return {
            '站点名称': station_name,
            **stats
        }
    except Exception as e:
        return {
            '站点名称': station_name,
            '__error__': str(e)
        }


# 计算所有站点的电价差统计
@st.cache_data(show_spinner=False)
def calculate_all_stations_price_spread(price_file_index):
    """并行计算所有站点的电价差统计数据，并复用磁盘缓存。"""
    if not price_file_index:
        return pd.DataFrame(), (), {"cached_count": 0, "recomputed_count": 0}

    disk_cache_entries = load_ranking_cache_entries()
    next_cache_entries = {}
    station_stats = []
    failed_stations = []
    pending_file_index = []
    cached_count = 0

    for file_meta in price_file_index:
        station_name, file_path_str, modified_time_ns, file_size = file_meta
        cached_entry = disk_cache_entries.get(file_path_str)

        if is_valid_ranking_cache_entry(cached_entry, modified_time_ns, file_size):
            cached_stats = {
                "站点名称": station_name,
                **cached_entry["stats"]
            }
            station_stats.append(cached_stats)
            next_cache_entries[file_path_str] = cached_entry
            cached_count += 1
            continue

        pending_file_index.append(file_meta)

    max_workers = min(len(pending_file_index), max(4, min(12, (os.cpu_count() or 4) + 2))) if pending_file_index else 0

    if pending_file_index:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(calculate_station_price_stats, pending_file_index))
    else:
        results = []

    for file_meta, result in zip(pending_file_index, results):
        station_name, file_path_str, modified_time_ns, file_size = file_meta
        if not result:
            continue
        if '__error__' in result:
            failed_stations.append((result['站点名称'], result['__error__']))
            continue

        station_stats.append(result)
        next_cache_entries[file_path_str] = {
            "modified_time_ns": modified_time_ns,
            "file_size": file_size,
            "stats": {
                "日均峰谷价差": result["日均峰谷价差"],
                "全年最高峰谷价差": result["全年最高峰谷价差"],
                "全年最低峰谷价差": result["全年最低峰谷价差"],
                "全年平均电价": result["全年平均电价"],
                "数据天数": result["数据天数"]
            }
        }

    if not station_stats:
        return pd.DataFrame(), tuple(failed_stations), {
            "cached_count": cached_count,
            "recomputed_count": len(pending_file_index)
        }

    merged_cache_entries = merge_ranking_cache_entries(
        disk_cache_entries,
        next_cache_entries,
        price_file_index,
    )

    if merged_cache_entries != disk_cache_entries:
        save_ranking_cache_entries(merged_cache_entries)

    stats_df = pd.DataFrame(station_stats)
    stats_df = stats_df.sort_values('日均峰谷价差', ascending=False).reset_index(drop=True)
    stats_df['排名'] = range(1, len(stats_df) + 1)
    stats_df = stats_df[['排名', '站点名称', '日均峰谷价差', '全年最高峰谷价差', '全年最低峰谷价差', '全年平均电价', '数据天数']]
    return stats_df, tuple(failed_stations), {
        "cached_count": cached_count,
        "recomputed_count": len(pending_file_index)
    }





def classify_factory_station_group(station_name):
    """根据站点名称判断是否属于电厂。"""
    station_name_str = str(station_name).strip()
    return FACTORY_STATION_LABEL if "厂" in station_name_str else NON_FACTORY_STATION_LABEL


def get_guangdong_period_type(date_value, time_str):
    """按广东省峰谷分时电价规则判定时段类型。"""
    try:
        hour_str, minute_str = str(time_str).split(':', 1)
        minutes = int(hour_str) * 60 + int(minute_str)
    except (TypeError, ValueError):
        return "未知"

    date_ts = pd.to_datetime(date_value, errors='coerce')
    # 广东尖峰还包括其他月份的高温天，但程序当前未接入气象数据，
    # 因此先按固定的 7-9 月整月识别尖峰时段。
    is_sharp_peak_month = pd.notna(date_ts) and date_ts.month in GUANGDONG_SHARP_PEAK_MONTHS

    if is_sharp_peak_month and (
        11 * 60 <= minutes < 12 * 60 or
        15 * 60 <= minutes < 17 * 60
    ):
        return "尖峰"

    if 10 * 60 <= minutes < 12 * 60 or 14 * 60 <= minutes < 19 * 60:
        return "高峰"

    if 0 <= minutes < 8 * 60:
        return "低谷"

    return "平段"


def build_station_group_mapping(station_info_df, group_column, station_names=None):
    """构造电站与分组字段的映射表。"""
    if group_column == FACTORY_GROUP_COLUMN:
        if station_names is None:
            return None

        station_names_list = (
            pd.Series(list(station_names))
            .dropna()
            .astype(str)
            .str.strip()
        )
        station_names_list = station_names_list[station_names_list != ""].drop_duplicates().tolist()

        if not station_names_list:
            return None

        return pd.DataFrame({
            '电站名': station_names_list,
            group_column: [classify_factory_station_group(name) for name in station_names_list]
        })

    if (
        station_info_df is None
        or '电站名' not in station_info_df.columns
        or group_column not in station_info_df.columns
    ):
        return None

    station_group_df = (
        station_info_df[['电站名', group_column]]
        .copy()
        .dropna(subset=['电站名'])
    )
    station_group_df['电站名'] = station_group_df['电站名'].astype(str).str.strip()
    station_group_df[group_column] = (
        station_group_df[group_column]
        .fillna('未分组')
        .astype(str)
        .str.strip()
        .replace('', '未分组')
    )
    station_group_df = station_group_df.drop_duplicates(subset=['电站名'], keep='first')
    return station_group_df


def get_available_group_values(station_info_df, group_column, station_names=None):
    """获取指定字段下可用的分组选项。"""
    station_group_df = build_station_group_mapping(station_info_df, group_column, station_names=station_names)
    if station_group_df is None:
        return []

    if group_column == FACTORY_GROUP_COLUMN:
        ordered_values = [FACTORY_STATION_LABEL, NON_FACTORY_STATION_LABEL]
        available_values = set(station_group_df[group_column].dropna().tolist())
        return [value for value in ordered_values if value in available_values]

    return sorted(station_group_df[group_column].dropna().unique().tolist())


def get_station_group_value(station_info_df, station_name, group_column):
    """获取单个站点对应的分组字段值。"""
    station_group_df = build_station_group_mapping(
        station_info_df,
        group_column,
        station_names=[station_name]
    )
    if station_group_df is None:
        return None

    station_name_str = str(station_name).strip()
    matched_df = station_group_df[station_group_df['电站名'] == station_name_str]
    if matched_df.empty:
        return None

    group_value = str(matched_df.iloc[0][group_column]).strip()
    if not group_value or group_value == '未分组':
        return None

    return group_value


def prepare_grouped_rankings(all_stations_stats, station_info_df, group_column):
    """为排名结果补充分组信息和组内排名。"""
    station_group_df = build_station_group_mapping(
        station_info_df,
        group_column,
        station_names=all_stations_stats['站点名称'].tolist()
    )
    if station_group_df is None:
        return None

    grouped_df = all_stations_stats.merge(
        station_group_df,
        left_on='站点名称',
        right_on='电站名',
        how='left'
    ).drop(columns=['电站名'])

    grouped_df[group_column] = grouped_df[group_column].fillna('未分组')
    grouped_df = grouped_df.sort_values(
        [group_column, '日均峰谷价差', '站点名称'],
        ascending=[True, False, True]
    ).reset_index(drop=True)
    grouped_df['组内排名'] = grouped_df.groupby(group_column).cumcount() + 1
    grouped_df[f'{group_column}站点数'] = grouped_df.groupby(group_column)['站点名称'].transform('count')

    return grouped_df


def _shared_opt_cache_key(station_name, P, battery_capacity, efficiency, data_source, year_or_path):
    """生成共享储能优化缓存键。"""
    raw = f"{station_name}|{P}|{battery_capacity}|{efficiency}|{data_source}|{year_or_path}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def load_shared_opt_cache():
    """读取共享储能优化缓存（含完整优化结果）。"""
    if not SHARED_OPT_CACHE_FILE.exists():
        return {}
    try:
        with open(SHARED_OPT_CACHE_FILE, "rb") as f:
            payload = pickle.load(f)
        if payload.get("version") != SHARED_OPT_CACHE_VERSION:
            return {}
        return payload.get("entries", {})
    except Exception:
        return {}


def save_shared_opt_cache(entries):
    """将共享储能优化缓存写入磁盘。"""
    payload = {"version": SHARED_OPT_CACHE_VERSION, "entries": entries}
    tmp = SHARED_OPT_CACHE_FILE.with_suffix(".tmp")
    with open(tmp, "wb") as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    os.replace(tmp, SHARED_OPT_CACHE_FILE)


def get_shared_opt_result(cache_key):
    """从共享缓存获取优化结果，返回 (date_list, all_results, all_summaries) 或 None。"""
    cache = load_shared_opt_cache()
    entry = cache.get(cache_key)
    if entry and isinstance(entry, dict) and 'all_results' in entry:
        return entry['date_list'], entry['all_results'], entry['all_summaries']
    return None


def put_shared_opt_result(cache_key, date_list, all_results, all_summaries):
    """将优化结果存入共享缓存（线程安全）。"""
    with SHARED_OPT_CACHE_LOCK:
        cache = load_shared_opt_cache()
        cache[cache_key] = {
            'date_list': date_list,
            'all_results': all_results,
            'all_summaries': all_summaries,
        }
        save_shared_opt_cache(cache)


def _run_optimization_for_station(all_prices, config):
    """对给定价格数据执行智能优化，返回 (date_list, all_results, all_summaries, total_revenue)。"""
    num_days = len(all_prices)
    all_results = []
    all_summaries = []
    total_revenue = 0.0
    current_soc = 0.0

    for day_idx in range(num_days):
        price = all_prices[day_idx, :]
        day_start_soc = current_soc
        prob, result = optimize_single_day(price, day_idx, current_soc, config)

        if pulp.LpStatus[prob.status] == "Optimal":
            var_dict = {v.name: v for v in prob.variables()}
            charge_power_values = [float(pulp.value(var_dict[f"charge_{i}"])) for i in range(96)]
            discharge_power_values = [float(pulp.value(var_dict[f"discharge_{i}"])) for i in range(96)]
            soc_values = [float(pulp.value(var_dict[f"soc_{i}"])) for i in range(97)]
            current_soc = float(pulp.value(var_dict[f"soc_{96}"]))
            day_revenue = float(pulp.value(prob.objective))
            total_revenue += day_revenue

            dt = config['dt']
            for i in range(96):
                hour = i * dt
                time_str = f"{int(hour):02d}:{int((hour - int(hour)) * 60):02d}"
                if charge_power_values[i] > 1e-4:
                    period_type = "充电"
                elif discharge_power_values[i] > 1e-4:
                    period_type = "放电"
                else:
                    period_type = "空闲"
                all_results.append({
                    '日期': f"第{day_idx+1}天", '时间': time_str,
                    '电价_元/kWh': float(price[i]),
                    '充电功率_kW': charge_power_values[i],
                    '放电功率_kW': discharge_power_values[i],
                    '净功率_kW': discharge_power_values[i] - charge_power_values[i],
                    '电池电量_kWh': soc_values[i + 1],
                    '时段类型': period_type
                })

            total_charge = sum(charge_power_values[i] * dt for i in range(96))
            total_discharge = sum(discharge_power_values[i] * dt for i in range(96))
            all_summaries.append({
                '日期': f"第{day_idx+1}天",
                '日收益_元': day_revenue,
                '初始电量_kWh': day_start_soc,
                '最终电量_kWh': current_soc,
                '充电量_kWh': total_charge,
                '放电量_kWh': total_discharge
            })

    date_list = [f"第{i+1}天" for i in range(num_days)]
    return date_list, all_results, all_summaries, total_revenue


# 储能优化函数
def optimize_single_day(price, day_index, start_soc, config):
    """优化单天的储能调度"""
    num = config['num']
    dt = config['dt']
    P = config['P']
    battery_capacity = config['battery_capacity']
    efficiency = config['efficiency']
    
    # 创建问题
    prob = pulp.LpProblem(f"BESS_Optimization_Day{day_index + 1}", pulp.LpMaximize)
    
    # 定义变量
    charge_power = pulp.LpVariable.dicts("charge", range(num), lowBound=0)
    discharge_power = pulp.LpVariable.dicts("discharge", range(num), lowBound=0)
    soc = pulp.LpVariable.dicts("soc", range(num + 1), lowBound=0, upBound=battery_capacity)
    
    # 二进制变量
    z_ch = pulp.LpVariable.dicts("is_charging", range(num), cat="Binary")
    z_dis = pulp.LpVariable.dicts("is_discharging", range(num), cat="Binary")
    y_ch = pulp.LpVariable.dicts("charge_start", range(num), cat="Binary")
    y_dis = pulp.LpVariable.dicts("discharge_start", range(num), cat="Binary")
    
    epsilon = 1e-3
    min_duration = 4
    min_on_power = 0.05 * P
    
    # 目标函数
    prob += pulp.lpSum([(discharge_power[i] - charge_power[i]) * price[i] * dt for i in range(num)])
    
    # 初始电量约束
    prob += soc[0] == start_soc
    
    # 电池状态转移方程
    for i in range(num):
        prob += soc[i + 1] == soc[i] + charge_power[i] * dt - discharge_power[i] * dt / efficiency
    
    # 功率-状态关联约束
    for i in range(num):
        prob += charge_power[i] <= P * z_ch[i]
        prob += discharge_power[i] <= P * z_dis[i]
        prob += charge_power[i] >= min_on_power * z_ch[i]
        prob += discharge_power[i] >= min_on_power * z_dis[i]
        prob += z_ch[i] + z_dis[i] <= 1
        
        if i == 0:
            prob += y_ch[i] == z_ch[i]
            prob += y_dis[i] == z_dis[i]
        else:
            prob += y_ch[i] >= z_ch[i] - z_ch[i-1]
            prob += y_ch[i] <= z_ch[i]
            prob += y_ch[i] <= 1 - z_ch[i-1]
            prob += y_dis[i] >= z_dis[i] - z_dis[i-1]
            prob += y_dis[i] <= z_dis[i]
            prob += y_dis[i] <= 1 - z_dis[i-1]
    
    # 最小连续时长约束
    for i in range(num):
        if i <= num - min_duration:
            prob += pulp.lpSum(z_ch[j] for j in range(i, i + min_duration)) >= min_duration * y_ch[i]
            prob += pulp.lpSum(z_dis[j] for j in range(i, i + min_duration)) >= min_duration * y_dis[i]
        else:
            prob += y_ch[i] == 0
            prob += y_dis[i] == 0
    
    # 累计能量平衡约束（允许电量留存）
    for k in range(1, num+1):
        prob += efficiency * pulp.lpSum([charge_power[i] * dt for i in range(k)]) + start_soc >= pulp.lpSum(
            [discharge_power[i] * dt for i in range(k)])
    
    # 连续充电能量约束
    M = int(math.ceil(battery_capacity / (P * dt))) + 1
    for s in range(0, 97 - M):
        end_idx = min(s + M - 1, 95)
        prob += pulp.lpSum([charge_power[j] * dt for j in range(s, end_idx + 1)]) <= battery_capacity
    
    # 求解
    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=120)
    result = prob.solve(solver)
    
    return prob, result


def simulate_custom_strategy(prices, charge_slots, discharge_slots, config):
    """模拟单天自定义充放电策略

    Args:
        prices: 96个时段的电价数组
        charge_slots: 充电时段索引列表 (0~95)
        discharge_slots: 放电时段索引列表 (0~95)
        config: {'P', 'battery_capacity', 'efficiency', 'dt', 'num'}

    Returns:
        dict: 模拟结果
    """
    num = config['num']
    dt = config['dt']
    P = config['P']
    battery_capacity = config['battery_capacity']
    efficiency = config['efficiency']

    charge_power = np.zeros(num)
    discharge_power = np.zeros(num)
    soc = np.zeros(num + 1)

    charge_set = set(charge_slots)
    discharge_set = set(discharge_slots)

    for i in range(num):
        soc_before = soc[i]

        if i in charge_set:
            max_charge_energy = (battery_capacity - soc_before) / efficiency
            actual_energy = min(P * dt, max_charge_energy)
            charge_power[i] = actual_energy / dt
            soc[i + 1] = soc_before + actual_energy * efficiency
        elif i in discharge_set:
            max_discharge_energy = soc_before * efficiency
            actual_energy = min(P * dt, max_discharge_energy)
            discharge_power[i] = actual_energy / dt
            soc[i + 1] = soc_before - actual_energy / efficiency
        else:
            soc[i + 1] = soc_before

    total_charge = np.sum(charge_power * dt)
    total_discharge = np.sum(discharge_power * dt)
    revenue = np.sum((discharge_power - charge_power) * prices * dt)

    return {
        'charge_power': charge_power,
        'discharge_power': discharge_power,
        'soc': soc[1:],
        'total_charge': total_charge,
        'total_discharge': total_discharge,
        'revenue': revenue,
    }


def run_custom_strategy(all_prices, date_list, charge_slots, discharge_slots, config):
    """遍历全年每天执行自定义策略

    Args:
        all_prices: (num_days, 96) 二维数组
        date_list: 日期列表
        charge_slots: 充电时段索引列表
        discharge_slots: 放电时段索引列表
        config: 储能参数

    Returns:
        (all_results, all_summaries)
    """
    all_results = []
    all_summaries = []

    for day_idx, prices in enumerate(all_prices):
        result = simulate_custom_strategy(prices, charge_slots, discharge_slots, config)

        dt = config['dt']
        for i in range(96):
            hour = i * dt
            time_str = f"{int(hour):02d}:{int((hour - int(hour)) * 60):02d}"
            if result['charge_power'][i] > 1e-4:
                period_type = "充电"
            elif result['discharge_power'][i] > 1e-4:
                period_type = "放电"
            else:
                period_type = "空闲"
            all_results.append({
                '日期': date_list[day_idx],
                '时间': time_str,
                '电价_元/kWh': float(prices[i]),
                '充电功率_kW': float(result['charge_power'][i]),
                '放电功率_kW': float(result['discharge_power'][i]),
                '净功率_kW': float(result['discharge_power'][i] - result['charge_power'][i]),
                '电池电量_kWh': float(result['soc'][i]),
                '时段类型': period_type
            })

        all_summaries.append({
            '日期': date_list[day_idx],
            '日收益_元': result['revenue'],
            '初始电量_kWh': 0.0,
            '最终电量_kWh': float(result['soc'][-1]),
            '充电量_kWh': result['total_charge'],
            '放电量_kWh': result['total_discharge']
        })

    return all_results, all_summaries


def _extract_day_results(prob, day_idx, day_date, price, config):
    """从优化结果中提取单日的详细数据和汇总。"""
    dt = config['dt']
    var_dict = {v.name: v for v in prob.variables()}
    charge_power_values = [float(pulp.value(var_dict[f"charge_{i}"])) for i in range(96)]
    discharge_power_values = [float(pulp.value(var_dict[f"discharge_{i}"])) for i in range(96)]
    soc_values = [float(pulp.value(var_dict[f"soc_{i}"])) for i in range(97)]
    final_soc = float(pulp.value(var_dict[f"soc_{96}"]))
    total_revenue = float(pulp.value(prob.objective))

    day_results = []
    for i in range(96):
        hour = i * dt
        time_str = f"{int(hour):02d}:{int((hour - int(hour)) * 60):02d}"
        if charge_power_values[i] > 1e-4:
            period_type = "充电"
        elif discharge_power_values[i] > 1e-4:
            period_type = "放电"
        else:
            period_type = "空闲"
        day_results.append({
            '日期': day_date, '时间': time_str,
            '电价_元/kWh': float(price[i]),
            '充电功率_kW': charge_power_values[i],
            '放电功率_kW': discharge_power_values[i],
            '净功率_kW': discharge_power_values[i] - charge_power_values[i],
            '电池电量_kWh': soc_values[i + 1],
            '时段类型': period_type
        })

    total_charge = sum(charge_power_values[i] * dt for i in range(96))
    total_discharge = sum(discharge_power_values[i] * dt for i in range(96))
    day_summary = {
        '日期': day_date,
        '日收益_元': total_revenue,
        '初始电量_kWh': 0.0,
        '最终电量_kWh': final_soc,
        '充电量_kWh': total_charge,
        '放电量_kWh': total_discharge
    }

    return day_results, day_summary


def _optimize_single_day_wrapper(args):
    """并行优化的包装函数，每天初始电量都从零开始。"""
    day_idx, price, day_date, config = args
    prob, result = optimize_single_day(price, day_idx, 0.0, config)
    if pulp.LpStatus[prob.status] == "Optimal":
        return _extract_day_results(prob, day_idx, day_date, price, config)
    return None


@st.cache_data(show_spinner="正在优化储能调度...")
def run_optimization_cached(file_path_or_none, P, battery_capacity, efficiency):
    """缓存化的多天储能优化，相同参数+相同文件只计算一次

    Args:
        P: 逆变器功率 (MW)
        battery_capacity: 电池容量 (MWh)
        efficiency: 放电效率
    """
    # 将MW/MWh转换为kW/kWh进行内部计算
    P_kw = P * 1000  # MW -> kW
    battery_capacity_kwh = battery_capacity * 1000  # MWh -> kWh
    
    config = {
        'P': P_kw, 'battery_capacity': battery_capacity_kwh,
        'initial_soc': 0, 'efficiency': efficiency,
        'dt': 0.25, 'num': 96
    }
    df = pd.read_excel(file_path_or_none)
    start_col = 1 if len(df.columns) > 96 and not pd.api.types.is_numeric_dtype(df.iloc[:, 0]) else 0
    all_prices = df.iloc[:, start_col:start_col + 96].values.astype(float)

    date_list = []
    if start_col == 1:
        for d in df.iloc[:, 0].values:
            if pd.isna(d):
                date_list.append('')
            elif isinstance(d, pd.Timestamp):
                date_list.append(d.strftime('%Y-%m-%d'))
            elif isinstance(d, str):
                ds = str(d).strip()
                date_list.append(ds.split('T')[0] if 'T' in ds else (ds.split(' ')[0] if ' ' in ds else ds[:10]))
            else:
                date_list.append(str(d)[:10])
    else:
        date_list = [f'第{i+1}天' for i in range(len(df))]

    num_days = len(all_prices)

    # 准备并行任务参数（每天初始电量都从零开始）
    tasks = [(day_idx, all_prices[day_idx], date_list[day_idx], config) for day_idx in range(num_days)]

    max_workers = min(num_days, max(4, min(12, (os.cpu_count() or 4) + 2)))
    all_results = []
    all_summaries = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(_optimize_single_day_wrapper, tasks))

    for day_idx, day_result in enumerate(results):
        if day_result is not None:
            day_results, day_summary = day_result
            all_results.extend(day_results)
            all_summaries.append(day_summary)

    return date_list, all_results, all_summaries


@st.cache_data(show_spinner="正在优化储能调度...")
def run_optimization_cached_from_db(station_name, year, P, battery_capacity, efficiency):
    """从数据库加载数据并进行储能优化
    
    Args:
        station_name: 站点名称
        year: 年份
        P: 逆变器功率 (MW)
        battery_capacity: 电池容量 (MWh)
        efficiency: 放电效率
    """
    # 将MW/MWh转换为kW/kWh进行内部计算
    P_kw = P * 1000  # MW -> kW
    battery_capacity_kwh = battery_capacity * 1000  # MWh -> kWh
    
    config = {
        'P': P_kw, 'battery_capacity': battery_capacity_kwh,
        'initial_soc': 0, 'efficiency': efficiency,
        'dt': 0.25, 'num': 96
    }
    
    # 从数据库加载数据
    df = load_price_data_from_db(station_name, year)
    if df is None:
        raise ValueError(f"无法从数据库加载 {station_name} 的 {year} 年数据")
    
    start_col = 1 if len(df.columns) > 96 and not pd.api.types.is_numeric_dtype(df.iloc[:, 0]) else 0
    all_prices = df.iloc[:, start_col:start_col + 96].values.astype(float)

    date_list = []
    if start_col == 1:
        for d in df.iloc[:, 0].values:
            if pd.isna(d):
                date_list.append('')
            elif isinstance(d, pd.Timestamp):
                date_list.append(d.strftime('%Y-%m-%d'))
            elif isinstance(d, str):
                ds = str(d).strip()
                date_list.append(ds.split('T')[0] if 'T' in ds else (ds.split(' ')[0] if ' ' in ds else ds[:10]))
            else:
                date_list.append(str(d)[:10])
    else:
        date_list = [f'第{i+1}天' for i in range(len(df))]

    num_days = len(all_prices)

    # 准备并行任务参数（每天初始电量都从零开始）
    tasks = [(day_idx, all_prices[day_idx], date_list[day_idx], config) for day_idx in range(num_days)]

    max_workers = min(num_days, max(4, min(12, (os.cpu_count() or 4) + 2)))
    all_results = []
    all_summaries = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(_optimize_single_day_wrapper, tasks))

    for day_idx, day_result in enumerate(results):
        if day_result is not None:
            day_results, day_summary = day_result
            all_results.extend(day_results)
            all_summaries.append(day_summary)

    return date_list, all_results, all_summaries


# 主程序
def main():
    # 加载电站信息
    station_info_df = load_station_info()

    # 数据源选择
    st.sidebar.header("📂 数据源配置")
    data_source = st.sidebar.radio(
        "选择数据源",
        options=["数据库", "Excel文件"],
        index=1,  # 默认为Excel文件模式
        help="选择从数据库或Excel文件读取电价数据"
    )

    # 检测数据库可用性
    db_available, db_error = check_db_available() if data_source == "数据库" else (False, None)
    
    # 数据库不可用时自动切换到Excel模式
    if data_source == "数据库" and not db_available:
        st.warning(f"数据库连接失败: {db_error}")
        st.info("已自动切换到Excel文件模式，请确保电价数据目录中有数据文件。")
        data_source = "Excel文件"

    if data_source == "数据库":
        # 从数据库加载数据
        db_stations = get_db_station_list()
        if not db_stations:
            st.warning("数据库中未找到站点数据！")
            st.info("请检查数据库连接配置或确认数据库中已有电价数据。")
            return
        
        db_years = get_db_available_years()
        if not db_years:
            st.warning("数据库中未找到可用的年份数据！")
            return
        
        # 年份选择
        selected_year = st.sidebar.selectbox(
            "选择年份",
            options=db_years,
            index=0,
            help="选择要查询的年份"
        )
        
        # 构建price_files字典（兼容现有逻辑）
        # 使用站点名称作为键，None作为文件路径（因为从数据库加载）
        price_files = {station: None for station in db_stations}
        
        # 保存数据源信息供后续使用
        st.session_state['data_source'] = 'database'
        st.session_state['selected_year'] = selected_year
        
        # 数据同步功能
        st.sidebar.divider()
        st.sidebar.subheader("💾 数据同步")
        
        # 选择要同步的站点
        sync_mode = st.sidebar.radio(
            "同步范围",
            options=["全部站点", "选择站点"],
            horizontal=True,
            help="选择要同步到本地的站点范围"
        )
        
        if sync_mode == "选择站点":
            sync_stations = st.sidebar.multiselect(
                "选择站点",
                options=db_stations,
                default=[],
                help="选择要同步的站点"
            )
        else:
            sync_stations = db_stations
        
        # 同步按钮
        if st.sidebar.button("📤 同步到本地Excel", type="primary", disabled=len(sync_stations) == 0):
            st.sidebar.divider()
            st.sidebar.subheader("📊 同步进度")
            progress_placeholder = st.sidebar.empty()
            
            # 执行同步
            success_count, fail_count, failed_stations = export_all_db_data_to_excel(
                sync_stations, selected_year, progress_placeholder
            )
            
            # 显示结果
            if fail_count == 0:
                st.sidebar.success(f"✅ 同步完成！成功导出 {success_count} 个站点的数据")
            else:
                st.sidebar.warning(f"⚠️ 同步完成：成功 {success_count} 个，失败 {fail_count} 个")
                if failed_stations:
                    with st.sidebar.expander("查看失败详情"):
                        for station, error in failed_stations[:10]:
                            st.write(f"- {station}: {error}")
                        if len(failed_stations) > 10:
                            st.write(f"... 还有 {len(failed_stations) - 10} 个站点")
            
            # 同步后刷新文件列表
            st.sidebar.button("🔄 刷新文件列表", on_click=st.cache_data.clear)
        
    else:
        # 从Excel文件加载数据（原有逻辑）
        price_dir_options = get_price_data_dir_options()
        if not price_dir_options:
            st.warning("未找到电价数据目录！")
            st.info(f"请先创建电价数据目录: `{PRICE_DATA_DIR}`")
            return

        price_dir_labels = [label for label, _ in price_dir_options]
        default_price_dir_index = 0

        selected_price_dir_label = st.sidebar.selectbox(
            "电价数据年份",
            price_dir_labels,
            index=default_price_dir_index,
            help='读取"电价数据"目录下对应年份文件夹中的站点电价表。',
        )
        selected_price_dir = dict(price_dir_options)[selected_price_dir_label]

        # 扫描电价文件
        price_files = scan_price_files(selected_price_dir)

        if not price_files:
            st.warning("未找到电价数据文件！")
            st.info(f"请在以下目录放置电价文件: `{selected_price_dir}`")
            st.markdown("""
            **文件格式要求：**
            - Excel格式 (.xlsx/.xls)
            - 文件放在 `电价数据/年份` 文件夹内，例如 `电价数据/2025`
            - 第一列为日期列
            - 后续96列为00:00到23:45的96个时间节点电价数据
            """)
            return
        
        st.session_state['data_source'] = 'excel'

    has_busbar_info = (
        station_info_df is not None
        and '母线' in station_info_df.columns
        and '电站名' in station_info_df.columns
    )
    has_city_info = (
        station_info_df is not None
        and '城市' in station_info_df.columns
        and '电站名' in station_info_df.columns
    )
    busbar_types = []
    city_types = []

    if has_busbar_info:
        busbar_types = get_available_group_values(station_info_df, '母线')

    if has_city_info:
        # 直接从电站名表格获取所有城市，不按电价文件过滤
        city_types = (
            station_info_df['城市']
            .dropna()
            .astype(str)
            .str.strip()
            .replace('', pd.NA)
            .dropna()
            .unique()
            .tolist()
        )
        city_types = sorted(city_types)

    factory_group_types = get_available_group_values(
        station_info_df,
        FACTORY_GROUP_COLUMN,
        station_names=price_files.keys()
    )

    st.divider()
    if view_mode == "📊 电价数据查询":
        st.sidebar.header("📍 站点选择")

        # 级联筛选：城市 → 厂站类型 → 母线 → 站点
        filtered_stations = list(price_files.keys())

        # 1. 选择城市
        if city_types:
            city_options = ["全部城市"] + sorted(city_types)
            selected_city = st.sidebar.selectbox(
                "选择城市",
                options=city_options,
                index=0,
                help="选择要查看的城市"
            )
            
            if selected_city != "全部城市":
                station_group_df = build_station_group_mapping(
                    station_info_df,
                    "城市",
                    station_names=price_files.keys()
                )
                if station_group_df is not None:
                    city_stations_df = station_group_df[station_group_df["城市"] == selected_city]
                    city_stations_list = (
                        city_stations_df['电站名']
                        .astype(str)
                        .str.strip()
                        .tolist()
                    )
                    filtered_stations = [s for s in city_stations_list if s in price_files.keys()]
                    st.sidebar.info(f"📊 {selected_city} 共有 {len(filtered_stations)} 个站点")

        # 2. 在筛选后的站点中选择厂站类型
        if factory_group_types and len(filtered_stations) > 0:
            factory_group_options = get_available_group_values(
                station_info_df,
                FACTORY_GROUP_COLUMN,
                station_names=filtered_stations
            )
            
            if factory_group_options:
                all_factory_option = f"全部{FACTORY_GROUP_COLUMN}"
                factory_options = [all_factory_option] + factory_group_options
                selected_factory = st.sidebar.selectbox(
                    f"选择{FACTORY_GROUP_COLUMN}",
                    options=factory_options,
                    index=0,
                    help=f"选择要查看的{FACTORY_GROUP_COLUMN}"
                )
                
                if selected_factory != all_factory_option:
                    station_group_df = build_station_group_mapping(
                        station_info_df,
                        FACTORY_GROUP_COLUMN,
                        station_names=filtered_stations
                    )
                    if station_group_df is not None:
                        factory_stations_df = station_group_df[station_group_df[FACTORY_GROUP_COLUMN] == selected_factory]
                        factory_stations_list = (
                            factory_stations_df['电站名']
                            .astype(str)
                            .str.strip()
                            .tolist()
                        )
                        filtered_stations = [s for s in factory_stations_list if s in filtered_stations]
                        st.sidebar.info(f"📊 {selected_factory} 共有 {len(filtered_stations)} 个站点")

        # 3. 在筛选后的站点中选择母线
        if busbar_types and len(filtered_stations) > 0:
            busbar_options_filtered = get_available_group_values(
                station_info_df,
                "母线",
                station_names=filtered_stations
            )
            
            if busbar_options_filtered:
                all_busbar_option = "全部母线"
                busbar_options = [all_busbar_option] + busbar_options_filtered
                selected_busbar = st.sidebar.selectbox(
                    "选择母线",
                    options=busbar_options,
                    index=0,
                    help="选择要查看的母线"
                )
                
                if selected_busbar != all_busbar_option:
                    station_group_df = build_station_group_mapping(
                        station_info_df,
                        "母线",
                        station_names=filtered_stations
                    )
                    if station_group_df is not None:
                        busbar_stations_df = station_group_df[station_group_df["母线"] == selected_busbar]
                        busbar_stations_list = (
                            busbar_stations_df['电站名']
                            .astype(str)
                            .str.strip()
                            .tolist()
                        )
                        filtered_stations = [s for s in busbar_stations_list if s in filtered_stations]
                        st.sidebar.info(f" {selected_busbar} 共有 {len(filtered_stations)} 个站点")

        # 4. 最终站点选择
        if len(filtered_stations) == 0:
            st.warning("当前筛选条件下没有可用的电价数据文件！")
            return

        # 切换年份时保留已选站点
        prev_station = st.session_state.get('selected_station')
        if prev_station and prev_station in filtered_stations:
            default_station_index = filtered_stations.index(prev_station)
        else:
            default_station_index = 0

        selected_station = st.sidebar.selectbox(
            "选择站点",
            options=filtered_stations,
            index=default_station_index,
            help="选择要查看的站点"
        )
        st.session_state['selected_station'] = selected_station

        selected_file = price_files[selected_station]

        # 加载数据
        data_source = st.session_state.get('data_source', 'excel')
        
        with st.spinner("正在加载数据..."):
            if data_source == 'database':
                # 从数据库加载数据
                selected_year = st.session_state.get('selected_year')
                df = load_price_data_from_db(selected_station, selected_year)
            else:
                # 从Excel文件加载数据
                df = load_price_data(*get_file_cache_key(selected_file))

        if df is None:
            st.error("数据加载失败！")
            return

        # 计算全省平均电价（带磁盘缓存）
        if data_source == 'database':
            # 数据库模式：直接计算全省平均
            selected_year = st.session_state.get('selected_year')
            all_db_keys = [(station, selected_year) for station in price_files.keys()]
            prov_dates, prov_time_cols, prov_avg, failed_stations = compute_provincial_average_from_db(all_db_keys)
            
            # 显示失败站点提示和重新加载按钮
            if failed_stations:
                st.warning(f"以下 {len(failed_stations)} 个站点数据加载失败，已跳过: {', '.join(failed_stations[:5])}{'...' if len(failed_stations) > 5 else ''}")
                
                col_reload, col_info = st.columns([1, 3])
                with col_reload:
                    if st.button("🔄 重新加载失败站点", key="reload_failed_stations"):
                        # 清除相关缓存
                        compute_provincial_average_from_db.clear()
                        for station in failed_stations:
                            load_price_data_from_db.clear(station, selected_year)
                        st.rerun()
                with col_info:
                    st.info("点击按钮重新加载失败的站点数据")
        else:
            # Excel模式：使用原有缓存逻辑
            all_file_keys = [get_file_cache_key(fp) for fp in price_files.values()]
            prov_dates, prov_time_cols, prov_avg = compute_provincial_average_with_cache(tuple(all_file_keys))

        # 显示基本信息
        station_city = get_station_group_value(station_info_df, selected_station, '城市')
        station_title = f"{station_city} - {selected_station}" if station_city else selected_station
        st.header(f"📊 {station_title} - 电价数据概览")

        col1, col2, col3, col4 = st.columns(4)

        # 解析日期列
        date_col = df.columns[0]

        # 确保日期列是datetime类型
        try:
            df[date_col] = pd.to_datetime(df[date_col])
        except:
            pass

        # 计算统计信息
        price_cols = df.columns[1:]  # 除了日期列外的所有列都是电价数据
        all_prices = df[price_cols].values.flatten()

        with col1:
            st.metric("数据天数 (天)", f"{len(df)}")
        with col2:
            st.metric("最低电价 (元/kWh)", f"{all_prices.min():.4f}")
        with col3:
            st.metric("最高电价 (元/kWh)", f"{all_prices.max():.4f}")
        with col4:
            st.metric("平均电价 (元/kWh)", f"{all_prices.mean():.4f}")

        st.divider()

        # 按日期+时间节点查询电价
        st.subheader("🔍 按日期+时间节点查询电价")

        # 获取可用日期范围
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            available_dates = df[date_col].dt.date
        else:
            available_dates = pd.to_datetime(df[date_col]).dt.date
        min_date = available_dates.min()
        max_date = available_dates.max()

        col_date, col_time, col_info = st.columns([1, 1, 2])
        with col_date:
            query_date = st.date_input(
                "选择日期",
                value=None,
                min_value=min_date,
                max_value=max_date,
                help="选择要查询的日期"
            )
        with col_time:
            query_time = st.time_input(
                "选择时间节点",
                value=None,
                help="选择要查询的时间（自动对齐到最近的15分钟节点）"
            )

        if query_date is not None and query_time is not None:
            # 对齐到最近的15分钟节点
            hour = query_time.hour
            minute = query_time.minute
            snapped_minute = (minute // 15) * 15
            snapped_time_str = f"{hour:02d}:{snapped_minute:02d}"

            # 在时间列中查找
            time_columns = price_cols.tolist()
            if snapped_time_str in time_columns:
                time_col_idx = time_columns.index(snapped_time_str)

                # 查找对应日期的行
                if pd.api.types.is_datetime64_any_dtype(df[date_col]):
                    mask = df[date_col].dt.date == query_date
                else:
                    mask = pd.to_datetime(df[date_col]).dt.date == query_date

                matched_rows = df[mask]

                with col_info:
                    st.info(f"查询：**{query_date} {snapped_time_str}**（已自动对齐到15分钟节点）")

                if not matched_rows.empty:
                    target_price = float(pd.to_numeric(matched_rows.iloc[0][price_cols[time_col_idx]], errors='coerce'))

                    # 同时刻全年统计
                    all_slot_prices = pd.to_numeric(df[price_cols[time_col_idx]], errors='coerce').dropna()

                    col_a, col_b, col_c, col_d = st.columns(4)
                    with col_a:
                        st.metric(f"{query_date} {snapped_time_str} 电价", f"{target_price:.4f} 元/kWh")
                    with col_b:
                        st.metric("全年该时刻均价", f"{all_slot_prices.mean():.4f} 元/kWh")
                    with col_c:
                        st.metric("全年该时刻最高", f"{all_slot_prices.max():.4f} 元/kWh")
                    with col_d:
                        delta = target_price - all_slot_prices.mean()
                        st.metric("与全年均价差", f"{delta:+.4f} 元/kWh")
                else:
                    with col_info:
                        st.warning(f"未找到 **{query_date}** 的数据，请确认该日期有电价记录。")
            else:
                with col_info:
                    st.warning(f"未找到时刻 **{snapped_time_str}** 对应的数据列，请检查时间输入。")

        st.divider()

        # 日期选择器
        st.subheader("📅 选择日期查看详细数据")

        # 创建日期选项列表
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            date_options = df[date_col].dt.strftime('%Y-%m-%d').tolist()
        else:
            date_options = df[date_col].astype(str).tolist()

        selected_date_idx = st.selectbox(
            "选择日期",
            options=range(len(date_options)),
            format_func=lambda x: date_options[x],
            index=0
        )

        # 获取选中日期的数据
        selected_row = df.iloc[selected_date_idx]
        selected_date = date_options[selected_date_idx]

        # 提取96个时间节点的电价
        time_columns = price_cols.tolist()
        prices = selected_row[price_cols].values.astype(float)

        # 显示选中日期的电价表格
        st.subheader(f" {selected_date} 电价明细表")

        # 创建展示用的DataFrame
        display_df = pd.DataFrame({
            '时间节点': time_columns,
            '电价 (元/kWh)': [round(p, 4) for p in prices]
        })

        # 添加时段类型标注
        selected_date_value = selected_row[date_col]
        display_df['时段类型'] = display_df['时间节点'].apply(
            lambda time_str: get_guangdong_period_type(selected_date_value, time_str)
        )

        # 显示表格（不使用样式）
        st.dataframe(display_df, use_container_width=True, height=400, column_config={
            "时间节点": st.column_config.TextColumn("时间节点", width="small"),
            "电价 (元/kWh)": st.column_config.TextColumn("电价 (元/kWh)", width="medium"),
            "时段类型": st.column_config.TextColumn("时段类型", width="small")
        }, hide_index=True)

        # 全省平均电价明细表
        if prov_avg is not None and prov_dates is not None:
            prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
            if selected_date in prov_date_to_row:
                prov_prices = prov_avg[prov_date_to_row[selected_date]]
                st.markdown("---")
                st.subheader(" 全省平均电价明细表")
                prov_display_df = pd.DataFrame({
                    '时间节点': time_columns,
                    '全省平均电价 (元/kWh)': np.round(prov_prices, 4)
                })
                prov_display_df['时段类型'] = prov_display_df['时间节点'].apply(
                    lambda time_str: get_guangdong_period_type(selected_date_value, time_str)
                )
                st.dataframe(prov_display_df, use_container_width=True, height=400, column_config={
                    "时间节点": st.column_config.TextColumn("时间节点", width="small"),
                    "全省平均电价 (元/kWh)": st.column_config.TextColumn("全省平均电价 (元/kWh)", width="medium"),
                    "时段类型": st.column_config.TextColumn("时段类型", width="small")
                }, hide_index=True)

        st.divider()

        # 可视化 - 电价曲线图
        st.subheader(" 电价曲线图")
        
        # 创建左右两列布局
        chart_col, metrics_col = st.columns([3, 1])
        
        with chart_col:
            # 创建Plotly图表
            fig = go.Figure()

            # 添加电价曲线
            fig.add_trace(go.Scatter(
                x=time_columns,
                y=prices,
                mode='lines+markers',
                name='电价',
                line=dict(color='#FF6B35', width=2),
                marker=dict(size=4)
            ))

            # 添加填充区域
            fig.add_trace(go.Scatter(
                x=time_columns,
                y=prices,
                mode='none',
                fill='tozeroy',
                fillcolor='rgba(255, 107, 53, 0.1)',
                name='电价区域'
            ))

            # 添加全省平均电价曲线
            if prov_avg is not None and prov_dates is not None:
                prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
                if selected_date in prov_date_to_row:
                    prov_prices = prov_avg[prov_date_to_row[selected_date]]
                    fig.add_trace(go.Scatter(
                        x=time_columns,
                        y=prov_prices,
                        mode='lines',
                        name='全省平均电价',
                        line=dict(color='blue', width=2, dash='dash')
                    ))

            # 更新布局
            fig.update_layout(
                title=f'{selected_date} 电价变化曲线（{selected_station} vs 全省平均）',
                xaxis_title='时间节点',
                yaxis_title='电价 (元/kWh)',
                hovermode='x unified',
                template='plotly_white',
                height=500,
                xaxis=dict(
                    tickangle=45,
                    tickvals=time_columns[::4],  # 每4个点显示一个标签
                    ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                )
            )

            st.plotly_chart(fig, use_container_width=True)
        
        with metrics_col:
            # 计算并显示站点电价与全省平均电价的RMSE
            if prov_avg is not None and prov_dates is not None:
                prov_date_to_row = {d: i for i, d in enumerate(prov_dates)}
                if selected_date in prov_date_to_row:
                    prov_prices = prov_avg[prov_date_to_row[selected_date]]
                    
                    # 计算RMSE（均方根误差）
                    price_diff = prices - prov_prices
                    rmse = np.sqrt(np.mean(price_diff ** 2))
                    
                    st.subheader("📊 电价差异分析")
                    st.divider()
                    
                    st.metric(
                        "RMSE (均方根误差)",
                        f"{rmse:.4f}",
                        help="站点电价与全省平均电价的均方根误差，越小表示越接近全省均价"
                    )
                    
                    # 计算其他统计指标
                    mae = np.mean(np.abs(price_diff))
                    st.metric(
                        "MAE (平均标准差)",
                        f"{mae:.4f}",
                        help="站点电价与全省平均电价的平均绝对误差"
                    )
                    
                    max_diff_idx = np.argmax(np.abs(price_diff))
                    st.metric(
                        "最大差异节点",
                        time_columns[max_diff_idx],
                        help=f"差异最大的时间节点（差值：{price_diff[max_diff_idx]:.4f} 元/kWh）"
                    )
                    
                    # 添加额外信息
                    st.divider()
                    st.markdown(f"**时间节点数：** {len(prices)}")
                    st.markdown(f"**最小差值：** {np.min(price_diff):.4f}")
                    st.markdown(f"**最大差值：** {np.max(price_diff):.4f}")

        # 时间段电价分析曲线（用户可选日期范围，支持跨年）
        st.divider()
        st.subheader("📈 时间段电价分析曲线")

        # 加载该站点所有年份数据
        if data_source == 'database':
            df_all_years = load_station_all_years_data(selected_station, 'database')
        else:
            price_dir_labels_key = tuple(label for label, _ in get_price_data_dir_options())
            df_all_years = load_station_all_years_data(selected_station, 'excel', _price_dir_labels=price_dir_labels_key)

        if df_all_years is None:
            df_all_years = df

        date_col_ay = df_all_years.columns[0]
        price_cols_ay = df_all_years.columns[1:]
        if not pd.api.types.is_datetime64_any_dtype(df_all_years[date_col_ay]):
            df_all_years[date_col_ay] = pd.to_datetime(df_all_years[date_col_ay], errors='coerce')

        if len(price_cols_ay) >= 96:
            price_cols_ay = price_cols_ay[:96]

        # 生成时间列（96个时段）
        time_columns_ay = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]
        if len(price_cols_ay) == 96:
            time_columns_chart = time_columns_ay
        else:
            time_columns_chart = list(price_cols_ay)

        if pd.api.types.is_datetime64_any_dtype(df_all_years[date_col_ay]):
            all_dates_analysis = df_all_years[date_col_ay].dt.date
            min_date_a = all_dates_analysis.min()
            max_date_a = all_dates_analysis.max()

            date_col_a1, date_col_a2 = st.columns(2)
            with date_col_a1:
                analysis_start_date = st.date_input(
                    "开始日期",
                    value=min_date_a,
                    min_value=min_date_a,
                    max_value=max_date_a,
                    key="analysis_start_date"
                )
            with date_col_a2:
                analysis_end_date = st.date_input(
                    "结束日期",
                    value=max_date_a,
                    min_value=min_date_a,
                    max_value=max_date_a,
                    key="analysis_end_date"
                )

            # 筛选日期范围
            analysis_mask = (all_dates_analysis >= analysis_start_date) & (all_dates_analysis <= analysis_end_date)
            df_analysis = df_all_years[analysis_mask]

            if df_analysis.empty:
                st.warning("所选日期范围内没有数据！")
            else:
                st.caption(f"📅 已筛选 {analysis_start_date} 至 {analysis_end_date}，共 {len(df_analysis)} 天数据")

                # 计算各时段跨天平均电价
                analysis_avg_prices = df_analysis[price_cols_ay].mean(axis=0).values.astype(float)

                # 全省均价（按相同日期范围过滤）
                analysis_prov_avg = None
                if prov_avg is not None and prov_dates is not None:
                    prov_date_to_row_a = {d: i for i, d in enumerate(prov_dates)}
                    filtered_prov_indices = [
                        prov_date_to_row_a[d] for d in prov_dates
                        if analysis_start_date <= pd.Timestamp(d).date() <= analysis_end_date
                    ]
                    if filtered_prov_indices:
                        analysis_prov_avg = np.mean(prov_avg[filtered_prov_indices], axis=0)

                # 左右两列布局
                analysis_chart_col, analysis_metrics_col = st.columns([3, 1])

                with analysis_chart_col:
                    fig_analysis = go.Figure()

                    # 站点平均电价曲线
                    fig_analysis.add_trace(go.Scatter(
                        x=time_columns_chart,
                        y=analysis_avg_prices,
                        mode='lines+markers',
                        name='站点平均电价',
                        line=dict(color='#FF6B35', width=2),
                        marker=dict(size=4)
                    ))

                    # 填充区域
                    fig_analysis.add_trace(go.Scatter(
                        x=time_columns_chart,
                        y=analysis_avg_prices,
                        mode='none',
                        fill='tozeroy',
                        fillcolor='rgba(255, 107, 53, 0.1)',
                        name='电价区域'
                    ))

                    # 全省平均电价曲线
                    if analysis_prov_avg is not None:
                        fig_analysis.add_trace(go.Scatter(
                            x=time_columns_chart,
                            y=analysis_prov_avg,
                            mode='lines',
                            name='全省平均电价',
                            line=dict(color='blue', width=2, dash='dash')
                        ))

                    fig_analysis.update_layout(
                        title=f'{analysis_start_date} 至 {analysis_end_date} 电价分析曲线（{selected_station} vs 全省平均）',
                        xaxis_title='时间节点',
                        yaxis_title='电价 (元/kWh)',
                        hovermode='x unified',
                        template='plotly_white',
                        height=500,
                        xaxis=dict(
                            tickangle=45,
                            tickvals=time_columns_chart[::4],
                            ticktext=[time_columns_chart[i] for i in range(0, len(time_columns_chart), 4)]
                        )
                    )

                    st.plotly_chart(fig_analysis, use_container_width=True)

                with analysis_metrics_col:
                    if analysis_prov_avg is not None:
                        analysis_diff = analysis_avg_prices - analysis_prov_avg
                        analysis_rmse = np.sqrt(np.mean(analysis_diff ** 2))

                        st.subheader("📊 电价差异分析")
                        st.divider()

                        st.metric(
                            "RMSE (均方根误差)",
                            f"{analysis_rmse:.4f}",
                            help="站点电价与全省平均电价的均方根误差，越小表示越接近全省均价"
                        )

                        analysis_mae = np.mean(np.abs(analysis_diff))
                        st.metric(
                            "MAE (平均标准差)",
                            f"{analysis_mae:.4f}",
                            help="站点电价与全省平均电价的平均绝对误差"
                        )

                        analysis_max_diff_idx = int(np.argmax(np.abs(analysis_diff)))
                        st.metric(
                            "最大差异节点",
                            time_columns_chart[analysis_max_diff_idx] if analysis_max_diff_idx < len(time_columns_chart) else "N/A",
                            help=f"差异最大的时间节点（差值：{analysis_diff[analysis_max_diff_idx]:.4f} 元/kWh）"
                        )

                        st.divider()
                        st.markdown(f"**时间节点数：** {len(analysis_avg_prices)}")
                        st.markdown(f"**最小差值：** {np.min(analysis_diff):.4f}")
                        st.markdown(f"**最大差值：** {np.max(analysis_diff):.4f}")

                # 各时间节点电价均价明细表
                st.divider()
                st.subheader("📋 各时间节点电价均价明细表")

                detail_df = pd.DataFrame({
                    '时间节点': time_columns_chart,
                    '站点平均电价 (元/kWh)': np.round(analysis_avg_prices, 4),
                })

                if analysis_prov_avg is not None:
                    detail_df['全省平均电价 (元/kWh)'] = np.round(analysis_prov_avg, 4)
                    detail_df['差值 (元/kWh)'] = np.round(analysis_avg_prices - analysis_prov_avg, 4)

                st.dataframe(detail_df, use_container_width=True, hide_index=True)

        st.divider()

        # 价差规律分析
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            st.subheader("🔍 价差规律分析")
            
            # 用户选择峰谷时段数
            n_peak = st.slider(
                "选择峰谷时段数",
                min_value=1,
                max_value=24,
                value=8,
                help="取电价最高N个时段和最低N个时段的平均值之差作为日均峰谷价差"
            )
            
            # 计算日均峰谷价差
            df_price_spread = df.copy()
            df_price_spread['日期'] = df_price_spread[date_col]
            
            # 时间段筛选
            all_dates = df_price_spread['日期'].dt.date
            min_date = all_dates.min()
            max_date = all_dates.max()
            
            date_col1, date_col2 = st.columns(2)
            with date_col1:
                start_date = st.date_input(
                    "开始日期",
                    value=min_date,
                    min_value=min_date,
                    max_value=max_date,
                    key="spread_start_date"
                )
            with date_col2:
                end_date = st.date_input(
                    "结束日期",
                    value=max_date,
                    min_value=min_date,
                    max_value=max_date,
                    key="spread_end_date"
                )
            
            # 按日期范围筛选
            mask = (df_price_spread['日期'].dt.date >= start_date) & (df_price_spread['日期'].dt.date <= end_date)
            df_price_spread = df_price_spread[mask].copy()
            
            if df_price_spread.empty:
                st.warning("所选日期范围内没有数据！")
                return
            
            st.info(f"📅 已筛选 {start_date} 至 {end_date}，共 {len(df_price_spread)} 天数据")
            
            # 计算每日峰谷价差（滑动窗口：连续N个时段均值最高 - 连续N个时段均值最低）
            def _calc_spread(row):
                p, v, s = _sliding_window_spread(row.values, n_peak)
                return pd.Series({'peak': p, 'valley': v, 'spread': s})

            spread_df = df_price_spread[price_cols].apply(_calc_spread, axis=1)
            df_price_spread['峰值均价'] = spread_df['peak']
            df_price_spread['谷值均价'] = spread_df['valley']
            df_price_spread['日均峰谷价差'] = spread_df['spread']
            df_price_spread['日均电价'] = df_price_spread[price_cols].mean(axis=1)
            df_price_spread['月份'] = df_price_spread['日期'].dt.month
            df_price_spread['季度'] = df_price_spread['日期'].dt.quarter
            df_price_spread['星期'] = df_price_spread['日期'].dt.dayofweek  # 0=周一, 6=周日
            df_price_spread['是否周末'] = df_price_spread['星期'].apply(lambda x: '周末' if x >= 5 else '工作日')
            
            # 创建Tabs
            tab1, tab2, tab3, tab4 = st.tabs(["日内价差规律", "工作日/周末对比", "月度价差趋势", "季度价差对比"])
            
            with tab1:
                st.markdown(f"### 📊 日均峰谷价差分布分析（统取{n_peak}个时段）")
                st.info(f"💡 **日均峰谷价差** = 每日电价最高的{n_peak}个时段均价 - 每日电价最低的{n_peak}个时段均价")
                
                # 获取日均峰谷价差数据
                daily_spread = df_price_spread['日均峰谷价差']
                
                # 统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("平均日均峰谷价差", f"{daily_spread.mean():.4f} 元/kWh")
                with col2:
                    st.metric("最大日均峰谷价差", f"{daily_spread.max():.4f} 元/kWh")
                with col3:
                    st.metric("最小日均峰谷价差", f"{daily_spread.min():.4f} 元/kWh")
                with col4:
                    st.metric("价差标准差", f"{daily_spread.std():.4f}")
                
                # 绘制日均峰谷价差分布直方图
                fig_hist = go.Figure()
                
                fig_hist.add_trace(go.Histogram(
                    x=daily_spread,
                    nbinsx=50,
                    name='日均峰谷价差分布',
                    marker_color='#FF6B35',
                    opacity=0.7
                ))
                
                # 添加平均线
                fig_hist.add_vline(
                    x=daily_spread.mean(),
                    line_dash="dash",
                    line_color="red",
                    annotation_text=f"平均值: {daily_spread.mean():.4f}",
                    annotation_position="top right"
                )
                
                fig_hist.update_layout(
                    title='日均峰谷价差分布直方图',
                    xaxis_title='日均峰谷价差 (元/kWh)',
                    yaxis_title='天数',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_hist, use_container_width=True)
                
                # 绘制箱线图
                st.markdown("### 📦 日均峰谷价差箱线图")
                
                fig_box = go.Figure()
                
                fig_box.add_trace(go.Box(
                    y=daily_spread,
                    name='日均峰谷价差',
                    marker_color='#FF6B35',
                    boxmean=True,
                    boxpoints='outliers'
                ))
                
                fig_box.update_layout(
                    title='日均峰谷价差箱线图（显示异常值）',
                    yaxis_title='日均峰谷价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_box, use_container_width=True)
                
                # 分位数分析
                st.markdown("### 📊 价差分位数统计")
                
                quantiles = [0.05, 0.10, 0.25, 0.50, 0.75, 0.90, 0.95]
                quantile_values = daily_spread.quantile(quantiles)
                
                quantile_df = pd.DataFrame({
                    '分位数': [f"{int(q*100)}%" for q in quantiles],
                    '价差值 (元/kWh)': quantile_values.round(4).values,
                    '说明': [
                        '仅5%的天数价差低于此值',
                        '仅10%的天数价差低于此值',
                        '25%的天数价差低于此值（下四分位数）',
                        '中位数（50%的天数价差低于此值）',
                        '75%的天数价差低于此值（上四分位数）',
                        '90%的天数价差低于此值',
                        '仅5%的天数价差高于此值'
                    ]
                })
                
                st.dataframe(quantile_df, use_container_width=True)
                
                # 关键发现
                st.markdown("### 🔍 关键发现")
                
                iqr = quantile_values[0.75] - quantile_values[0.25]
                lower_bound = quantile_values[0.25] - 1.5 * iqr
                upper_bound = quantile_values[0.75] + 1.5 * iqr
                
                outlier_count = ((daily_spread < lower_bound) | (daily_spread > upper_bound)).sum()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.success(f" **中位数价差：** {quantile_values[0.50]:.4f} 元/kWh")
                    st.info(f" **75%的天数价差低于：** {quantile_values[0.75]:.4f} 元/kWh")
                with col2:
                    st.warning(f" **异常值天数：** {outlier_count} 天（占总天数的 {outlier_count/len(daily_spread)*100:.1f}%）")
                    st.markdown(f"- **正常价差范围：** {lower_bound:.4f} ~ {upper_bound:.4f} 元/kWh")
            
            with tab2:
                st.markdown("### 📊 工作日 vs 周末价差对比")
                
                # 按工作日/周末分组统计
                weekday_stats = df_price_spread.groupby('是否周末').agg({
                    '日均峰谷价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                weekday_stats.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                weekday_stats = weekday_stats.reset_index()
                
                # 显示统计表格
                st.dataframe(weekday_stats, use_container_width=True)
                
                # 绘制对比图
                fig_weekday = go.Figure()
                
                fig_weekday.add_trace(go.Bar(
                    x=weekday_stats['是否周末'],
                    y=weekday_stats['平均价差'],
                    name='平均价差',
                    marker_color=['#2196F3', '#FF6B35'],
                    text=weekday_stats['平均价差'].round(4),
                    textposition='auto'
                ))
                
                fig_weekday.update_layout(
                    title='工作日与周末平均价差对比',
                    xaxis_title='日期类型',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_weekday, use_container_width=True)
            
            with tab3:
                st.markdown("### 📊 月度价差趋势")
                
                # 按月份统计价差
                monthly_spread = df_price_spread.groupby('月份').agg({
                    '日均峰谷价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                monthly_spread.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                monthly_spread = monthly_spread.reset_index()
                
                # 显示统计表格
                st.dataframe(monthly_spread, use_container_width=True)
                
                # 绘制月度趋势图
                fig_monthly = go.Figure()
                
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'],
                    mode='lines+markers',
                    name='平均价差',
                    line=dict(color='#FF6B35', width=2),
                    marker=dict(size=8)
                ))
                
                # 添加误差线（标准差）
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'] + monthly_spread['价差标准差'],
                    mode='lines',
                    name='+标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    showlegend=False
                ))
                
                fig_monthly.add_trace(go.Scatter(
                    x=monthly_spread['月份'],
                    y=monthly_spread['平均价差'] - monthly_spread['价差标准差'],
                    mode='lines',
                    name='-标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    fill='tonexty',
                    fillcolor='rgba(255, 107, 53, 0.1)',
                    showlegend=False
                ))
                
                fig_monthly.update_layout(
                    title='月度平均价差趋势',
                    xaxis_title='月份',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(tickmode='linear', dtick=1)
                )
                
                st.plotly_chart(fig_monthly, use_container_width=True)
            
            with tab4:
                st.markdown("### 📊 季度价差对比")
                
                # 按季度统计价差
                quarterly_spread = df_price_spread.groupby('季度').agg({
                    '日均峰谷价差': ['mean', 'std', 'min', 'max', 'count']
                }).round(4)
                quarterly_spread.columns = ['平均价差', '价差标准差', '最小价差', '最大价差', '天数']
                quarterly_spread = quarterly_spread.reset_index()
                quarterly_spread['季度名称'] = quarterly_spread['季度'].apply(lambda x: f'Q{x}')
                
                # 显示统计表格
                st.dataframe(quarterly_spread[['季度名称', '平均价差', '价差标准差', '最小价差', '最大价差', '天数']], use_container_width=True)
                
                # 绘制季度对比图
                fig_quarterly = go.Figure()
                
                fig_quarterly.add_trace(go.Bar(
                    x=quarterly_spread['季度名称'],
                    y=quarterly_spread['平均价差'],
                    name='平均价差',
                    marker_color=['#FF6B35', '#2196F3', '#4CAF50', '#9C27B0'],
                    text=quarterly_spread['平均价差'].round(4),
                    textposition='auto'
                ))
                
                fig_quarterly.update_layout(
                    title='季度平均价差对比',
                    xaxis_title='季度',
                    yaxis_title='平均价差 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_quarterly, use_container_width=True)
        
        st.divider()

        # 时段特征及波动趋势分析
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            st.subheader("⚡ 时段特征及波动趋势")
            
            # 为每个时间节点添加时段类型标签
            df_time_features = df.copy()
            df_time_features['日期'] = df_time_features[date_col]
            
            # 计算每个时间节点的统计特征
            time_stats_list = []
            for time_col in time_columns:
                time_data = df_time_features[time_col].dropna()
                time_stats = {
                    '时间节点': time_col,
                    '平均电价': time_data.mean(),
                    '电价中位数': time_data.median(),
                    '电价标准差': time_data.std(),
                    '电价变异系数': time_data.std() / time_data.mean() if time_data.mean() != 0 else 0,
                    '最高电价': time_data.max(),
                    '最低电价': time_data.min(),
                    '电价极差': time_data.max() - time_data.min(),
                    '数据天数': len(time_data)
                }
                time_stats_list.append(time_stats)
            
            time_stats_df = pd.DataFrame(time_stats_list).round(4)
            
            # 添加时段类型
            # 假设第一个日期来判定时段类型
            sample_date = df_time_features[date_col].iloc[0]
            time_stats_df['时段类型'] = time_stats_df['时间节点'].apply(
                lambda t: get_guangdong_period_type(sample_date, t)
            )
            
            # 创建Tabs
            tab1, tab2, tab3 = st.tabs(["峰平谷时段特征", "电价波动趋势", "波动特征统计"])
            
            with tab1:
                st.markdown("###  峰、平、谷时段电价特征")
                st.info("💡 **电价变异系数** = 电价标准差 / 电价平均值，衡量电价的相对波动程度（值越大波动越剧烈）")
                
                # 按时段类型分组统计
                period_stats = time_stats_df.groupby('时段类型').agg({
                    '平均电价': 'mean',
                    '电价标准差': 'mean',
                    '电价变异系数': 'mean',
                    '最高电价': 'max',
                    '最低电价': 'min',
                    '电价极差': 'mean'
                }).round(4).reset_index()
                
                # 显示统计表格
                st.dataframe(period_stats, use_container_width=True)
                
                # 绘制峰平谷对比图
                fig_period = go.Figure()
                
                # 按顺序排列时段类型（与广东省分时电价规则一致）
                period_order = ['尖峰', '高峰', '平段', '低谷']
                period_stats['排序'] = period_stats['时段类型'].apply(
                    lambda x: period_order.index(x) if x in period_order else len(period_order)
                )
                period_stats = period_stats.sort_values('排序')
                
                colors = {'尖峰': '#FF0000', '高峰': '#FF6B35', '平段': '#FFC107', '低谷': '#2196F3'}
                
                fig_period.add_trace(go.Bar(
                    x=period_stats['时段类型'],
                    y=period_stats['平均电价'],
                    name='平均电价',
                    marker_color=[colors.get(p, '#999999') for p in period_stats['时段类型']],
                    text=period_stats['平均电价'].round(4),
                    textposition='auto'
                ))
                
                fig_period.update_layout(
                    title='各时段类型平均电价对比',
                    xaxis_title='时段类型',
                    yaxis_title='平均电价 (元/kWh)',
                    template='plotly_white',
                    height=400
                )
                
                st.plotly_chart(fig_period, use_container_width=True)
                
                # 显示关键发现
                col1, col2, col3 = st.columns(3)
                with col1:
                    max_period = period_stats.loc[period_stats['平均电价'].idxmax()]
                    st.success(f" **最高均价时段：** {max_period['时段类型']} ({max_period['平均电价']:.4f} 元/kWh)")
                with col2:
                    min_period = period_stats.loc[period_stats['平均电价'].idxmin()]
                    st.info(f"🟢 **最低均价时段：** {min_period['时段类型']} ({min_period['平均电价']:.4f} 元/kWh)")
                with col3:
                    max_spread_period = period_stats.loc[period_stats['电价极差'].idxmax()]
                    st.warning(f"🟡 **波动最大时段：** {max_spread_period['时段类型']} (极差：{max_spread_period['电价极差']:.4f})")
            
            with tab2:
                st.markdown("### 📊 各时间节点电价波动趋势")
                
                # 绘制每个时间节点的均价和标准差
                fig_trend = go.Figure()
                
                # 均价曲线
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'],
                    mode='lines+markers',
                    name='平均电价',
                    line=dict(color='#FF6B35', width=2),
                    marker=dict(size=6)
                ))
                
                # 标准差带
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'] + time_stats_df['电价标准差'],
                    mode='lines',
                    name='+标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    showlegend=False
                ))
                
                fig_trend.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['平均电价'] - time_stats_df['电价标准差'],
                    mode='lines',
                    name='-标准差',
                    line=dict(color='rgba(255, 107, 53, 0.3)', width=1, dash='dash'),
                    fill='tonexty',
                    fillcolor='rgba(255, 107, 53, 0.1)',
                    showlegend=False
                ))
                
                fig_trend.update_layout(
                    title='各时间节点平均电价及波动范围',
                    xaxis_title='时间节点',
                    yaxis_title='电价 (元/kWh)',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(
                        tickangle=45,
                        tickvals=time_columns[::4],
                        ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                    )
                )
                
                st.plotly_chart(fig_trend, use_container_width=True)
                
                # 变异系数趋势图
                fig_cv = go.Figure()
                
                fig_cv.add_trace(go.Scatter(
                    x=time_stats_df['时间节点'],
                    y=time_stats_df['电价变异系数'],
                    mode='lines+markers',
                    name='变异系数',
                    line=dict(color='#2196F3', width=2),
                    marker=dict(size=6)
                ))
                
                fig_cv.update_layout(
                    title='各时间节点电价变异系数（相对波动程度）',
                    xaxis_title='时间节点',
                    yaxis_title='变异系数',
                    template='plotly_white',
                    height=400,
                    xaxis=dict(
                        tickangle=45,
                        tickvals=time_columns[::4],
                        ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                    )
                )
                
                st.plotly_chart(fig_cv, use_container_width=True)
            
            with tab3:
                st.markdown("###  波动特征统计")
                
                # 显示详细的时间节点统计数据
                st.dataframe(time_stats_df, use_container_width=True, height=400)
                
                # 找出波动最大和最小的时间段
                col1, col2 = st.columns(2)
                with col1:
                    max_std_row = time_stats_df.loc[time_stats_df['电价标准差'].idxmax()]
                    st.warning(f" **波动最大的时间节点：** {max_std_row['时间节点']}")
                    st.markdown(f"- 平均电价：{max_std_row['平均电价']:.4f} 元/kWh")
                    st.markdown(f"- 标准差：{max_std_row['电价标准差']:.4f}")
                    st.markdown(f"- 变异系数：{max_std_row['电价变异系数']:.4f}")
                    st.markdown(f"- 极差：{max_std_row['电价极差']:.4f}")
                
                with col2:
                    min_std_row = time_stats_df.loc[time_stats_df['电价标准差'].idxmin()]
                    st.success(f" **波动最小的时间节点：** {min_std_row['时间节点']}")
                    st.markdown(f"- 平均电价：{min_std_row['平均电价']:.4f} 元/kWh")
                    st.markdown(f"- 标准差：{min_std_row['电价标准差']:.4f}")
                    st.markdown(f"- 变异系数：{min_std_row['电价变异系数']:.4f}")
                    st.markdown(f"- 极差：{min_std_row['电价极差']:.4f}")
        
        st.divider()

        # 年度统计
        st.subheader(" 年度统计分析")

        # 计算每天的平均电价
        daily_avg_prices = df[price_cols].mean(axis=1)

        # 添加日期列
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            analysis_df = pd.DataFrame({
                '日期': df[date_col].dt.strftime('%Y-%m-%d'),
                '日均电价': daily_avg_prices.values
            })
        else:
            analysis_df = pd.DataFrame({
                '日期': df[date_col].astype(str),
                '日均电价': daily_avg_prices.values
            })

        # 绘制年度电价趋势图
        fig_trend = go.Figure()

        fig_trend.add_trace(go.Scatter(
            x=analysis_df['日期'],
            y=analysis_df['日均电价'],
            mode='lines',
            name=f'{selected_station} 日均电价',
            line=dict(color='#2196F3', width=1.5)
        ))

        # 添加全省平均日均电价曲线
        if prov_avg is not None and prov_dates is not None:
            prov_daily_avg = prov_avg.mean(axis=1)
            fig_trend.add_trace(go.Scatter(
                x=prov_dates,
                y=prov_daily_avg,
                mode='lines',
                name='全省平均日均电价',
                line=dict(color='red', width=1.5, dash='dash')
            ))

        fig_trend.update_layout(
            title=f'{selected_station} vs 全省平均 日均电价趋势对比',
            xaxis_title='日期',
            yaxis_title='日均电价 (元/kWh)',
            template='plotly_white',
            height=400,
            xaxis=dict(
                tickangle=45,
                tickvals=analysis_df['日期'][::30],  # 每30天显示一个标签
                ticktext=[analysis_df['日期'].iloc[i] for i in range(0, len(analysis_df), 30)]
            )
        )

        st.plotly_chart(fig_trend, use_container_width=True)

        # 年度电价差异分析指标
        if prov_avg is not None and prov_dates is not None:
            st.subheader("📊 年度电价差异分析")
            
            # 计算全省日均电价
            prov_daily_avg = prov_avg.mean(axis=1)
            prov_date_to_avg = {d: avg for d, avg in zip(prov_dates, prov_daily_avg)}
            
            # 匹配日期计算差异
            station_dates = analysis_df['日期'].tolist()
            station_daily_avg = analysis_df['日均电价'].values
            
            all_diffs = []
            all_diffs_squared = []
            for date_str, station_avg in zip(station_dates, station_daily_avg):
                if date_str in prov_date_to_avg:
                    prov_avg_val = prov_date_to_avg[date_str]
                    diff = station_avg - prov_avg_val
                    all_diffs.append(abs(diff))
                    all_diffs_squared.append(diff ** 2)
            
            if all_diffs:
                all_diffs = np.array(all_diffs)
                all_diffs_squared = np.array(all_diffs_squared)
                
                # 计算RMSE（均方根误差）
                rmse = np.sqrt(np.mean(all_diffs_squared))
                # 计算MAE（平均绝对误差）
                mae = np.mean(all_diffs)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric(
                        "RMSE (均方根误差)",
                        f"{rmse:.4f}",
                        help="站点电价与全省平均电价的均方根误差，越小表示越接近全省均价"
                    )
                with col2:
                    st.metric(
                        "MAE (平均标准差)",
                        f"{mae:.4f}",
                        help="站点电价与全省平均电价的平均绝对误差"
                    )
                with col3:
                    max_diff = np.max(all_diffs)
                    st.metric(
                        "最大日差异",
                        f"{max_diff:.4f}",
                        help="全年中站点电价与全省均价差异最大的一天"
                    )
                with col4:
                    min_diff = np.min(all_diffs)
                    st.metric(
                        "最小日差异",
                        f"{min_diff:.4f}",
                        help="全年中站点电价与全省均价差异最小的一天"
                    )
                
                # 差异分布直方图
                fig_diff_hist = go.Figure()
                fig_diff_hist.add_trace(go.Histogram(
                    x=all_diffs,
                    nbinsx=50,
                    name='差异分布',
                    marker_color='#FF6B35',
                    opacity=0.7
                ))
                fig_diff_hist.add_vline(
                    x=mae, line_dash="dash", line_color="red",
                    annotation_text=f"MAE: {mae:.4f}"
                )
                fig_diff_hist.update_layout(
                    title='站点电价与全省均价差异分布',
                    xaxis_title='绝对差异 (元/kWh)',
                    yaxis_title='天数',
                    template='plotly_white',
                    height=300
                )
                st.plotly_chart(fig_diff_hist, use_container_width=True)

        # 月度统计
        st.subheader("📅 月度统计摘要")

        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            # 提取月份
            df_copy = df.copy()
            df_copy['月份'] = df_copy[date_col].dt.to_period('M').astype(str)

            # 计算每天的平均电价
            df_copy['日均电价'] = df_copy[price_cols].mean(axis=1)

            # 按月统计
            monthly_stats = df_copy.groupby('月份')['日均电价'].agg(['mean', 'min', 'max', 'std'])
            monthly_stats.columns = ['月均电价', '月最低电价', '月最高电价', '电价标准差']
            monthly_stats = monthly_stats.round(4)

            st.dataframe(monthly_stats, use_container_width=True)
        else:
            st.info("无法进行月度统计（日期格式不支持）")

        st.divider()

        # 数据导出
        st.subheader("💾 数据导出")

        export_option = st.radio(
            "选择导出内容",
            ["单日详细数据", "全年完整数据"],
            horizontal=True
        )

        if export_option == "单日详细数据":
            # 重新创建单日数据
            single_day_df = pd.DataFrame({
                '时间节点': time_columns,
                '电价 (元/kWh)': prices
            })
            single_day_df.insert(0, '日期', selected_date)
            csv_data = single_day_df.to_csv(index=False, encoding='utf-8-sig')
            file_name = f"{selected_station}_{selected_date}_电价数据.csv"
        else:
            export_df = df.copy()
            csv_data = export_df.to_csv(index=False, encoding='utf-8-sig')
            file_name = f"{selected_station}_全年电价数据.csv"

        st.download_button(
            label="📥 下载CSV文件",
            data=csv_data,
            file_name=file_name,
            mime="text/csv"
        )

    # 电价差排名页
    elif view_mode == "📈 电价差排名":
        # 侧边栏：全局筛选
        selected_city_filter = None
        if city_types:
            st.sidebar.divider()
            st.sidebar.header("🌍 全局筛选")
            all_city_option = "全部城市"
            selected_city_filter = st.sidebar.selectbox(
                "选择城市",
                options=[all_city_option] + city_types,
                index=0,
                help="选择城市后，所有排名都将限制在该城市范围内"
            )
            if selected_city_filter == all_city_option:
                selected_city_filter = None
        
        # 如果选择了城市，先过滤站点
        filtered_price_files = price_files
        if selected_city_filter:
            city_station_df = build_station_group_mapping(
                station_info_df,
                '城市',
                station_names=list(price_files.keys())
            )
            if city_station_df is not None:
                city_stations = city_station_df[
                    city_station_df['城市'] == selected_city_filter
                ]['电站名'].astype(str).str.strip().tolist()
                filtered_price_files = {k: v for k, v in price_files.items() if k in city_stations}
        
        # 根据数据源选择计算方式
        data_source = st.session_state.get('data_source', 'excel')
        
        if data_source == 'database':
            # 数据库模式
            selected_year = st.session_state.get('selected_year')
            station_year_pairs = [(station, selected_year) for station in filtered_price_files.keys()]
            st.caption("从数据库计算电价差排名。")
            
            if selected_city_filter:
                st.caption(f"🌍 当前筛选城市：{selected_city_filter}")

            with st.spinner("正在计算所有站点电价差..."):
                all_stations_stats, failed_stations, cache_summary = calculate_all_stations_price_spread_from_db(station_year_pairs)

            st.caption(
                f"本次从数据库计算 {cache_summary['recomputed_count']} 个站点。"
            )
            
            # 显示失败站点提示和重新加载按钮
            if failed_stations:
                st.warning(f"有 {len(failed_stations)} 个站点统计失败，已自动跳过: {', '.join([s[0] if isinstance(s, tuple) else s for s in failed_stations[:5]])}{'...' if len(failed_stations) > 5 else ''}")
                
                col_reload, col_info = st.columns([1, 3])
                with col_reload:
                    if st.button("🔄 重新加载失败站点", key="reload_failed_rankings"):
                        # 清除相关缓存
                        calculate_all_stations_price_spread_from_db.clear()
                        for station_info in failed_stations:
                            station_name = station_info[0] if isinstance(station_info, tuple) else station_info
                            load_price_data_from_db.clear(station_name, selected_year)
                        st.rerun()
                with col_info:
                    st.info("点击按钮重新加载失败的站点数据")
        else:
            # Excel模式
            price_file_index = build_price_file_index(filtered_price_files)
            st.caption("排名结果会自动缓存到本地，应用重启后也能复用；只有变更过的 Excel 才会重算。")
            
            if selected_city_filter:
                st.caption(f"🌍 当前筛选城市：{selected_city_filter}")

            with st.spinner("正在计算所有站点电价差..."):
                all_stations_stats, failed_stations, cache_summary = calculate_all_stations_price_spread(price_file_index)

            st.caption(
                f"本次命中本地缓存 {cache_summary['cached_count']} 个文件，"
                f"重新计算 {cache_summary['recomputed_count']} 个文件。"
            )
            
            # Excel模式下显示失败站点提示
            if failed_stations:
                st.warning(f"有 {len(failed_stations)} 个站点文件统计失败，已自动跳过。")

        if len(all_stations_stats) > 0:
            # 为所有统计数据添加厂站类型信息
            all_stations_with_factory = prepare_grouped_rankings(
                all_stations_stats, station_info_df, FACTORY_GROUP_COLUMN
            )
            
            grouped_ranking_configs = {}

            if busbar_types:
                grouped_busbar_df = prepare_grouped_rankings(all_stations_stats, station_info_df, '母线')
                if grouped_busbar_df is not None:
                    grouped_ranking_configs["按母线排名"] = {
                        "dataframe": grouped_busbar_df,
                        "group_column": "母线"
                    }

            ranking_mode_options = ["全部站点总排名"] + list(grouped_ranking_configs.keys())
            if not grouped_ranking_configs:
                st.info("未找到可用的分组信息，当前仅显示全部站点总排名。")

            ranking_mode = st.radio(
                "排名视图",
                options=ranking_mode_options,
                horizontal=True
            )

            if ranking_mode == "全部站点总排名":
                # 添加厂站类型筛选器
                factory_filter_options = ["全部"] + factory_group_types if factory_group_types else ["全部"]
                selected_factory_filter = st.radio(
                    "厂站类型筛选",
                    options=factory_filter_options,
                    horizontal=True,
                    index=0
                )
                
                # 根据厂站类型筛选数据
                if selected_factory_filter == "全部":
                    filtered_stats_df = all_stations_stats.copy()
                else:
                    filtered_stats_df = all_stations_with_factory[
                        all_stations_with_factory[FACTORY_GROUP_COLUMN] == selected_factory_filter
                    ].copy()
                
                if len(filtered_stats_df) == 0:
                    st.warning(f"当前筛选条件下没有{selected_factory_filter}的数据！")
                    return
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总站点数", f"{len(filtered_stats_df)} 个")
                with col2:
                    max_spread_station = filtered_stats_df.iloc[0]
                    st.metric("最高电价差", f"{max_spread_station['日均峰谷价差']:.4f} 元/kWh",
                             delta=max_spread_station['站点名称'])
                with col3:
                    min_spread_station = filtered_stats_df.iloc[-1]
                    st.metric("最低电价差", f"{min_spread_station['日均峰谷价差']:.4f} 元/kWh",
                             delta=min_spread_station['站点名称'], delta_color="inverse")
                with col4:
                    avg_spread_all = filtered_stats_df['日均峰谷价差'].mean()
                    st.metric("平均电价差", f"{avg_spread_all:.4f} 元/kWh")

                table_title = f"📋 {'全部站点' if selected_factory_filter == '全部' else selected_factory_filter}总排名明细表"
                display_df = filtered_stats_df.copy()
                chart_source_df = display_df.copy()
                chart_x_col = '站点名称'
                chart_title_prefix = "前"
                chart_title_suffix = "个站点日均峰谷价差对比"
                chart_x_title = "站点名称"
                export_file_name = f"站点电价差总排名{'_' + selected_factory_filter if selected_factory_filter != '全部' else ''}.csv"
                chart_caption = "表格保留全部站点，图表默认只展示前若干名，避免一次性渲染过多柱子影响速度。"
                chart_hover_station = False
                chart_group_label = None
            else:
                grouped_config = grouped_ranking_configs[ranking_mode]
                group_column = grouped_config["group_column"]
                group_rank_column_label = f"{group_column}内排名"
                grouped_display_df = grouped_config["dataframe"].rename(columns={'排名': '总排名'}).copy()
                
                # 为分组排名数据添加厂站类型信息
                factory_group_df = build_station_group_mapping(
                    station_info_df,
                    FACTORY_GROUP_COLUMN,
                    station_names=grouped_display_df['站点名称'].tolist()
                )
                if factory_group_df is not None:
                    grouped_display_df = grouped_display_df.merge(
                        factory_group_df[['电站名', FACTORY_GROUP_COLUMN]],
                        left_on='站点名称',
                        right_on='电站名',
                        how='left'
                    ).drop(columns=['电站名'])
                    grouped_display_df[FACTORY_GROUP_COLUMN] = grouped_display_df[FACTORY_GROUP_COLUMN].fillna('未分组')
                
                # 添加厂站类型筛选器
                factory_filter_options = ["全部"] + factory_group_types if factory_group_types else ["全部"]
                selected_factory_filter = st.radio(
                    "厂站类型筛选",
                    options=factory_filter_options,
                    horizontal=True,
                    index=0
                )
                
                # 根据厂站类型筛选数据
                if selected_factory_filter != "全部":
                    grouped_display_df = grouped_display_df[
                        grouped_display_df[FACTORY_GROUP_COLUMN] == selected_factory_filter
                    ].copy()
                    # 重新计算组内排名
                    grouped_display_df = grouped_display_df.sort_values(
                        [group_column, '日均峰谷价差', '站点名称'],
                        ascending=[True, False, True]
                    ).reset_index(drop=True)
                    grouped_display_df['组内排名'] = grouped_display_df.groupby(group_column).cumcount() + 1
                
                available_groups = sorted(grouped_display_df[group_column].unique().tolist())
                selected_rank_group = st.selectbox(
                    f"选择{group_column}",
                    options=[f"全部{group_column}"] + available_groups,
                    index=0
                )

                if selected_rank_group == f"全部{group_column}":
                    champion_df = (
                        grouped_display_df[grouped_display_df['组内排名'] == 1]
                        .sort_values('日均峰谷价差', ascending=False)
                        .reset_index(drop=True)
                    )

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric(f"{group_column}组数", f"{len(available_groups)} 个")
                    with col2:
                        st.metric("总站点数", f"{len(grouped_display_df)} 个")
                    with col3:
                        champion_station = champion_df.iloc[0]
                        st.metric(f"最强{group_column}第1名", f"{champion_station['日均峰谷价差']:.4f} 元/kWh",
                                 delta=f"{champion_station[group_column]} - {champion_station['站点名称']}")
                    with col4:
                        st.metric("冠军平均电价差", f"{champion_df['日均峰谷价差'].mean():.4f} 元/kWh")

                    table_title = f"📋 各{group_column}排名明细表"
                    display_df = grouped_display_df[
                        [group_column, '组内排名', '总排名', '站点名称', '日均峰谷价差', '全年最高峰谷价差', '全年最低峰谷价差', '全年平均电价', '数据天数']
                    ].copy()
                    display_df = display_df.sort_values([group_column, '组内排名']).reset_index(drop=True)
                    display_df = display_df.rename(columns={'组内排名': group_rank_column_label})
                    chart_source_df = champion_df.copy()
                    chart_x_col = group_column
                    chart_title_prefix = "前"
                    chart_title_suffix = f"个{group_column}冠军电价差对比"
                    chart_x_title = group_column
                    export_file_name = f"{group_column}排名_全部{group_column}.csv"
                    chart_caption = f"表格显示全部{group_column}的排名，图表展示各{group_column}第1名的对比。"
                    chart_hover_station = True
                    chart_group_label = group_column
                else:
                    single_group_df = grouped_display_df[
                        grouped_display_df[group_column] == selected_rank_group
                    ].sort_values('组内排名').reset_index(drop=True)

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric(f"当前{group_column}站点数", f"{len(single_group_df)} 个")
                    with col2:
                        top_station = single_group_df.iloc[0]
                        st.metric(f"{group_column}第1名", f"{top_station['日均峰谷价差']:.4f} 元/kWh",
                                 delta=top_station['站点名称'])
                    with col3:
                        bottom_station = single_group_df.iloc[-1]
                        st.metric(f"{group_column}最后1名", f"{bottom_station['日均峰谷价差']:.4f} 元/kWh",
                                 delta=bottom_station['站点名称'], delta_color="inverse")
                    with col4:
                        st.metric(f"当前{group_column}平均电价差", f"{single_group_df['日均峰谷价差'].mean():.4f} 元/kWh")

                    table_title = f"📋 {selected_rank_group} 排名明细表"
                    display_df = single_group_df[
                        ['组内排名', '总排名', '站点名称', '日均峰谷价差', '全年最高峰谷价差', '全年最低峰谷价差', '全年平均电价', '数据天数']
                    ].copy()
                    display_df = display_df.rename(columns={'组内排名': group_rank_column_label})
                    chart_source_df = single_group_df.copy()
                    chart_x_col = '站点名称'
                    chart_title_prefix = selected_rank_group
                    chart_title_suffix = "排名电价差对比"
                    chart_x_title = "站点名称"
                    export_file_name = f"{selected_rank_group}_{group_column}排名.csv"
                    chart_caption = f"表格和图表都按当前{group_column}展示排名。"
                    chart_hover_station = False
                    chart_group_label = None

            st.subheader(table_title)
            st.dataframe(display_df, use_container_width=True, height=400)

            st.subheader("📈 电价差分布图")
            st.caption(chart_caption)

            if len(chart_source_df) == 0:
                st.info("当前条件下没有可用于绘图的数据。")
            else:
                chart_options = []
                for option in [20, 50, 100, 200, len(chart_source_df)]:
                    if option <= len(chart_source_df) and option not in chart_options:
                        chart_options.append(option)

                if len(chart_options) == 1:
                    chart_station_count = chart_options[0]
                    st.caption(f"当前仅有 {chart_station_count} 条数据，图表已自动展示全部。")
                else:
                    default_chart_count = 50 if 50 in chart_options else chart_options[-1]
                    chart_station_count = st.select_slider(
                        "图表展示数量",
                        options=chart_options,
                        value=default_chart_count
                    )

                chart_df = chart_source_df.head(chart_station_count).copy()
                show_value_labels = chart_station_count <= 50

                fig_spread = go.Figure()
                bar_kwargs = dict(
                    x=chart_df[chart_x_col],
                    y=chart_df['日均峰谷价差'],
                    name='日均峰谷价差',
                    marker_color='rgb(55, 83, 109)'
                )
                if show_value_labels:
                    bar_kwargs['text'] = [f"{x:.4f}" for x in chart_df['日均峰谷价差']]
                    bar_kwargs['textposition'] = 'auto'
                if chart_hover_station:
                    bar_kwargs['customdata'] = chart_df[['站点名称']].to_numpy()
                    bar_kwargs['hovertemplate'] = f"{chart_group_label}: %{{x}}<br>站点: %{{customdata[0]}}<br>日均峰谷价差: %{{y:.4f}} 元/kWh<extra></extra>"

                fig_spread.add_trace(go.Bar(**bar_kwargs))

                if ranking_mode == "全部站点总排名" or chart_x_col != '站点名称':
                    chart_title = f"{chart_title_prefix} {chart_station_count} {chart_title_suffix}"
                else:
                    chart_title = f"{chart_title_prefix}{chart_title_suffix}"

                fig_spread.update_layout(
                    title=chart_title,
                    xaxis_title=chart_x_title,
                    yaxis_title='日均峰谷价差 (元/kWh)',
                    template='plotly_white',
                    height=500,
                    xaxis=dict(tickangle=45),
                    showlegend=False
                )

                st.plotly_chart(fig_spread, use_container_width=True)

            st.subheader("💾 导出排名数据")
            csv_data = display_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载排名数据 (CSV)",
                data=csv_data,
                file_name=export_file_name,
                mime="text/csv"
            )
        else:
            st.warning("暂无电价差数据！")

    # 储能配储优化页
    elif view_mode == "🔋 储能配储优化":
        st.header(" 储能配储优化")
        st.markdown("""
        基于站点一年电价数据，进行储能配置优化，计算最优充放电策略和年收益。
        """)
            
        # 侧边栏：选择站点和储能参数
        st.sidebar.divider()
        st.sidebar.header("🔋 储能配储优化配置")
        
        # 站点筛选方式
        st.sidebar.subheader("📍 站点选择")
        filter_mode_options = ["全部站点"]
        if busbar_types:
            filter_mode_options.append("按母线查询")
        if city_types:
            filter_mode_options.append("按城市查询")
        if factory_group_types:
            filter_mode_options.append("按厂站类型查询")
        
        selected_filter_mode = st.sidebar.selectbox(
            "查询方式",
            options=filter_mode_options,
            index=0,
            help="选择站点筛选方式"
        )
        
        filtered_stations = list(price_files.keys())
        
        # 根据筛选方式过滤站点
        if selected_filter_mode == "按母线查询":
            group_options = busbar_types
            active_group_column = "母线"
        elif selected_filter_mode == "按城市查询":
            group_options = city_types
            active_group_column = "城市"
        elif selected_filter_mode == "按厂站类型查询":
            group_options = factory_group_types
            active_group_column = FACTORY_GROUP_COLUMN
        else:
            group_options = []
            active_group_column = None
        
        if active_group_column and group_options:
            all_group_option = f"全部{active_group_column}"
            active_group_value = st.sidebar.selectbox(
                f"选择{active_group_column}",
                options=[all_group_option] + group_options,
                index=0,
                help=f"选择要查看的{active_group_column}"
            )
            
            if active_group_value != all_group_option:
                station_group_df = build_station_group_mapping(
                    station_info_df,
                    active_group_column,
                    station_names=price_files.keys()
                )
                filtered_stations_df = station_group_df[
                    station_group_df[active_group_column] == active_group_value
                ]
                filtered_stations_list = (
                    filtered_stations_df['电站名']
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                filtered_stations = [s for s in filtered_stations_list if s in price_files.keys()]
                st.sidebar.info(f"📊 {active_group_value} 共有 {len(filtered_stations)} 个站点")
        
        if len(filtered_stations) == 0:
            st.sidebar.warning("当前筛选条件下没有可用的站点！")
            selected_station = None
        else:
            # 切换年份时保留已选站点
            prev_station = st.session_state.get('selected_station')
            if prev_station and prev_station in filtered_stations:
                default_station_index = filtered_stations.index(prev_station)
            else:
                default_station_index = 0

            # 选择站点
            selected_station = st.sidebar.selectbox(
                "选择站点",
                options=filtered_stations,
                index=default_station_index,
                help="选择要进行储能优化的站点"
            )
            st.session_state['selected_station'] = selected_station
            
        # 储能参数配置
        st.sidebar.subheader("⚙️ 储能参数")
        P = st.sidebar.number_input(
            "逆变器功率 (MW)",
            value=STORAGE_CONFIG['P'],
            min_value=1,
            step=10,
            help="储能逆变器额定功率"
        )
        battery_capacity = st.sidebar.number_input(
            "电池容量 (MWh)",
            value=STORAGE_CONFIG['battery_capacity'],
            min_value=1,
            step=10,
            help="电池总容量"
        )
        efficiency = st.sidebar.number_input(
            "放电效率",
            value=STORAGE_CONFIG['efficiency'],
            min_value=0.5,
            max_value=1.0,
            step=0.01,
            help="充放电循环效率"
        )

        # 策略模式选择
        st.sidebar.divider()
        st.sidebar.subheader("📋 策略模式")
        strategy_mode = st.sidebar.radio(
            "选择策略模式",
            options=["🚀 智能优化", "📝 自定义策略"],
            horizontal=True,
            help="智能优化：由算法自动寻找最优充放电策略；自定义策略：手动指定充放电时段"
        )

        # 自定义策略：充放电时段选择
        custom_charge_slots = []
        custom_discharge_slots = []
        if strategy_mode == "📝 自定义策略":
            st.sidebar.subheader("⏰ 充放电时段设置")
            time_slot_options = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]
            num_slots = len(time_slot_options)

            # 充放电轮数控制
            if "num_cycles" not in st.session_state:
                st.session_state.num_cycles = 1

            cycle_col1, cycle_col2 = st.sidebar.columns(2)
            with cycle_col1:
                if st.button("➕ 增加轮数", use_container_width=True, disabled=st.session_state.num_cycles >= 5):
                    st.session_state.num_cycles += 1
            with cycle_col2:
                if st.button("➖ 减少轮数", use_container_width=True, disabled=st.session_state.num_cycles <= 1):
                    st.session_state.num_cycles -= 1

            num_cycles = st.session_state.num_cycles
            st.sidebar.caption(f"当前共 {num_cycles} 轮充放电")

            # 已占用时段集合（用于排除重叠）
            all_occupied_slots = set()
            all_charge_slots_lists = []
            all_discharge_slots_lists = []
            has_error = False

            for cycle_idx in range(num_cycles):
                st.sidebar.markdown(f"---")
                st.sidebar.markdown(f"**第 {cycle_idx + 1} 轮**")

                # 充电时段
                available_charge_slots = [t for t in time_slot_options if t not in all_occupied_slots]
                if not available_charge_slots:
                    st.sidebar.error(f"第 {cycle_idx + 1} 轮：无可用充电时段！")
                    has_error = True
                    continue

                # 默认充电时段 11:00-13:00
                charge_start_default = "11:00" if "11:00" in available_charge_slots else available_charge_slots[0]
                charge_end_default = "13:00" if "13:00" in available_charge_slots else available_charge_slots[min(len(available_charge_slots) - 1, 31)]
                if available_charge_slots.index(charge_start_default) > available_charge_slots.index(charge_end_default):
                    charge_start_default = available_charge_slots[0]
                    charge_end_default = available_charge_slots[min(len(available_charge_slots) - 1, 31)]

                charge_start, charge_end = st.sidebar.select_slider(
                    f"充电起止时间",
                    options=available_charge_slots,
                    value=(charge_start_default, charge_end_default),
                    key=f"charge_slider_{cycle_idx}",
                    help="选择充电的起止时间（含两端）"
                )
                charge_start_idx = time_slot_options.index(charge_start)
                charge_end_idx = time_slot_options.index(charge_end)
                current_charge = time_slot_options[charge_start_idx:charge_end_idx + 1]
                all_charge_slots_lists.append(list(range(charge_start_idx, charge_end_idx + 1)))
                all_occupied_slots.update(current_charge)
                st.sidebar.caption(f"充电：{charge_start} ~ {charge_end}（{len(current_charge)} 个时段）")

                # 放电时段：排除所有已占用时段
                available_discharge_slots = [t for t in time_slot_options if t not in all_occupied_slots]
                if not available_discharge_slots:
                    st.sidebar.error(f"第 {cycle_idx + 1} 轮：无可用放电时段！")
                    has_error = True
                    continue

                # 默认放电时段 19:00-21:00
                discharge_start_default = "19:00" if "19:00" in available_discharge_slots else available_discharge_slots[0]
                discharge_end_default = "21:00" if "21:00" in available_discharge_slots else available_discharge_slots[min(len(available_discharge_slots) - 1, 19)]
                if available_discharge_slots.index(discharge_start_default) > available_discharge_slots.index(discharge_end_default):
                    discharge_start_default = available_discharge_slots[0]
                    discharge_end_default = available_discharge_slots[min(len(available_discharge_slots) - 1, 19)]

                discharge_start, discharge_end = st.sidebar.select_slider(
                    f"放电起止时间",
                    options=available_discharge_slots,
                    value=(discharge_start_default, discharge_end_default),
                    key=f"discharge_slider_{cycle_idx}",
                    help="选择放电的起止时间（含两端），不可与其他时段重叠"
                )
                discharge_start_idx = time_slot_options.index(discharge_start)
                discharge_end_idx = time_slot_options.index(discharge_end)
                current_discharge = time_slot_options[discharge_start_idx:discharge_end_idx + 1]
                all_discharge_slots_lists.append(list(range(discharge_start_idx, discharge_end_idx + 1)))
                all_occupied_slots.update(current_discharge)
                st.sidebar.caption(f"放电：{discharge_start} ~ {discharge_end}（{len(current_discharge)} 个时段）")

            # 合并所有轮次的时段
            for slots in all_charge_slots_lists:
                custom_charge_slots.extend(slots)
            for slots in all_discharge_slots_lists:
                custom_discharge_slots.extend(slots)

            # 在主区域展示各时间节点电价波动趋势，辅助用户选择充放电时段
            st.divider()
            st.subheader("📊 各时间节点电价波动趋势（辅助选时段）")

            with st.spinner("正在加载电价数据生成趋势图..."):
                if data_source == 'database':
                    _trend_year = st.session_state.get('selected_year')
                    _trend_df = load_price_data_from_db(selected_station, _trend_year)
                else:
                    _trend_df = load_price_data(*get_file_cache_key(price_files[selected_station]))

            if _trend_df is not None and len(_trend_df.columns) >= 97:
                _trend_time_cols = _trend_df.columns[1:97]
                _trend_time_columns = list(_trend_time_cols)

                _trend_stats_list = []
                for tc in _trend_time_cols:
                    td = _trend_df[tc].dropna()
                    _trend_stats_list.append({
                        '时间节点': tc,
                        '平均电价': td.mean(),
                        '电价标准差': td.std(),
                        '最低电价': td.min(),
                        '最高电价': td.max(),
                        '电价极差': td.max() - td.min()
                    })
                _trend_stats_df = pd.DataFrame(_trend_stats_list).round(4)

                # 均价曲线 + 标准差带
                fig_price_trend = go.Figure()
                fig_price_trend.add_trace(go.Scatter(
                    x=_trend_stats_df['时间节点'], y=_trend_stats_df['平均电价'],
                    mode='lines+markers', name='平均电价',
                    line=dict(color='#FF6B35', width=2), marker=dict(size=5)
                ))
                fig_price_trend.add_trace(go.Scatter(
                    x=_trend_stats_df['时间节点'],
                    y=_trend_stats_df['平均电价'] + _trend_stats_df['电价标准差'],
                    mode='lines', name='+标准差',
                    line=dict(color='rgba(255,107,53,0.3)', width=1, dash='dash'),
                    showlegend=False
                ))
                fig_price_trend.add_trace(go.Scatter(
                    x=_trend_stats_df['时间节点'],
                    y=_trend_stats_df['平均电价'] - _trend_stats_df['电价标准差'],
                    mode='lines', name='-标准差',
                    line=dict(color='rgba(255,107,53,0.3)', width=1, dash='dash'),
                    fill='tonexty', fillcolor='rgba(255,107,53,0.1)',
                    showlegend=False
                ))

                # 高亮充电时段（蓝色）和放电时段（红色）
                charge_time_strs = [time_slot_options[i] for i in custom_charge_slots]
                discharge_time_strs = [time_slot_options[i] for i in custom_discharge_slots]
                if charge_time_strs:
                    charge_prices = [_trend_stats_df.loc[_trend_stats_df['时间节点'] == t, '平均电价'].values[0]
                                     if len(_trend_stats_df[_trend_stats_df['时间节点'] == t]) > 0 else 0
                                     for t in charge_time_strs]
                    fig_price_trend.add_trace(go.Scatter(
                        x=charge_time_strs, y=charge_prices,
                        mode='markers', name='充电时段',
                        marker=dict(color='blue', size=10, symbol='triangle-up')
                    ))
                if discharge_time_strs:
                    discharge_prices = [_trend_stats_df.loc[_trend_stats_df['时间节点'] == t, '平均电价'].values[0]
                                        if len(_trend_stats_df[_trend_stats_df['时间节点'] == t]) > 0 else 0
                                        for t in discharge_time_strs]
                    fig_price_trend.add_trace(go.Scatter(
                        x=discharge_time_strs, y=discharge_prices,
                        mode='markers', name='放电时段',
                        marker=dict(color='red', size=10, symbol='triangle-down')
                    ))

                fig_price_trend.update_layout(
                    title=f'{selected_station} 各时间节点平均电价及波动范围',
                    xaxis_title='时间节点', yaxis_title='电价 (元/kWh)',
                    template='plotly_white', height=400,
                    xaxis=dict(tickangle=45, tickvals=_trend_time_columns[::4],
                               ticktext=[_trend_time_columns[i] for i in range(0, len(_trend_time_columns), 4)])
                )
                st.plotly_chart(fig_price_trend, use_container_width=True)
                st.caption(" 三角向上（蓝色）= 充电时段 | 三角向下（红色）= 放电时段 | 阴影区域 = ±1倍标准差")
            else:
                st.warning("无法加载电价数据生成趋势图")

        # 初始化 session_state
        if "opt_cache_key" not in st.session_state:
            st.session_state.opt_cache_key = None
        if "opt_results" not in st.session_state:
            st.session_state.opt_results = None

        # 根据数据源和策略模式生成缓存键
        data_source = st.session_state.get('data_source', 'excel')
        is_custom_mode = (strategy_mode == "📝 自定义策略")
        if data_source == 'database':
            selected_year = st.session_state.get('selected_year')
            if is_custom_mode:
                cache_key = (f"db_{selected_station}_{selected_year}", P, battery_capacity, efficiency, "custom", tuple(custom_charge_slots), tuple(custom_discharge_slots))
            else:
                cache_key = (f"db_{selected_station}_{selected_year}", P, battery_capacity, efficiency, "opt")
        else:
            if is_custom_mode:
                cache_key = (str(price_files.get(selected_station, '')), P, battery_capacity, efficiency, "custom", tuple(custom_charge_slots), tuple(custom_discharge_slots))
            else:
                cache_key = (str(price_files.get(selected_station, '')), P, battery_capacity, efficiency, "opt")
        
        params_changed = (st.session_state.opt_cache_key != cache_key)

        # 按钮文字根据模式切换
        button_label = "📊 开始计算" if is_custom_mode else "🚀 开始优化"
        run_opt = st.button(button_label, type="primary", use_container_width=True)

        if run_opt:
            # 自定义策略校验
            if is_custom_mode:
                if not custom_charge_slots:
                    st.error("请至少选择一个充电时段！")
                    st.stop()
                if not custom_discharge_slots:
                    st.error("请至少选择一个放电时段！")
                    st.stop()
                overlap = set(custom_charge_slots) & set(custom_discharge_slots)
                if overlap:
                    overlap_str = ", ".join([time_slot_options[i] for i in sorted(overlap)])
                    st.error(f"充电时段和放电时段不能重叠！重叠时段：{overlap_str}")
                    st.stop()

            with st.spinner("正在加载电价数据并计算..."):
                try:
                    if is_custom_mode:
                        # 自定义策略模式
                        if data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            df_data = load_price_data_from_db(selected_station, selected_year)
                        else:
                            df_data = load_price_data(*get_file_cache_key(price_files[selected_station]))

                        if df_data is None:
                            st.error("数据加载失败！")
                            st.stop()

                        start_col = 1 if len(df_data.columns) > 96 and not pd.api.types.is_numeric_dtype(df_data.iloc[:, 0]) else 0
                        all_prices = df_data.iloc[:, start_col:start_col + 96].values.astype(float)

                        date_list = []
                        if start_col == 1:
                            for d in df_data.iloc[:, 0].values:
                                if pd.isna(d):
                                    date_list.append('')
                                elif isinstance(d, pd.Timestamp):
                                    date_list.append(d.strftime('%Y-%m-%d'))
                                elif isinstance(d, str):
                                    ds = str(d).strip()
                                    date_list.append(ds.split('T')[0] if 'T' in ds else (ds.split(' ')[0] if ' ' in ds else ds[:10]))
                                else:
                                    date_list.append(str(d)[:10])
                        else:
                            date_list = [f'第{i+1}天' for i in range(len(df_data))]

                        config = {
                            'P': P * 1000,
                            'battery_capacity': battery_capacity * 1000,
                            'efficiency': efficiency,
                            'dt': 0.25,
                            'num': 96
                        }
                        all_results, all_summaries = run_custom_strategy(
                            all_prices, date_list, custom_charge_slots, custom_discharge_slots, config
                        )
                    else:
                        # 智能优化模式：先查共享缓存
                        if data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            shared_key = _shared_opt_cache_key(selected_station, P, battery_capacity, efficiency, 'database', selected_year)
                        else:
                            shared_key = _shared_opt_cache_key(selected_station, P, battery_capacity, efficiency, 'excel', str(price_files[selected_station]))

                        shared_result = get_shared_opt_result(shared_key)
                        if shared_result is not None:
                            date_list, all_results, all_summaries = shared_result
                        elif data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            date_list, all_results, all_summaries = run_optimization_cached_from_db(
                                selected_station, selected_year, P, battery_capacity, efficiency
                            )
                            put_shared_opt_result(shared_key, date_list, all_results, all_summaries)
                        else:
                            date_list, all_results, all_summaries = run_optimization_cached(
                                price_files[selected_station], P, battery_capacity, efficiency
                            )
                            put_shared_opt_result(shared_key, date_list, all_results, all_summaries)
                    
                    st.session_state.opt_results = {
                        'date_list': date_list,
                        'all_results': all_results,
                        'all_summaries': all_summaries,
                        'is_custom': is_custom_mode,
                        'num_cycles': st.session_state.num_cycles if is_custom_mode else None,
                    }
                    st.session_state.opt_cache_key = cache_key
                    params_changed = False
                except Exception as e:
                    st.error(f"计算失败：{str(e)}")
                    st.exception(e)
                    st.stop()

        # 显示已缓存的结果（按钮按下后或之前已计算过且参数未变）
        if st.session_state.opt_results is not None and not params_changed:
            date_list = st.session_state.opt_results['date_list']
            all_results = st.session_state.opt_results['all_results']
            all_summaries = st.session_state.opt_results['all_summaries']
            result_is_custom = st.session_state.opt_results.get('is_custom', False)
            num_days = len(all_summaries)

            # 计算年收益
            total_yearly_revenue = sum([s['日收益_元'] for s in all_summaries])

            # 显示年收益
            if result_is_custom:
                st.success(f"🎉 自定义策略计算完成！该站点年收益：**{total_yearly_revenue:,.2f} 元**")
            else:
                st.success(f"🎉 优化完成！该站点年收益：**{total_yearly_revenue:,.2f} 元**")

            # 显示参数和收益摘要
            if result_is_custom:
                result_num_cycles = st.session_state.opt_results.get('num_cycles', 1)
                avg_daily_revenue = total_yearly_revenue / num_days
                implied_spread = (avg_daily_revenue / (battery_capacity * 1000)) / result_num_cycles

                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("逆变器功率", f"{P:.0f} MW")
                with col2:
                    st.metric("电池容量", f"{battery_capacity:.0f} MWh")
                with col3:
                    st.metric("年收益", f"{total_yearly_revenue:,.0f} 元")
                with col4:
                    st.metric("日均收益", f"{avg_daily_revenue:.0f} 元")
                with col5:
                    st.metric(
                        "反算电价差",
                        f"{implied_spread:.4f} 元/kWh",
                        help=f"计算公式：(日均收益 / 电池容量) / 充放电轮数 = ({avg_daily_revenue:.0f} / {battery_capacity * 1000:.0f}) / {result_num_cycles}"
                    )
            else:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("逆变器功率", f"{P:.0f} MW")
                with col2:
                    st.metric("电池容量", f"{battery_capacity:.0f} MWh")
                with col3:
                    st.metric("年收益", f"{total_yearly_revenue:,.0f} 元")
                with col4:
                    st.metric("日均收益", f"{total_yearly_revenue/num_days:.0f} 元")

            # 每日收益折线图
            st.divider()
            st.subheader("📈 每日收益趋势")

            # 将日期转换为中文月日格式
            def to_chinese_date(d):
                try:
                    parts = str(d).strip().split('-')
                    if len(parts) >= 3:
                        return f"{int(parts[1])}月{int(parts[2])}日"
                    return str(d)
                except (ValueError, IndexError):
                    return str(d)

            chinese_dates = [to_chinese_date(d) for d in date_list]

            daily_revenues = [s['日收益_元'] for s in all_summaries]
            fig_daily_rev = go.Figure()
            fig_daily_rev.add_trace(go.Scatter(
                x=chinese_dates,
                y=daily_revenues,
                mode='lines+markers',
                name='日收益',
                line=dict(color='steelblue', width=2),
                marker=dict(size=4),
                fill='tozeroy',
                fillcolor='rgba(70,130,180,0.1)'
            ))
            # 添加均值参考线
            avg_daily = np.mean(daily_revenues)
            fig_daily_rev.add_hline(
                y=avg_daily, line_dash="dash", line_color="red",
                annotation_text=f"日均: {avg_daily:.0f} 元"
            )
            fig_daily_rev.update_layout(
                title="每日收益折线图",
                xaxis_title='日期',
                yaxis_title='收益 (元)',
                template='plotly_white',
                height=400,
                xaxis=dict(tickangle=45)
            )
            st.plotly_chart(fig_daily_rev, use_container_width=True)

            # 选择日期查看策略
            st.divider()
            st.subheader("📅 查看单日配储策略")

            date_options = [s['日期'] for s in all_summaries]

            # 用 session_state 记住选中的日期索引
            if "selected_date_idx" not in st.session_state or st.session_state.get("prev_num_days") != num_days:
                st.session_state.selected_date_idx = 0

            selected_date_idx = st.selectbox(
                "选择日期",
                options=range(len(date_options)),
                format_func=lambda x: date_options[x],
                index=st.session_state.selected_date_idx,
                key="date_selector"
            )
            st.session_state.selected_date_idx = selected_date_idx
            st.session_state.prev_num_days = num_days

            selected_date = date_options[selected_date_idx]
            selected_summary = all_summaries[selected_date_idx]

            # 显示该日统计
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("日收益", f"{selected_summary['日收益_元']:.2f} 元")
            with col2:
                st.metric("充电量", f"{selected_summary['充电量_kWh']:.2f} kWh")
            with col3:
                st.metric("放电量", f"{selected_summary['放电量_kWh']:.2f} kWh")

            # 显示该日详细数据
            day_results = [r for r in all_results if r['日期'] == selected_date]
            day_df = pd.DataFrame(day_results)

            # 充放电功率曲线（合并到一张图）
            fig_power = go.Figure()
            fig_power.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['充电功率_kW'],
                mode='lines',
                name='充电功率',
                line=dict(color='blue', width=2),
                fill='tozeroy',
                fillcolor='rgba(0,0,255,0.1)'
            ))
            fig_power.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['放电功率_kW'],
                mode='lines',
                name='放电功率',
                line=dict(color='red', width=2),
                fill='tozeroy',
                fillcolor='rgba(255,0,0,0.1)'
            ))
            fig_power.update_layout(
                title=f"{selected_date} 充放电功率曲线",
                xaxis_title='时间',
                yaxis_title='功率 (kW)',
                template='plotly_white',
                height=400,
                xaxis=dict(tickangle=45, tickvals=day_df['时间'][::8])
            )
            st.plotly_chart(fig_power, use_container_width=True)

            # 电池电量曲线
            fig_soc = go.Figure()
            fig_soc.add_trace(go.Scatter(
                x=day_df['时间'],
                y=day_df['电池电量_kWh'],
                mode='lines',
                name='电池电量',
                line=dict(color='green', width=2),
                fill='tozeroy',
                fillcolor='rgba(0,255,0,0.1)'
            ))
            fig_soc.update_layout(
                title=f"{selected_date} 电池电量曲线",
                xaxis_title='时间',
                yaxis_title='电池电量 (kWh)',
                template='plotly_white',
                height=350,
                xaxis=dict(tickangle=45, tickvals=day_df['时间'][::8])
            )
            st.plotly_chart(fig_soc, use_container_width=True)

            # 显示详细表格
            st.subheader("📊 详细数据表")
            st.dataframe(day_df, use_container_width=True, height=400)

            # 导出结果
            st.divider()
            st.subheader("💾 导出结果")

            results_df = pd.DataFrame(all_results)
            summary_df = pd.DataFrame(all_summaries)

            mode_suffix = "自定义策略" if result_is_custom else "储能优化"

            csv_data = results_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载详细结果 (CSV)",
                data=csv_data,
                file_name=f"{selected_station}_{mode_suffix}_详细结果.csv",
                mime="text/csv"
            )

            summary_csv = summary_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载每日收益汇总 (CSV)",
                data=summary_csv,
                file_name=f"{selected_station}_{mode_suffix}_每日收益汇总.csv",
                mime="text/csv"
            )

        elif st.session_state.opt_results is not None and params_changed:
            st.info("参数已变更，请点击上方按钮重新计算。")

    # 区域节点电价对比分析页
    elif view_mode == "📊 区域节点电价对比":
        st.header("📊 区域节点电价对比分析")
        st.markdown("""
        对珠三角、粤北、粤东、粤西等区域节点电价进行对比分析。
        """)

        # 获取可用区域
        has_region_info = (
            station_info_df is not None
            and '城市' in station_info_df.columns
            and '电站名' in station_info_df.columns
        )

        if not has_region_info:
            st.warning("无法获取站点城市信息，请检查电站名.xlsx文件。")
            return

        # 侧边栏：区域选择
        st.sidebar.divider()
        st.sidebar.header("🌍 区域对比配置")

        # 构建城市到区域的映射
        city_to_region = {}
        for region, cities in REGION_MAPPING.items():
            for city in cities:
                city_to_region[city] = region

        # 获取可用区域
        available_cities = station_info_df['城市'].dropna().unique().tolist()
        available_regions = sorted(set(city_to_region.get(city, '其他') for city in available_cities if city in city_to_region))

        if not available_regions:
            st.warning("没有找到可匹配的区域数据。")
            return

        # 选择对比区域
        selected_regions = st.sidebar.multiselect(
            "选择对比区域",
            options=available_regions,
            default=available_regions[:2] if len(available_regions) >= 2 else available_regions[:1],
            help="选择要对比的区域（可多选）"
        )

        if not selected_regions:
            st.warning("请至少选择一个区域进行对比。")
            return

        # 选择对比指标
        comparison_metric = st.sidebar.radio(
            "对比指标",
            options=["平均电价", "峰谷价差", "最高电价", "最低电价"],
            index=0,
            help="选择要对比的电价指标"
        )

        # 获取选中区域的站点列表
        region_stations = {}
        for region in selected_regions:
            cities_in_region = REGION_MAPPING.get(region, [])
            stations = station_info_df[
                station_info_df['城市'].isin(cities_in_region)
            ]['电站名'].astype(str).str.strip().tolist()
            # 过滤出有电价数据的站点
            stations_with_data = [s for s in stations if s in price_files.keys()]
            region_stations[region] = stations_with_data

        # 显示区域站点统计
        st.subheader("📋 区域站点统计")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("对比区域数", f"{len(selected_regions)} 个")
        with col2:
            total_stations = sum(len(v) for v in region_stations.values())
            st.metric("总站点数", f"{total_stations} 个")
        with col3:
            max_region = max(region_stations.items(), key=lambda x: len(x[1]))
            st.metric("最多站点区域", f"{max_region[0]} ({len(max_region[1])} 个)")
        with col4:
            min_region = min(region_stations.items(), key=lambda x: len(x[1]))
            st.metric("最少站点区域", f"{min_region[0]} ({len(min_region[1])} 个)")

        # 显示各区域站点列表
        with st.expander("查看各区域站点列表"):
            for region in selected_regions:
                st.markdown(f"**{region}** ({len(region_stations[region])} 个站点): {', '.join(region_stations[region][:10])}{'...' if len(region_stations[region]) > 10 else ''}")

        st.divider()

        # 计算区域平均电价
        @st.cache_data(show_spinner="正在计算区域电价数据...")
        def compute_regional_averages(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域的平均电价数据"""
            regional_data = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                # 收集所有站点的96时段均值（每个站点先算自己的时段均值）
                station_time_avgs = []
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is not None and len(df.columns) >= 97:
                            price_cols = df.columns[1:97]
                            # 计算该站点每个时段的平均电价（跨天平均）
                            time_avg = df[price_cols].mean(axis=0).values.astype(float)
                            if len(time_avg) == 96 and not np.all(np.isnan(time_avg)):
                                station_time_avgs.append(time_avg)
                    except Exception:
                        continue

                if station_time_avgs:
                    # 计算该区域所有站点的时段平均值
                    regional_avg = np.nanmean(station_time_avgs, axis=0)
                    regional_data[region] = regional_avg

            return regional_data

        @st.cache_data(show_spinner=False)
        def compute_regional_monthly_averages(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域每月的均价数据，返回 {region: {月份int: 均价float}}"""
            regional_monthly = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                # 收集每个站点每月的均价
                month_prices = defaultdict(list)
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is None or len(df.columns) < 2:
                            continue

                        date_col = df.columns[0]
                        price_cols = df.columns[1:]
                        df_copy = df[[date_col]].copy()
                        df_copy['_month'] = pd.to_datetime(df_copy[date_col]).dt.month
                        df_copy['_avg_price'] = df[price_cols].apply(pd.to_numeric, errors='coerce').mean(axis=1)

                        for month_val, group in df_copy.groupby('_month'):
                            month_avg = group['_avg_price'].mean()
                            if not np.isnan(month_avg):
                                month_prices[int(month_val)].append(month_avg)
                    except Exception:
                        continue

                if month_prices:
                    regional_monthly[region] = {
                        m: round(float(np.nanmean(prices)), 4)
                        for m, prices in sorted(month_prices.items())
                    }

            return regional_monthly

        @st.cache_data(show_spinner=False)
        def compute_regional_daily_averages(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域每日的均价数据，返回 {region: [日均价列表]}"""
            regional_daily = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                daily_avgs = []
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is None or len(df.columns) < 2:
                            continue

                        price_cols = df.columns[1:]
                        day_means = df[price_cols].apply(pd.to_numeric, errors='coerce').mean(axis=1)
                        daily_avgs.extend(day_means.dropna().tolist())
                    except Exception:
                        continue

                if daily_avgs:
                    regional_daily[region] = daily_avgs

            return regional_daily

        @st.cache_data(show_spinner=False)
        def compute_regional_daily_volatility(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域每日电价标准差（日波动率），返回 {region: [每日std列表]}"""
            regional_volatility = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                daily_stds = []
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is None or len(df.columns) < 2:
                            continue

                        price_cols = df.columns[1:]
                        day_stds = df[price_cols].apply(pd.to_numeric, errors='coerce').std(axis=1)
                        daily_stds.extend(day_stds.dropna().tolist())
                    except Exception:
                        continue

                if daily_stds:
                    regional_volatility[region] = daily_stds

            return regional_volatility

        @st.cache_data(show_spinner=False)
        def compute_regional_monthly_volatility(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域每月的电价波动率（月内每日标准差的均值），返回 {region: {月份int: 波动率float}}"""
            regional_monthly_vol = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                month_stds = defaultdict(list)
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is None or len(df.columns) < 2:
                            continue

                        date_col = df.columns[0]
                        price_cols = df.columns[1:]
                        df_copy = df[[date_col]].copy()
                        df_copy['_month'] = pd.to_datetime(df_copy[date_col]).dt.month
                        df_copy['_daily_std'] = df[price_cols].apply(pd.to_numeric, errors='coerce').std(axis=1)

                        for month_val, group in df_copy.groupby('_month'):
                            month_avg_std = group['_daily_std'].mean()
                            if not np.isnan(month_avg_std):
                                month_stds[int(month_val)].append(month_avg_std)
                    except Exception:
                        continue

                if month_stds:
                    regional_monthly_vol[region] = {
                        m: round(float(np.nanmean(stds)), 4)
                        for m, stds in sorted(month_stds.items())
                    }

            return regional_monthly_vol

        @st.cache_data(show_spinner=False)
        def compute_regional_monthly_peak_valley_spread(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域每月的峰谷价差（月内每日峰谷价差的均值），返回 {region: {月份int: 峰谷价差float}}"""
            n_peak = 8  # 滑动窗口大小
            regional_monthly_spread = {}

            for region, stations in region_stations_dict.items():
                if not stations:
                    continue

                month_spreads = defaultdict(list)
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue

                        if df is None or len(df.columns) < 2:
                            continue

                        date_col = df.columns[0]
                        price_cols = df.columns[1:]
                        df_copy = df[[date_col]].copy()
                        df_copy['_month'] = pd.to_datetime(df_copy[date_col]).dt.month

                        # 计算每日峰谷价差
                        daily_spreads = []
                        for _, row in df.iterrows():
                            prices = pd.to_numeric(row[price_cols], errors='coerce').dropna().values
                            if len(prices) >= n_peak * 2:
                                _, _, spread = _sliding_window_spread(prices, n_peak)
                                daily_spreads.append(spread)
                            else:
                                daily_spreads.append(np.nan)
                        df_copy['_daily_spread'] = daily_spreads

                        for month_val, group in df_copy.groupby('_month'):
                            month_avg_spread = group['_daily_spread'].dropna().mean()
                            if not np.isnan(month_avg_spread):
                                month_spreads[int(month_val)].append(month_avg_spread)
                    except Exception:
                        continue

                if month_spreads:
                    regional_monthly_spread[region] = {
                        m: round(float(np.nanmean(spreads)), 4)
                        for m, spreads in sorted(month_spreads.items())
                    }

            return regional_monthly_spread

        @st.cache_data(show_spinner=False)
        def compute_regional_pair_spread_timeseries(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域对在每个时点（96时段）的价差时间序列及其标准差。

            返回:
                pair_spread_data: {
                    (r1, r2): {
                        'spread_series': np.array(96),
                        'std': float,
                        'mean': float,
                        'max_spread': float,
                        'min_spread': float,
                        'max_slot': str,
                        'min_slot': str,
                    }
                }
            """
            # 先计算各区域的96时段均价
            regional_avgs = {}
            for region, stations in region_stations_dict.items():
                if not stations:
                    continue
                station_time_avgs = []
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue
                        if df is not None and len(df.columns) >= 97:
                            price_cols = df.columns[1:97]
                            time_avg = df[price_cols].mean(axis=0).values.astype(float)
                            if len(time_avg) == 96 and not np.all(np.isnan(time_avg)):
                                station_time_avgs.append(time_avg)
                    except Exception:
                        continue
                if station_time_avgs:
                    regional_avgs[region] = np.nanmean(station_time_avgs, axis=0)

            time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]
            from itertools import combinations as _comb
            pair_spread_data = {}
            for r1, r2 in _comb(regional_avgs.keys(), 2):
                d1 = regional_avgs[r1]
                d2 = regional_avgs[r2]
                spread = d1 - d2
                valid = spread[~np.isnan(spread)]
                if len(valid) == 0:
                    continue
                max_idx = int(np.nanargmax(spread))
                min_idx = int(np.nanargmin(spread))
                pair_spread_data[(r1, r2)] = {
                    'spread_series': spread,
                    'std': round(float(np.std(valid)), 4),
                    'mean': round(float(np.mean(valid)), 4),
                    'max_spread': round(float(np.max(valid)), 4),
                    'min_spread': round(float(np.min(valid)), 4),
                    'max_slot': time_columns[max_idx] if max_idx < len(time_columns) else "N/A",
                    'min_slot': time_columns[min_idx] if min_idx < len(time_columns) else "N/A",
                }

            return pair_spread_data

        @st.cache_data(show_spinner=False)
        def compute_regional_pair_monthly_spread_std(region_stations_dict, data_source, selected_year=None, price_files_key=None):
            """计算各区域对每月的价差时序标准差，用于分析趋势变化。

            返回:
                {(r1, r2): {月份int: 月内价差时序标准差float}}
            """
            # 按月按区域计算96时段均价
            regional_monthly_avgs = {}  # {region: {month: np.array(96)}}
            for region, stations in region_stations_dict.items():
                if not stations:
                    continue
                month_slot_prices = defaultdict(list)  # {month: [np.array(96), ...]}
                for station in stations:
                    try:
                        if data_source == 'database':
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            if station in price_files and price_files[station] is not None:
                                df = load_price_data(*get_file_cache_key(price_files[station]))
                            else:
                                continue
                        if df is None or len(df.columns) < 97:
                            continue
                        date_col = df.columns[0]
                        price_cols = df.columns[1:97]
                        df_copy = df.copy()
                        df_copy['_month'] = pd.to_datetime(df_copy[date_col]).dt.month
                        for month_val, group in df_copy.groupby('_month'):
                            slot_avg = group[price_cols].apply(pd.to_numeric, errors='coerce').mean(axis=0).values.astype(float)
                            if len(slot_avg) == 96 and not np.all(np.isnan(slot_avg)):
                                month_slot_prices[int(month_val)].append(slot_avg)
                    except Exception:
                        continue
                if month_slot_prices:
                    regional_monthly_avgs[region] = {
                        m: np.nanmean(arrs, axis=0) for m, arrs in sorted(month_slot_prices.items())
                    }

            from itertools import combinations as _comb
            pair_monthly_std = {}
            regions = [r for r in regional_monthly_avgs.keys()]
            for r1, r2 in _comb(regions, 2):
                m1 = regional_monthly_avgs.get(r1, {})
                m2 = regional_monthly_avgs.get(r2, {})
                common_months = sorted(set(m1.keys()) & set(m2.keys()))
                monthly_stds = {}
                for m in common_months:
                    spread = m1[m] - m2[m]
                    valid = spread[~np.isnan(spread)]
                    if len(valid) > 0:
                        monthly_stds[m] = round(float(np.std(valid)), 4)
                if monthly_stds:
                    pair_monthly_std[(r1, r2)] = monthly_stds

            return pair_monthly_std

        data_source = st.session_state.get('data_source', 'excel')
        selected_year = st.session_state.get('selected_year', None)

        # 生成文件路径缓存键，确保切换年份时缓存失效
        price_files_key = tuple(sorted(
            (name, str(path)) for name, path in price_files.items() if path is not None
        )) if data_source != 'database' else (selected_year,)

        with st.spinner("正在计算区域电价数据..."):
            regional_data = compute_regional_averages(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_monthly_data = compute_regional_monthly_averages(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_daily_data = compute_regional_daily_averages(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_volatility_data = compute_regional_daily_volatility(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_monthly_vol_data = compute_regional_monthly_volatility(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_monthly_pv_spread_data = compute_regional_monthly_peak_valley_spread(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_pair_spread_ts = compute_regional_pair_spread_timeseries(
                region_stations, data_source, selected_year, price_files_key
            )
            regional_pair_monthly_std = compute_regional_pair_monthly_spread_std(
                region_stations, data_source, selected_year, price_files_key
            )

        if not regional_data:
            st.warning("选中区域没有可用的电价数据。")
            return

        # 生成时间列
        time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]

        # 颜色配置
        region_colors = {'珠三角': '#FF6B35', '粤北': '#2196F3', '粤东': '#4CAF50', '粤西': '#9C27B0'}

        # 根据对比指标进行分析
        if comparison_metric == "平均电价":
            st.subheader("📊 区域平均电价对比")

            # 计算各区域全天平均电价
            region_avg_prices = {}
            for region, data in regional_data.items():
                region_avg_prices[region] = float(np.nanmean(data))

            # 创建对比表格
            comparison_df = pd.DataFrame({
                '区域': list(region_avg_prices.keys()),
                '平均电价 (元/kWh)': [round(float(v), 4) for v in region_avg_prices.values()]
            })
            comparison_df = comparison_df.sort_values('平均电价 (元/kWh)', ascending=False)

            # 显示指标卡片
            col1, col2, col3 = st.columns(3)
            with col1:
                highest_region = comparison_df.iloc[0]
                st.metric("最高均价区域", f"{highest_region['区域']}", 
                         delta=f"{highest_region['平均电价 (元/kWh)']:.4f} 元/kWh")
            with col2:
                lowest_region = comparison_df.iloc[-1]
                st.metric("最低均价区域", f"{lowest_region['区域']}", 
                         delta=f"{lowest_region['平均电价 (元/kWh)']:.4f} 元/kWh", delta_color="inverse")
            with col3:
                price_diff = highest_region['平均电价 (元/kWh)'] - lowest_region['平均电价 (元/kWh)']
                st.metric("区域间价差", f"{price_diff:.4f} 元/kWh")

            # 绘制柱状图
            fig_bar = go.Figure()
            for i, (_, row) in enumerate(comparison_df.iterrows()):
                fig_bar.add_trace(go.Bar(
                    x=[row['区域']],
                    y=[row['平均电价 (元/kWh)']],
                    name=row['区域'],
                    marker_color=region_colors.get(row['区域'], '#999999'),
                    text=[f"{row['平均电价 (元/kWh)']:.4f}"],
                    textposition='auto'
                ))
            fig_bar.update_layout(
                title='区域平均电价对比',
                xaxis_title='区域', yaxis_title='平均电价 (元/kWh)',
                template='plotly_white', height=400, showlegend=False
            )
            st.plotly_chart(fig_bar, use_container_width=True)
            st.dataframe(comparison_df, use_container_width=True)

        elif comparison_metric == "峰谷价差":
            st.subheader("📊 区域峰谷价差对比")

            # 计算各区域峰谷价差
            peak_valley_data = []
            for region, data in regional_data.items():
                peak_price = float(np.nanmax(data))
                valley_price = float(np.nanmin(data))
                spread = peak_price - valley_price
                peak_idx = int(np.nanargmax(data))
                valley_idx = int(np.nanargmin(data))
                peak_time = time_columns[peak_idx] if peak_idx < len(time_columns) else "N/A"
                valley_time = time_columns[valley_idx] if valley_idx < len(time_columns) else "N/A"
                peak_valley_data.append({
                    '区域': region,
                    '峰值电价': round(peak_price, 4),
                    '谷值电价': round(valley_price, 4),
                    '峰谷价差': round(spread, 4),
                    '峰值时段': peak_time,
                    '谷值时段': valley_time
                })

            pv_df = pd.DataFrame(peak_valley_data).sort_values('峰谷价差', ascending=False)

            # 显示指标卡片
            col1, col2, col3 = st.columns(3)
            with col1:
                highest_pv = pv_df.iloc[0]
                st.metric("最大峰谷价差区域", f"{highest_pv['区域']}", 
                         delta=f"{highest_pv['峰谷价差']:.4f} 元/kWh")
            with col2:
                lowest_pv = pv_df.iloc[-1]
                st.metric("最小峰谷价差区域", f"{lowest_pv['区域']}", 
                         delta=f"{lowest_pv['峰谷价差']:.4f} 元/kWh", delta_color="inverse")
            with col3:
                pv_diff = highest_pv['峰谷价差'] - lowest_pv['峰谷价差']
                st.metric("价差差距", f"{pv_diff:.4f} 元/kWh")

            # 绘制柱状图
            fig_pv = go.Figure()
            for _, row in pv_df.iterrows():
                fig_pv.add_trace(go.Bar(
                    x=[row['区域']],
                    y=[row['峰谷价差']],
                    name=row['区域'],
                    marker_color=region_colors.get(row['区域'], '#999999'),
                    text=[f"{row['峰谷价差']:.4f}"],
                    textposition='auto'
                ))
            fig_pv.update_layout(
                title='区域峰谷价差对比',
                xaxis_title='区域', yaxis_title='峰谷价差 (元/kWh)',
                template='plotly_white', height=400, showlegend=False
            )
            st.plotly_chart(fig_pv, use_container_width=True)
            st.dataframe(pv_df, use_container_width=True)

        elif comparison_metric == "最高电价":
            st.subheader("📊 区域最高电价对比")

            max_price_data = []
            for region, data in regional_data.items():
                max_price = float(np.nanmax(data))
                max_idx = int(np.nanargmax(data))
                max_time = time_columns[max_idx] if max_idx < len(time_columns) else "N/A"
                max_price_data.append({
                    '区域': region,
                    '最高电价': round(max_price, 4),
                    '出现时段': max_time
                })

            max_df = pd.DataFrame(max_price_data).sort_values('最高电价', ascending=False)

            col1, col2 = st.columns(2)
            with col1:
                top = max_df.iloc[0]
                st.metric("最高电价区域", f"{top['区域']}", 
                         delta=f"{top['最高电价']:.4f} 元/kWh (时段: {top['出现时段']})")
            with col2:
                bottom = max_df.iloc[-1]
                st.metric("最低最高价区域", f"{bottom['区域']}", 
                         delta=f"{bottom['最高电价']:.4f} 元/kWh", delta_color="inverse")

            fig_max = go.Figure()
            for _, row in max_df.iterrows():
                fig_max.add_trace(go.Bar(
                    x=[row['区域']], y=[row['最高电价']],
                    name=row['区域'],
                    marker_color=region_colors.get(row['区域'], '#999999'),
                    text=[f"{row['最高电价']:.4f}"], textposition='auto'
                ))
            fig_max.update_layout(
                title='区域最高电价对比', xaxis_title='区域', yaxis_title='最高电价 (元/kWh)',
                template='plotly_white', height=400, showlegend=False
            )
            st.plotly_chart(fig_max, use_container_width=True)
            st.dataframe(max_df, use_container_width=True)

        elif comparison_metric == "最低电价":
            st.subheader("📊 区域最低电价对比")

            min_price_data = []
            for region, data in regional_data.items():
                min_price = float(np.nanmin(data))
                min_idx = int(np.nanargmin(data))
                min_time = time_columns[min_idx] if min_idx < len(time_columns) else "N/A"
                min_price_data.append({
                    '区域': region,
                    '最低电价': round(min_price, 4),
                    '出现时段': min_time
                })

            min_df = pd.DataFrame(min_price_data).sort_values('最低电价', ascending=True)

            col1, col2 = st.columns(2)
            with col1:
                top = min_df.iloc[0]
                st.metric("最低电价区域", f"{top['区域']}", 
                         delta=f"{top['最低电价']:.4f} 元/kWh (时段: {top['出现时段']})")
            with col2:
                bottom = min_df.iloc[-1]
                st.metric("最高最低价区域", f"{bottom['区域']}", 
                         delta=f"{bottom['最低电价']:.4f} 元/kWh", delta_color="inverse")

            fig_min = go.Figure()
            for _, row in min_df.iterrows():
                fig_min.add_trace(go.Bar(
                    x=[row['区域']], y=[row['最低电价']],
                    name=row['区域'],
                    marker_color=region_colors.get(row['区域'], '#999999'),
                    text=[f"{row['最低电价']:.4f}"], textposition='auto'
                ))
            fig_min.update_layout(
                title='区域最低电价对比', xaxis_title='区域', yaxis_title='最低电价 (元/kWh)',
                template='plotly_white', height=400, showlegend=False
            )
            st.plotly_chart(fig_min, use_container_width=True)
            st.dataframe(min_df, use_container_width=True)

        # 时段趋势对比（所有指标模式下都显示）
        st.divider()
        st.subheader("📈 区域电价时段趋势对比")

        time_comparison_data = []
        for region, data in regional_data.items():
            for i, time_col in enumerate(time_columns[:96]):
                if i < len(data):
                    time_comparison_data.append({
                        '区域': region, '时段': time_col,
                        '平均电价': round(float(data[i]), 4)
                    })

        if time_comparison_data:
            time_comp_df = pd.DataFrame(time_comparison_data)
            fig_time = go.Figure()
            for region in selected_regions:
                region_data = time_comp_df[time_comp_df['区域'] == region]
                if not region_data.empty:
                    fig_time.add_trace(go.Scatter(
                        x=region_data['时段'], y=region_data['平均电价'],
                        mode='lines+markers', name=region,
                        line=dict(color=region_colors.get(region, '#999999'), width=2),
                        marker=dict(size=4)
                    ))
            fig_time.update_layout(
                title='区域电价时段趋势对比', xaxis_title='时段', yaxis_title='平均电价 (元/kWh)',
                template='plotly_white', height=500,
                xaxis=dict(tickangle=45, tickvals=time_columns[::4],
                           ticktext=[time_columns[i] for i in range(0, min(len(time_columns), 96), 4)]),
                hovermode='x unified'
            )
            st.plotly_chart(fig_time, use_container_width=True)

        # 月度均价对比
        if regional_monthly_data:
            st.divider()
            st.subheader("📅 区域月度均价对比")

            month_labels = [f"{m}月" for m in range(1, 13)]
            fig_monthly = go.Figure()
            for region in selected_regions:
                monthly = regional_monthly_data.get(region, {})
                if not monthly:
                    continue
                x_months = [f"{m}月" for m in sorted(monthly.keys())]
                y_prices = [monthly[m] for m in sorted(monthly.keys())]
                fig_monthly.add_trace(go.Scatter(
                    x=x_months, y=y_prices,
                    mode='lines+markers', name=region,
                    line=dict(color=region_colors.get(region, '#999999'), width=2),
                    marker=dict(size=6),
                    text=[f"{p:.4f}" for p in y_prices],
                    hovertemplate='%{x}: %{y:.4f} 元/kWh<extra>%{fullData.name}</extra>'
                ))
            fig_monthly.update_layout(
                title='区域月度均价对比曲线',
                xaxis_title='月份', yaxis_title='月度均价 (元/kWh)',
                template='plotly_white', height=450,
                xaxis=dict(tickmode='array', tickvals=month_labels),
                hovermode='x unified'
            )
            st.plotly_chart(fig_monthly, use_container_width=True)

            # 月度均价明细表
            monthly_table_data = []
            for month in range(1, 13):
                row = {'月份': f"{month}月"}
                for region in selected_regions:
                    val = regional_monthly_data.get(region, {}).get(month)
                    row[region] = round(val, 4) if val is not None else None
                monthly_table_data.append(row)
            monthly_table_df = pd.DataFrame(monthly_table_data)
            st.dataframe(monthly_table_df, use_container_width=True)

        # 区域月度峰谷价差对比
        if regional_monthly_pv_spread_data:
            st.divider()
            st.subheader("📊 区域月度峰谷价差对比")
            st.markdown("月度峰谷价差 = 当月各天峰谷价差（连续8时段均值最高 - 连续8时段均值最低）的均值。")

            month_labels = [f"{m}月" for m in range(1, 13)]
            fig_pv_monthly = go.Figure()
            pv_table_rows = []
            for region in selected_regions:
                monthly = regional_monthly_pv_spread_data.get(region, {})
                if not monthly:
                    continue
                x_months = [f"{m}月" for m in sorted(monthly.keys())]
                y_vals = [monthly[m] for m in sorted(monthly.keys())]
                fig_pv_monthly.add_trace(go.Scatter(
                    x=x_months, y=y_vals,
                    mode='lines+markers', name=region,
                    line=dict(color=region_colors.get(region, '#999999'), width=2),
                    marker=dict(size=6),
                    hovertemplate='%{x}: %{y:.4f} 元/kWh<extra>%{fullData.name}</extra>'
                ))
                row = {'区域': region}
                for m in sorted(monthly.keys()):
                    row[f"{m}月"] = monthly[m]
                pv_table_rows.append(row)

            fig_pv_monthly.update_layout(
                title='各区域月度峰谷价差对比曲线',
                xaxis_title='月份', yaxis_title='月度峰谷价差 (元/kWh)',
                template='plotly_white', height=450,
                xaxis=dict(tickmode='array', tickvals=month_labels),
                hovermode='x unified',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
            )
            st.plotly_chart(fig_pv_monthly, use_container_width=True)

            if pv_table_rows:
                pv_table_df = pd.DataFrame(pv_table_rows)
                st.dataframe(pv_table_df, use_container_width=True, hide_index=True)

        # 区域对价差概率密度分布
        if regional_daily_data and len(selected_regions) >= 2:
            st.divider()
            st.subheader("📊 区域对价差概率密度分布")
            st.markdown("各区域两两配对的日均价差分布，标注均值、P5、P95 分位数。")

            from itertools import combinations as _combinations
            region_pairs = list(_combinations(selected_regions, 2))

            for r1, r2 in region_pairs:
                d1 = regional_daily_data.get(r1, [])
                d2 = regional_daily_data.get(r2, [])
                if not d1 or not d2:
                    continue

                min_len = min(len(d1), len(d2))
                spread = np.array(d1[:min_len]) - np.array(d2[:min_len])
                spread = spread[~np.isnan(spread)]
                if len(spread) == 0:
                    continue

                mean_val = np.mean(spread)
                p5 = np.percentile(spread, 5)
                p95 = np.percentile(spread, 95)

                st.markdown(f"**{r1} − {r2}** （共 {len(spread)} 天）")

                fig_hist = go.Figure()
                fig_hist.add_trace(go.Histogram(
                    x=spread, nbinsx=60, name='价差分布',
                    histnorm='probability density',
                    marker_color='rgba(33, 150, 243, 0.6)',
                    marker_line=dict(width=0.5, color='rgba(33, 150, 243, 1)')
                ))
                fig_hist.add_vline(x=mean_val, line_dash='solid', line_color='#FF6B35', line_width=2,
                                   annotation_text=f'均值 {mean_val:.4f}', annotation_position='top right',
                                   annotation_font_size=11)
                fig_hist.add_vline(x=p5, line_dash='dash', line_color='#4CAF50', line_width=2,
                                   annotation_text=f'P5 {p5:.4f}', annotation_position='top left',
                                   annotation_font_size=11)
                fig_hist.add_vline(x=p95, line_dash='dash', line_color='#9C27B0', line_width=2,
                                   annotation_text=f'P95 {p95:.4f}', annotation_position='top right',
                                   annotation_font_size=11)
                fig_hist.update_layout(
                    title=f'{r1} − {r2} 日均价差概率密度分布',
                    xaxis_title='价差 (元/kWh)', yaxis_title='概率密度',
                    template='plotly_white', height=380,
                    showlegend=False
                )
                st.plotly_chart(fig_hist, use_container_width=True)

                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("均值", f"{mean_val:.4f}")
                with col2:
                    st.metric("P5", f"{p5:.4f}")
                with col3:
                    st.metric("P95", f"{p95:.4f}")
                with col4:
                    st.metric("标准差", f"{np.std(spread):.4f}")

        # 区域对关键价差时序分析
        if regional_pair_spread_ts and len(selected_regions) >= 2:
            st.divider()
            st.subheader("📊 区域对关键价差时序分析")
            st.markdown(
                "对每对区域，计算各时点（96时段）的价差（区域A均价 − 区域B均价），形成价差时间序列，"
                "再计算标准差。标准差越大说明价差在不同时段波动越大；"
                "结合月度标准差趋势可判断价差是否呈现趋势性扩大或缩小。"
            )

            time_columns_local = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]

            # 统计表格
            spread_ts_rows = []
            for (r1, r2), info in sorted(regional_pair_spread_ts.items()):
                spread_ts_rows.append({
                    '区域对': f'{r1} − {r2}',
                    '价差标准差': info['std'],
                    '价差均值': info['mean'],
                    '最大价差': info['max_spread'],
                    '最大价差时段': info['max_slot'],
                    '最小价差': info['min_spread'],
                    '最小价差时段': info['min_slot'],
                })

            if spread_ts_rows:
                st.markdown("**各区域对价差时序统计**")
                spread_ts_df = pd.DataFrame(spread_ts_rows)
                st.dataframe(spread_ts_df, use_container_width=True, hide_index=True)

                # 柱状图：各区域对价差标准差对比
                fig_std_bar = go.Figure()
                fig_std_bar.add_trace(go.Bar(
                    x=[r['区域对'] for r in spread_ts_rows],
                    y=[r['价差标准差'] for r in spread_ts_rows],
                    marker_color='#FF6B35',
                    text=[f"{r['价差标准差']:.4f}" for r in spread_ts_rows],
                    textposition='auto'
                ))
                fig_std_bar.update_layout(
                    title='各区域对价差时序标准差对比',
                    xaxis_title='区域对', yaxis_title='价差标准差 (元/kWh)',
                    template='plotly_white', height=400, showlegend=False
                )
                st.plotly_chart(fig_std_bar, use_container_width=True)

            # 价差时序曲线图
            st.markdown("**各区域对价差时序曲线（96时段）**")
            fig_spread_ts = go.Figure()
            for (r1, r2), info in sorted(regional_pair_spread_ts.items()):
                spread_series = info['spread_series']
                fig_spread_ts.add_trace(go.Scatter(
                    x=time_columns_local[:len(spread_series)],
                    y=spread_series,
                    mode='lines+markers',
                    name=f'{r1} − {r2}',
                    line=dict(width=2),
                    marker=dict(size=3)
                ))
            fig_spread_ts.add_hline(y=0, line_dash='dash', line_color='gray', line_width=1)
            fig_spread_ts.update_layout(
                title='各区域对价差时序曲线',
                xaxis_title='时段', yaxis_title='价差 (元/kWh)',
                template='plotly_white', height=450,
                xaxis=dict(tickangle=45, tickvals=time_columns_local[::4],
                           ticktext=[time_columns_local[i] for i in range(0, min(len(time_columns_local), 96), 4)]),
                hovermode='x unified',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
            )
            st.plotly_chart(fig_spread_ts, use_container_width=True)

        # 月度价差标准差趋势分析
        if regional_pair_monthly_std and len(selected_regions) >= 2:
            st.divider()
            st.subheader("📈 月度价差标准差趋势")
            st.markdown(
                "每月分别计算各区域对的96时段价差序列标准差。"
                "若标准差逐月增大 → 价差波动加剧（趋势性扩大）；"
                "若标准差逐月减小 → 价差趋于收敛（趋势性缩小）。"
            )

            month_labels = [f"{m}月" for m in range(1, 13)]
            fig_monthly_std = go.Figure()
            monthly_std_table_rows = []

            pair_colors = [
                '#FF6B35', '#2196F3', '#4CAF50', '#9C27B0', '#FF9800', '#00BCD4'
            ]
            for idx, ((r1, r2), monthly_stds) in enumerate(sorted(regional_pair_monthly_std.items())):
                color = pair_colors[idx % len(pair_colors)]
                label = f'{r1} − {r2}'
                x_months = [f"{m}月" for m in sorted(monthly_stds.keys())]
                y_vals = [monthly_stds[m] for m in sorted(monthly_stds.keys())]
                fig_monthly_std.add_trace(go.Scatter(
                    x=x_months, y=y_vals,
                    mode='lines+markers', name=label,
                    line=dict(color=color, width=2),
                    marker=dict(size=6),
                    hovertemplate='%{x}: %{y:.4f} 元/kWh<extra>' + label + '</extra>'
                ))
                row = {'区域对': label}
                for m in sorted(monthly_stds.keys()):
                    row[f"{m}月"] = monthly_stds[m]
                monthly_std_table_rows.append(row)

            fig_monthly_std.update_layout(
                title='各区域对月度价差时序标准差趋势',
                xaxis_title='月份', yaxis_title='价差时序标准差 (元/kWh)',
                template='plotly_white', height=450,
                xaxis=dict(tickmode='array', tickvals=month_labels),
                hovermode='x unified',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
            )
            st.plotly_chart(fig_monthly_std, use_container_width=True)

            if monthly_std_table_rows:
                st.markdown("**月度价差标准差明细表**")
                monthly_std_df = pd.DataFrame(monthly_std_table_rows)
                st.dataframe(monthly_std_df, use_container_width=True, hide_index=True)

                # 趋势判断提示
                for (r1, r2), monthly_stds in sorted(regional_pair_monthly_std.items()):
                    months = sorted(monthly_stds.keys())
                    if len(months) >= 3:
                        vals = [monthly_stds[m] for m in months]
                        first_half = np.mean(vals[:len(vals)//2])
                        second_half = np.mean(vals[len(vals)//2:])
                        diff_pct = (second_half - first_half) / first_half * 100 if first_half != 0 else 0
                        label = f'{r1} − {r2}'
                        if diff_pct > 10:
                            st.warning(f"**{label}**: 价差标准差后半段比前半段上升 {diff_pct:.1f}%，呈趋势性扩大。")
                        elif diff_pct < -10:
                            st.success(f"**{label}**: 价差标准差后半段比前半段下降 {abs(diff_pct):.1f}%，呈趋势性缩小。")
                        else:
                            st.info(f"**{label}**: 价差标准差变化 {diff_pct:.1f}%，整体相对稳定。")

        # 区域间时段占比对比
        if regional_data and len(selected_regions) >= 2:
            st.divider()
            st.subheader("📊 区域间高价时段占比")

            from itertools import combinations as _comb2
            pairs = list(_comb2(selected_regions, 2))

            # 构建对比矩阵数据
            ratio_rows = []
            for r1, r2 in pairs:
                d1 = regional_data.get(r1)
                d2 = regional_data.get(r2)
                if d1 is None or d2 is None:
                    continue
                min_len = min(len(d1), len(d2))
                arr1 = np.array(d1[:min_len], dtype=float)
                arr2 = np.array(d2[:min_len], dtype=float)
                valid = ~(np.isnan(arr1) | np.isnan(arr2))
                if valid.sum() == 0:
                    continue
                a1, a2 = arr1[valid], arr2[valid]
                higher = int(np.sum(a1 > a2))
                lower = int(np.sum(a1 < a2))
                equal = int(np.sum(a1 == a2))
                total = len(a1)
                ratio_rows.append({
                    '区域对': f'{r1} > {r2}',
                    '对比方向': f'{r1} 高于 {r2}',
                    f'{r1} 高价时段数': higher,
                    f'{r2} 高价时段数': lower,
                    '持平时段数': equal,
                    f'{r1} 高价占比': round(higher / total * 100, 1),
                    f'{r2} 高价占比': round(lower / total * 100, 1),
                })

            if ratio_rows:
                # 汇总表格
                st.markdown("**各区域对高价时段占比（基于 96 时段日均电价）**")
                ratio_df = pd.DataFrame(ratio_rows)
                st.dataframe(ratio_df, use_container_width=True, hide_index=True)

                # 分组柱状图：每个区域对的双方高价占比
                fig_ratio = go.Figure()
                pair_labels = [r['区域对'] for r in ratio_rows]
                for region in selected_regions:
                    y_vals = []
                    for r in ratio_rows:
                        key = f'{region} 高价占比'
                        y_vals.append(r.get(key, 0))
                    fig_ratio.add_trace(go.Bar(
                        x=pair_labels, y=y_vals, name=region,
                        marker_color=region_colors.get(region, '#999999'),
                        text=[f"{v}%" for v in y_vals],
                        textposition='auto'
                    ))
                fig_ratio.update_layout(
                    title='各区域对中双方高价时段占比',
                    xaxis_title='区域对', yaxis_title='高价时段占比 (%)',
                    template='plotly_white', height=420,
                    barmode='group',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig_ratio, use_container_width=True)

        # 峰平谷分时均价统计
        if regional_data:
            st.divider()
            st.subheader("📊 各区域峰/平/谷分时均价")

            # 构建 96 时段 → 峰/平/谷 映射（按广东省分时规则，尖峰归入峰）
            def _slot_period_type(idx):
                """根据 0~95 时段索引返回 '峰'/'平'/'谷'。"""
                hour = idx * 15 // 60   # 0~23
                if 10 <= hour < 12 or 14 <= hour < 19:
                    return "峰"
                if hour < 8:
                    return "谷"
                return "平"

            slot_types = [_slot_period_type(i) for i in range(96)]
            period_labels = ["峰", "平", "谷"]
            period_colors = {"峰": "#FF6B35", "平": "#2196F3", "谷": "#4CAF50"}

            # 统计各区域在峰/平/谷时段的均价
            period_stats_rows = []
            for region in selected_regions:
                data = regional_data.get(region)
                if data is None or len(data) < 96:
                    continue
                arr = np.array(data[:96], dtype=float)
                row = {"区域": region}
                for period in period_labels:
                    mask = np.array([slot_types[i] == period for i in range(96)])
                    period_prices = arr[mask]
                    period_prices = period_prices[~np.isnan(period_prices)]
                    if len(period_prices) > 0:
                        row[f"{period}时段均价"] = round(float(np.mean(period_prices)), 4)
                        row[f"{period}时段数"] = int(np.sum(mask))
                    else:
                        row[f"{period}时段均价"] = None
                        row[f"{period}时段数"] = 0
                period_stats_rows.append(row)

            if period_stats_rows:
                period_stats_df = pd.DataFrame(period_stats_rows)
                st.dataframe(period_stats_df, use_container_width=True, hide_index=True)

                # 分组柱状图
                fig_period = go.Figure()
                for period in period_labels:
                    col_name = f"{period}时段均价"
                    y_vals = [r.get(col_name) for r in period_stats_rows]
                    x_labels = [r["区域"] for r in period_stats_rows]
                    fig_period.add_trace(go.Bar(
                        x=x_labels, y=y_vals, name=f"{period}时段",
                        marker_color=period_colors.get(period, '#999999'),
                        text=[f"{v:.4f}" if v is not None else "" for v in y_vals],
                        textposition='auto'
                    ))
                fig_period.update_layout(
                    title='各区域峰/平/谷分时均价对比',
                    xaxis_title='区域', yaxis_title='平均电价 (元/kWh)',
                    template='plotly_white', height=420,
                    barmode='group',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
                )
                st.plotly_chart(fig_period, use_container_width=True)

        # 日波动率箱线图
        if regional_volatility_data and len(selected_regions) >= 2:
            st.divider()
            st.subheader("📊 各区域日波动率对比（箱线图）")
            st.markdown("日波动率 = 每日 96 个时段电价的标准差，反映日内价格波动幅度。")

            fig_box = go.Figure()
            vol_stats_rows = []
            for region in selected_regions:
                stds = regional_volatility_data.get(region, [])
                if not stds:
                    continue
                fig_box.add_trace(go.Box(
                    y=stds, name=region,
                    marker_color=region_colors.get(region, '#999999'),
                    boxmean=True,
                    boxpoints='outliers'
                ))
                arr = np.array(stds)
                vol_stats_rows.append({
                    '区域': region,
                    '天数': len(arr),
                    '均值': round(float(np.mean(arr)), 4),
                    '中位数': round(float(np.median(arr)), 4),
                    'P5': round(float(np.percentile(arr, 5)), 4),
                    'P25': round(float(np.percentile(arr, 25)), 4),
                    'P75': round(float(np.percentile(arr, 75)), 4),
                    'P95': round(float(np.percentile(arr, 95)), 4),
                    '标准差': round(float(np.std(arr)), 4),
                })

            fig_box.update_layout(
                title='各区域日波动率（每日电价标准差）箱线图',
                xaxis_title='区域', yaxis_title='日波动率 (元/kWh)',
                template='plotly_white', height=480,
                showlegend=False
            )
            st.plotly_chart(fig_box, use_container_width=True)

            if vol_stats_rows:
                vol_stats_df = pd.DataFrame(vol_stats_rows)
                st.dataframe(vol_stats_df, use_container_width=True, hide_index=True)

        # 月度波动率曲线对比
        if regional_monthly_vol_data:
            st.divider()
            st.subheader("📈 各区域月度电价波动率曲线")
            st.markdown("月度波动率 = 当月各天日波动率（96时段标准差）的均值，反映月内价格波动幅度变化趋势。")

            month_labels = [f"{m}月" for m in range(1, 13)]
            fig_monthly_vol = go.Figure()
            vol_table_rows = []
            for region in selected_regions:
                monthly = regional_monthly_vol_data.get(region, {})
                if not monthly:
                    continue
                x_months = [f"{m}月" for m in sorted(monthly.keys())]
                y_vals = [monthly[m] for m in sorted(monthly.keys())]
                fig_monthly_vol.add_trace(go.Scatter(
                    x=x_months, y=y_vals,
                    mode='lines+markers', name=region,
                    line=dict(color=region_colors.get(region, '#999999'), width=2),
                    marker=dict(size=6),
                    hovertemplate='%{x}: %{y:.4f} 元/kWh<extra>%{fullData.name}</extra>'
                ))
                row = {'区域': region}
                for m in sorted(monthly.keys()):
                    row[f"{m}月"] = monthly[m]
                vol_table_rows.append(row)

            fig_monthly_vol.update_layout(
                title='各区域月度电价波动率对比曲线',
                xaxis_title='月份', yaxis_title='月度波动率 (元/kWh)',
                template='plotly_white', height=450,
                xaxis=dict(tickmode='array', tickvals=month_labels),
                hovermode='x unified',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
            )
            st.plotly_chart(fig_monthly_vol, use_container_width=True)

            if vol_table_rows:
                vol_table_df = pd.DataFrame(vol_table_rows)
                st.dataframe(vol_table_df, use_container_width=True, hide_index=True)

        # 数据导出
        st.divider()
        st.subheader("💾 导出对比数据")

        # 准备导出数据
        export_data = []
        for region, data in regional_data.items():
            for i, time_col in enumerate(time_columns[:96]):
                if i < len(data):
                    export_data.append({
                        '区域': region,
                        '时段': time_col,
                        '平均电价': round(float(data[i]), 4)
                    })

        if export_data:
            export_df = pd.DataFrame(export_data)
            csv_data = export_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 下载区域对比数据 (CSV)",
                data=csv_data,
                file_name=f"区域电价对比_{'_'.join(selected_regions)}.csv",
                mime="text/csv"
            )

    # 多站点电价对比页
    elif view_mode == "📈 多站点电价对比":
        st.header("📈 多站点电价对比分析")
        st.markdown("选择多个站点进行电价曲线对比分析，支持单日对比和日期范围对比两种模式。")

        # 配色方案
        MC_COLORS = [
            '#FF6B35', '#2196F3', '#4CAF50', '#9C27B0',
            '#FF9800', '#00BCD4', '#E91E63', '#795548'
        ]

        # ===== 侧边栏配置 =====
        st.sidebar.divider()
        st.sidebar.header("📊 多站点对比配置")

        # 1. 站点筛选（城市 → 厂站类型）
        mc_filtered_stations = list(price_files.keys())

        # 城市筛选
        if city_types:
            mc_city_options = ["全部城市"] + city_types
            mc_selected_city = st.sidebar.selectbox(
                "选择城市",
                options=mc_city_options,
                index=0,
                key="mc_city"
            )
            if mc_selected_city != "全部城市":
                mc_city_station_df = build_station_group_mapping(
                    station_info_df, '城市', station_names=mc_filtered_stations
                )
                if mc_city_station_df is not None:
                    mc_city_stations = mc_city_station_df[
                        mc_city_station_df['城市'] == mc_selected_city
                    ]['电站名'].astype(str).str.strip().tolist()
                    mc_filtered_stations = [s for s in mc_city_stations if s in price_files.keys()]
                    st.sidebar.info(f"🏙️ {mc_selected_city} 共有 {len(mc_filtered_stations)} 个站点")

        # 厂站类型筛选
        if factory_group_types and len(mc_filtered_stations) > 0:
            mc_factory_options = get_available_group_values(
                station_info_df, FACTORY_GROUP_COLUMN, station_names=mc_filtered_stations
            )
            if mc_factory_options:
                mc_all_factory_option = f"全部{FACTORY_GROUP_COLUMN}"
                mc_factory_list = [mc_all_factory_option] + mc_factory_options
                mc_selected_factory = st.sidebar.selectbox(
                    f"选择{FACTORY_GROUP_COLUMN}",
                    options=mc_factory_list,
                    index=0,
                    key="mc_factory"
                )
                if mc_selected_factory != mc_all_factory_option:
                    mc_factory_station_df = build_station_group_mapping(
                        station_info_df, FACTORY_GROUP_COLUMN, station_names=mc_filtered_stations
                    )
                    if mc_factory_station_df is not None:
                        mc_factory_stations = mc_factory_station_df[
                            mc_factory_station_df[FACTORY_GROUP_COLUMN] == mc_selected_factory
                        ]['电站名'].astype(str).str.strip().tolist()
                        mc_filtered_stations = [s for s in mc_factory_stations if s in mc_filtered_stations]
                        st.sidebar.info(f"🏭 {mc_selected_factory} 共有 {len(mc_filtered_stations)} 个站点")

        # 2. 多站点选择（最多8个）
        st.sidebar.divider()
        st.sidebar.subheader("📍 选择对比站点")

        if len(mc_filtered_stations) == 0:
            st.sidebar.warning("当前筛选条件下没有可用的站点！")
            mc_selected_stations = []
        else:
            # 默认选择前3个站点
            default_stations = mc_filtered_stations[:min(3, len(mc_filtered_stations))]
            mc_selected_stations = st.sidebar.multiselect(
                "选择对比站点（最多8个）",
                options=mc_filtered_stations,
                default=default_stations,
                max_selections=8,
                key="mc_stations",
                help="选择2-8个站点进行电价对比分析"
            )

        # 3. 对比模式选择
        st.sidebar.divider()
        st.sidebar.subheader("📅 对比模式")
        mc_compare_mode = st.sidebar.radio(
            "选择对比模式",
            options=["单日对比", "日期范围对比"],
            horizontal=True,
            key="mc_mode",
            help="单日对比：选择某一天查看各站点96时段电价曲线；日期范围对比：选择日期范围查看各站点平均电价曲线"
        )

        # 4. 日期选择
        data_source = st.session_state.get('data_source', 'excel')

        if mc_selected_stations:
            # 获取第一个站点的日期范围作为参考
            first_station = mc_selected_stations[0]
            if data_source == 'database':
                selected_year = st.session_state.get('selected_year')
                mc_ref_df = load_price_data_from_db(first_station, selected_year)
            else:
                mc_ref_df = load_price_data(*get_file_cache_key(price_files[first_station]))

            if mc_ref_df is not None and len(mc_ref_df.columns) >= 2:
                mc_date_col = mc_ref_df.columns[0]
                if pd.api.types.is_datetime64_any_dtype(mc_ref_df[mc_date_col]):
                    mc_available_dates = mc_ref_df[mc_date_col].dt.date
                else:
                    mc_available_dates = pd.to_datetime(mc_ref_df[mc_date_col]).dt.date
                mc_min_date = mc_available_dates.min()
                mc_max_date = mc_available_dates.max()

                if mc_compare_mode == "单日对比":
                    mc_selected_date = st.sidebar.date_input(
                        "选择日期",
                        value=mc_max_date,
                        min_value=mc_min_date,
                        max_value=mc_max_date,
                        key="mc_date",
                        help="选择要对比的日期"
                    )
                else:
                    mc_col1, mc_col2 = st.sidebar.columns(2)
                    with mc_col1:
                        mc_start_date = st.date_input(
                            "开始日期",
                            value=mc_min_date,
                            min_value=mc_min_date,
                            max_value=mc_max_date,
                            key="mc_start"
                        )
                    with mc_col2:
                        mc_end_date = st.date_input(
                            "结束日期",
                            value=mc_max_date,
                            min_value=mc_min_date,
                            max_value=mc_max_date,
                            key="mc_end"
                        )
            else:
                mc_selected_date = None
                mc_start_date = None
                mc_end_date = None
        else:
            mc_selected_date = None
            mc_start_date = None
            mc_end_date = None

        # ===== 主区域 =====
        st.divider()

        if not mc_selected_stations:
            st.warning("请在侧边栏选择至少一个站点进行对比。")
            st.info("💡 提示：可选择2-8个站点进行电价曲线对比分析。")
        else:
            # 显示选中站点信息
            st.info(f"📊 已选择 **{len(mc_selected_stations)}** 个站点进行对比：{', '.join(mc_selected_stations)}")

            # 加载各站点数据
            mc_station_data = {}
            mc_failed_stations = []

            with st.spinner("正在加载各站点电价数据..."):
                for station in mc_selected_stations:
                    try:
                        if data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            df = load_price_data_from_db(station, selected_year)
                        else:
                            df = load_price_data(*get_file_cache_key(price_files[station]))

                        if df is not None and len(df.columns) >= 2:
                            mc_station_data[station] = df
                        else:
                            mc_failed_stations.append(station)
                    except Exception as e:
                        mc_failed_stations.append(station)

            if mc_failed_stations:
                st.warning(f"以下站点数据加载失败，已跳过：{', '.join(mc_failed_stations)}")

            if not mc_station_data:
                st.error("所有选中站点的数据加载失败，无法进行对比分析。")
            else:
                # 获取时间列（96个时段）
                time_columns = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 15, 30, 45)]

                # 加载全省平均数据
                mc_prov_avg = None
                mc_prov_dates = None

                with st.spinner("正在计算全省平均电价..."):
                    try:
                        if data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            all_db_keys = [(s, selected_year) for s in price_files.keys()]
                            mc_prov_dates, _, mc_prov_avg, _ = compute_provincial_average_from_db(all_db_keys)
                        else:
                            all_file_keys = [get_file_cache_key(fp) for fp in price_files.values()]
                            mc_prov_dates, _, mc_prov_avg = compute_provincial_average_with_cache(tuple(all_file_keys))
                    except Exception:
                        pass

                # 绘制对比图
                if mc_compare_mode == "单日对比" and mc_selected_date:
                    st.subheader(f"📈 {mc_selected_date} 各站点电价对比")

                    fig = go.Figure()

                    # 添加各站点电价曲线
                    for i, (station, df) in enumerate(mc_station_data.items()):
                        mc_date_col = df.columns[0]
                        mc_price_cols = df.columns[1:]

                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            mask = df[mc_date_col].dt.date == mc_selected_date
                        else:
                            mask = pd.to_datetime(df[mc_date_col]).dt.date == mc_selected_date

                        day_data = df[mask]
                        if not day_data.empty:
                            prices = day_data[mc_price_cols].values.flatten().astype(float)
                            if len(prices) >= 96:
                                prices = prices[:96]
                                color = MC_COLORS[i % len(MC_COLORS)]
                                fig.add_trace(go.Scatter(
                                    x=time_columns,
                                    y=prices,
                                    mode='lines+markers',
                                    name=station,
                                    line=dict(color=color, width=2),
                                    marker=dict(size=3)
                                ))

                    # 添加全省平均线
                    if mc_prov_avg is not None and mc_prov_dates is not None:
                        prov_date_to_row = {d: i for i, d in enumerate(mc_prov_dates)}
                        date_str = mc_selected_date.strftime('%Y-%m-%d') if hasattr(mc_selected_date, 'strftime') else str(mc_selected_date)
                        if date_str in prov_date_to_row:
                            prov_prices = mc_prov_avg[prov_date_to_row[date_str]]
                            fig.add_trace(go.Scatter(
                                x=time_columns,
                                y=prov_prices,
                                mode='lines',
                                name='全省平均',
                                line=dict(color='red', width=2, dash='dash')
                            ))

                    fig.update_layout(
                        title=f'{mc_selected_date} 各站点电价对比',
                        xaxis_title='时间节点',
                        yaxis_title='电价 (元/kWh)',
                        hovermode='x unified',
                        template='plotly_white',
                        height=500,
                        xaxis=dict(
                            tickangle=45,
                            tickvals=time_columns[::4],
                            ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                        ),
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.02,
                            xanchor="right",
                            x=1
                        )
                    )
                    st.plotly_chart(fig, use_container_width=True)

                elif mc_compare_mode == "日期范围对比" and mc_start_date and mc_end_date:
                    st.subheader(f"📈 {mc_start_date} 至 {mc_end_date} 各站点平均电价对比")

                    fig = go.Figure()

                    # 计算各站点在日期范围内的平均电价曲线
                    for i, (station, df) in enumerate(mc_station_data.items()):
                        mc_date_col = df.columns[0]
                        mc_price_cols = df.columns[1:]

                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            all_dates = df[mc_date_col].dt.date
                        else:
                            all_dates = pd.to_datetime(df[mc_date_col]).dt.date

                        mask = (all_dates >= mc_start_date) & (all_dates <= mc_end_date)
                        range_data = df[mask]

                        if not range_data.empty:
                            avg_prices = range_data[mc_price_cols].mean(axis=0).values.astype(float)
                            if len(avg_prices) >= 96:
                                avg_prices = avg_prices[:96]
                                color = MC_COLORS[i % len(MC_COLORS)]
                                fig.add_trace(go.Scatter(
                                    x=time_columns,
                                    y=avg_prices,
                                    mode='lines+markers',
                                    name=station,
                                    line=dict(color=color, width=2),
                                    marker=dict(size=3)
                                ))

                    # 添加全省平均线
                    if mc_prov_avg is not None and mc_prov_dates is not None:
                        prov_date_to_row = {d: i for i, d in enumerate(mc_prov_dates)}
                        filtered_prov_indices = [
                            prov_date_to_row[d] for d in mc_prov_dates
                            if mc_start_date <= pd.Timestamp(d).date() <= mc_end_date
                        ]
                        if filtered_prov_indices:
                            prov_avg_prices = np.mean(mc_prov_avg[filtered_prov_indices], axis=0)
                            fig.add_trace(go.Scatter(
                                x=time_columns,
                                y=prov_avg_prices,
                                mode='lines',
                                name='全省平均',
                                line=dict(color='red', width=2, dash='dash')
                            ))

                    days_count = (mc_end_date - mc_start_date).days + 1
                    fig.update_layout(
                        title=f'{mc_start_date} 至 {mc_end_date} 各站点平均电价对比（{days_count}天）',
                        xaxis_title='时间节点',
                        yaxis_title='电价 (元/kWh)',
                        hovermode='x unified',
                        template='plotly_white',
                        height=500,
                        xaxis=dict(
                            tickangle=45,
                            tickvals=time_columns[::4],
                            ticktext=[time_columns[i] for i in range(0, len(time_columns), 4)]
                        ),
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.02,
                            xanchor="right",
                            x=1
                        )
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("请在侧边栏选择日期后开始对比。")

                # 统计信息表格
                st.divider()
                st.subheader("📊 各站点统计信息")

                mc_stats_data = []
                for i, (station, df) in enumerate(mc_station_data.items()):
                    mc_date_col = df.columns[0]
                    mc_price_cols = df.columns[1:]

                    # 根据对比模式筛选数据
                    if mc_compare_mode == "单日对比" and mc_selected_date:
                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            mask = df[mc_date_col].dt.date == mc_selected_date
                        else:
                            mask = pd.to_datetime(df[mc_date_col]).dt.date == mc_selected_date
                        filtered_df = df[mask]
                    elif mc_compare_mode == "日期范围对比" and mc_start_date and mc_end_date:
                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            all_dates = df[mc_date_col].dt.date
                        else:
                            all_dates = pd.to_datetime(df[mc_date_col]).dt.date
                        mask = (all_dates >= mc_start_date) & (all_dates <= mc_end_date)
                        filtered_df = df[mask]
                    else:
                        filtered_df = df

                    if not filtered_df.empty:
                        all_prices = filtered_df[mc_price_cols].values.flatten().astype(float)

                        # 计算峰谷价差
                        n_peak = 8
                        daily_spreads = []
                        for _, row in filtered_df.iterrows():
                            row_prices = pd.to_numeric(row[mc_price_cols], errors='coerce').dropna().values
                            if len(row_prices) >= n_peak * 2:
                                _, _, spread = _sliding_window_spread(row_prices, n_peak)
                                daily_spreads.append(spread)

                        mc_stats_data.append({
                            '站点名称': station,
                            '平均电价 (元/kWh)': round(all_prices.mean(), 4),
                            '最高电价 (元/kWh)': round(all_prices.max(), 4),
                            '最低电价 (元/kWh)': round(all_prices.min(), 4),
                            '日均峰谷价差 (元/kWh)': round(np.mean(daily_spreads), 4) if daily_spreads else 'N/A',
                            '数据天数': len(filtered_df),
                            '颜色': MC_COLORS[i % len(MC_COLORS)]
                        })

                if mc_stats_data:
                    mc_stats_df = pd.DataFrame(mc_stats_data)

                    # 显示带颜色标记的统计表格
                    mc_display_df = mc_stats_df.drop(columns=['颜色'])
                    st.dataframe(mc_display_df, use_container_width=True, hide_index=True)

                    # 显示各站点均价柱状图
                    st.subheader("📊 各站点平均电价对比")
                    fig_bar = go.Figure()
                    fig_bar.add_trace(go.Bar(
                        x=mc_stats_df['站点名称'],
                        y=mc_stats_df['平均电价 (元/kWh)'],
                        marker_color=mc_stats_df['颜色'],
                        text=[f"{x:.4f}" for x in mc_stats_df['平均电价 (元/kWh)']],
                        textposition='auto'
                    ))

                    # 添加全省平均参考线
                    if mc_prov_avg is not None:
                        if mc_compare_mode == "单日对比" and mc_selected_date and mc_prov_dates:
                            prov_date_to_row = {d: i for i, d in enumerate(mc_prov_dates)}
                            date_str = mc_selected_date.strftime('%Y-%m-%d') if hasattr(mc_selected_date, 'strftime') else str(mc_selected_date)
                            if date_str in prov_date_to_row:
                                prov_overall_avg = np.mean(mc_prov_avg[prov_date_to_row[date_str]])
                                fig_bar.add_hline(
                                    y=prov_overall_avg, line_dash="dash", line_color="red",
                                    annotation_text=f"全省均价: {prov_overall_avg:.4f}"
                                )
                        elif mc_compare_mode == "日期范围对比" and mc_start_date and mc_end_date and mc_prov_dates:
                            prov_date_to_row = {d: i for i, d in enumerate(mc_prov_dates)}
                            filtered_prov_indices = [
                                prov_date_to_row[d] for d in mc_prov_dates
                                if mc_start_date <= pd.Timestamp(d).date() <= mc_end_date
                            ]
                            if filtered_prov_indices:
                                prov_overall_avg = np.mean(mc_prov_avg[filtered_prov_indices])
                                fig_bar.add_hline(
                                    y=prov_overall_avg, line_dash="dash", line_color="red",
                                    annotation_text=f"全省均价: {prov_overall_avg:.4f}"
                                )

                    fig_bar.update_layout(
                        title='各站点平均电价对比',
                        xaxis_title='站点名称',
                        yaxis_title='平均电价 (元/kWh)',
                        template='plotly_white',
                        height=400,
                        showlegend=False
                    )
                    st.plotly_chart(fig_bar, use_container_width=True)

                # 数据导出
                st.divider()
                st.subheader("💾 导出对比数据")

                # 准备导出数据
                export_data = []
                for station, df in mc_station_data.items():
                    mc_date_col = df.columns[0]
                    mc_price_cols = df.columns[1:]

                    # 根据对比模式筛选数据
                    if mc_compare_mode == "单日对比" and mc_selected_date:
                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            mask = df[mc_date_col].dt.date == mc_selected_date
                        else:
                            mask = pd.to_datetime(df[mc_date_col]).dt.date == mc_selected_date
                        filtered_df = df[mask]
                    elif mc_compare_mode == "日期范围对比" and mc_start_date and mc_end_date:
                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            all_dates = df[mc_date_col].dt.date
                        else:
                            all_dates = pd.to_datetime(df[mc_date_col]).dt.date
                        mask = (all_dates >= mc_start_date) & (all_dates <= mc_end_date)
                        filtered_df = df[mask]
                    else:
                        filtered_df = df

                    for _, row in filtered_df.iterrows():
                        date_val = row[mc_date_col]
                        if pd.api.types.is_datetime64_any_dtype(df[mc_date_col]):
                            date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
                        else:
                            date_str = str(date_val)[:10]

                        for j, price_col in enumerate(mc_price_cols[:96]):
                            export_data.append({
                                '站点': station,
                                '日期': date_str,
                                '时间节点': time_columns[j] if j < len(time_columns) else price_col,
                                '电价 (元/kWh)': float(row[price_col]) if pd.notna(row[price_col]) else None
                            })

                if export_data:
                    mc_export_df = pd.DataFrame(export_data)
                    csv_data = mc_export_df.to_csv(index=False, encoding='utf-8-sig')

                    mode_suffix = "单日对比" if mc_compare_mode == "单日对比" else f"{mc_start_date}_{mc_end_date}"
                    file_name = f"多站点电价对比_{mode_suffix}.csv"

                    st.download_button(
                        label="📥 下载对比数据 (CSV)",
                        data=csv_data,
                        file_name=file_name,
                        mime="text/csv"
                    )

    # 光伏上网电费计算页
    elif view_mode == "💰 光伏上网电费计算":
        st.header("💰 光伏上网电费计算")
        st.markdown("""
        上传上网电量表格文件，选择站点电价，计算上网电费。
        """)

        # 计算方式选择
        calc_method = st.radio(
            "选择计算方式",
            options=["下一时间节点电价", "当前时间节点电价"],
            horizontal=True,
            help="下一时间节点电价：E(n)×电价(C(n+1))；当前时间节点电价：E(n)×电价(C(n))"
        )

        if calc_method == "下一时间节点电价":
            st.markdown("""
            **计算公式**：`上网电费 = Σ(E(n) × 电价(C(n+1))) × 综合倍率`
            
            - E(n)：第 n 行的上网电量（表格第E列）
            - C(n+1)：第 n+1 行的时间节点对应的站点电价
            """)
        else:
            st.markdown("""
            **计算公式**：`上网电费 = Σ(E(n) × 电价(C(n))) × 综合倍率`
            
            - E(n)：第 n 行的上网电量（表格第E列）
            - C(n)：第 n 行的时间节点对应的站点电价
            """)

        # 1. 上传文件
        uploaded_file = st.file_uploader(
            "上传上网电量表格文件",
            type=['xlsx', 'xls'],
            help="上传类似'荣拓四月上网电量.xlsx'格式的文件"
        )

        if uploaded_file is not None:
            try:
                uploaded_df = pd.read_excel(uploaded_file, header=0)

                # 验证必要列和第E列
                if '数据时间' not in uploaded_df.columns:
                    st.error("上传文件缺少必要列：数据时间")
                elif len(uploaded_df.columns) < 5:
                    st.error("上传文件列数不足，需要至少5列（第E列为上网电量）")
                else:
                    st.success(f"文件上传成功，共 {len(uploaded_df)} 条记录")

                    # 显示文件预览
                    with st.expander("查看文件预览"):
                        st.dataframe(uploaded_df.head(10), use_container_width=True)

                    st.divider()

                    # 2. 用户配置
                    col1, col2 = st.columns(2)
                    with col1:
                        # 站点选择
                        available_stations = list(price_files.keys())
                        if not available_stations:
                            st.warning("当前没有可用的电价数据站点")
                            return
                        selected_station = st.selectbox(
                            "选择站点（用于查询电价）",
                            options=available_stations,
                            help="选择要查询电价的站点"
                        )
                    with col2:
                        # 综合倍率输入
                        multiplier = st.number_input(
                            "综合倍率",
                            min_value=0.0,
                            value=1.0,
                            step=0.01,
                            help="输入综合倍率（如 CT 变比）"
                        )

                    st.divider()

                    # 3. 加载站点电价数据
                    data_source = st.session_state.get('data_source', 'excel')
                    with st.spinner("正在加载站点电价数据..."):
                        if data_source == 'database':
                            selected_year = st.session_state.get('selected_year')
                            price_df = load_price_data_from_db(selected_station, selected_year)
                        else:
                            station_file = price_files.get(selected_station)
                            if station_file is not None:
                                price_df = load_price_data(*get_file_cache_key(station_file))
                            else:
                                price_df = None

                    if price_df is None:
                        st.error(f"无法加载站点 {selected_station} 的电价数据")
                    else:
                        # 4. 构建电价查找表：日期+时间 → 电价
                        price_date_col = price_df.columns[0]
                        price_time_cols = price_df.columns[1:]
                        time_labels = price_time_cols.tolist()  # ['00:00', '00:15', ...]

                        # 将电价数据展开为长格式：(date, time_str) → price
                        price_lookup = {}
                        for _, row in price_df.iterrows():
                            date_val = row[price_date_col]
                            try:
                                date_str = pd.Timestamp(date_val).strftime('%Y-%m-%d')
                            except Exception:
                                date_str = str(date_val)[:10]
                            for i, time_col in enumerate(price_time_cols):
                                price_val = pd.to_numeric(row[time_col], errors='coerce')
                                if not np.isnan(price_val):
                                    price_lookup[(date_str, time_labels[i])] = float(price_val)

                        # 5. 计算上网电费（使用第E列，即索引4）
                        uploaded_df['数据时间'] = pd.to_datetime(uploaded_df['数据时间'])
                        uploaded_df['日期'] = uploaded_df['数据时间'].dt.strftime('%Y-%m-%d')
                        uploaded_df['时间'] = uploaded_df['数据时间'].dt.strftime('%H:%M')
                        uploaded_df['上网电量_kWh'] = pd.to_numeric(uploaded_df.iloc[:, 4], errors='coerce')

                        # 对齐到15分钟节点
                        def snap_to_15min(time_str):
                            try:
                                h, m = map(int, time_str.split(':'))
                                snapped_m = (m // 15) * 15
                                return f"{h:02d}:{snapped_m:02d}"
                            except Exception:
                                return time_str

                        uploaded_df['时间_对齐'] = uploaded_df['时间'].apply(snap_to_15min)

                        # 计算函数：根据计算方式返回结果
                        def calc_fee(method):
                            results = []
                            total_fee = 0.0
                            matched_count = 0
                            unmatched_count = 0

                            for idx in range(len(uploaded_df)):
                                e_n = uploaded_df.iloc[idx]['上网电量_kWh']
                                if pd.isna(e_n) or e_n == 0:
                                    continue

                                if method == "下一时间节点电价":
                                    # C(n+1)：取下一行的时间
                                    if idx + 1 < len(uploaded_df):
                                        next_row = uploaded_df.iloc[idx + 1]
                                        ref_date = next_row['日期']
                                        ref_time = next_row['时间_对齐']
                                        ref_label = '下一时间'
                                        ref_value = next_row['数据时间']
                                    else:
                                        continue
                                else:
                                    # C(n)：取当前行的时间
                                    ref_date = uploaded_df.iloc[idx]['日期']
                                    ref_time = uploaded_df.iloc[idx]['时间_对齐']
                                    ref_label = '当前时间'
                                    ref_value = uploaded_df.iloc[idx]['数据时间']

                                # 查找电价
                                price = price_lookup.get((ref_date, ref_time))
                                if price is not None:
                                    fee = e_n * price * multiplier
                                    total_fee += fee
                                    matched_count += 1
                                    results.append({
                                        '行号': idx + 1,
                                        '时间': uploaded_df.iloc[idx]['数据时间'],
                                        '上网电量_kWh': round(e_n, 5),
                                        ref_label: ref_value,
                                        '电价_元/kWh': round(price, 4),
                                        '电费_元': round(fee, 5)
                                    })
                                else:
                                    unmatched_count += 1

                            return results, total_fee, matched_count, unmatched_count

                        # 执行计算
                        results, total_fee, matched_count, unmatched_count = calc_fee(calc_method)

                        # 6. 显示结果
                        st.subheader("📊 计算结果")

                        col_a, col_b, col_c, col_d = st.columns(4)
                        with col_a:
                            total_electricity = sum(r['上网电量_kWh'] for r in results) * multiplier
                            st.metric("总上网电量", f"{total_electricity:,.2f} kWh")
                        with col_b:
                            st.metric("总上网电费", f"{total_fee:,.2f} 元")
                        with col_c:
                            st.metric("匹配成功", f"{matched_count} 条")
                        with col_d:
                            st.metric("未匹配", f"{unmatched_count} 条")

                        if results:
                            st.divider()
                            st.subheader("📋 明细数据")
                            result_df = pd.DataFrame(results)
                            st.dataframe(result_df, use_container_width=True, hide_index=True)

                            # 导出
                            csv_data = result_df.to_csv(index=False, encoding='utf-8-sig')
                            st.download_button(
                                label="📥 下载计算明细 (CSV)",
                                data=csv_data,
                                file_name=f"上网电费明细_{selected_station}.csv",
                                mime="text/csv"
                            )

            except Exception as e:
                st.error(f"文件处理失败：{str(e)}")

if __name__ == "__main__":
    main()
